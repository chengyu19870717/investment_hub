from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field
from typing import Optional
from datetime import datetime
from pathlib import Path
from uuid import uuid4
import sqlite3, json, re, os, subprocess, struct, asyncio, sys, threading, hashlib
import urllib.parse
from pypinyin import lazy_pinyin, Style
from project_inventory import DEFAULT_PROJECT_ROOT, list_project_inventory, resolve_project_root
from requirement_docs import (
    DEFAULT_REQUIREMENT_DOC_DIR,
    MODEL_OPTIONS,
    SECTION_OPTIONS,
    build_cli_optimization_instruction,
    create_requirement_doc,
    list_requirement_documents,
    next_requirement_version,
    parse_requirement_docx,
    parse_requirement_description,
    resolve_save_dir,
)
from wechat_assistant.comfyui_client import ComfyUIError, choose_checkpoint, find_speed_lora, generate_comfy_image, list_checkpoints
from wechat_assistant.config import load_config
from wechat_assistant.hot_topics import fetch_hot_topics
from wechat_assistant.llm import LLMError
from wechat_assistant.service import WechatContentAssistant
from file_organizer import (
    DEFAULT_RULES as ORGANIZER_DEFAULT_RULES,
    SUGGESTED_DIRS as ORGANIZER_SUGGESTED_DIRS,
    OrganizerError,
    dir_summary as organizer_dir_summary,
    organize as organizer_organize,
    plan as organizer_plan,
    safe_dir as organizer_safe_dir,
    undo as organizer_undo,
)
from investment_snapshot import (
    load_industry_data as _snap_load_industry_data,
    build_exposures as _snap_build_exposures,
    stock_dimensions as _snap_stock_dimensions,
    code_keys as _snap_code_keys,
    load_thresholds as _snap_load_thresholds,
)

app = FastAPI()

BASE_DIR     = Path(__file__).parent
DB_PATH      = Path.home() / ".baibao" / "baibao.db"
REPORT_DIR   = Path.home() / "project" / "quant_trading" / "reports"
UPLOAD_DIR   = BASE_DIR / "uploads"
OUTPUT_DIR   = BASE_DIR / "output"
INDUSTRY_CHAIN_PATH = BASE_DIR / "static" / "data" / "industry_chain.json"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

app.mount("/static",  StaticFiles(directory=BASE_DIR / "static"),  name="static")
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR),           name="uploads")
app.mount("/output",  StaticFiles(directory=OUTPUT_DIR),           name="output")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

@app.get("/upstream", response_class=HTMLResponse)
def upstream_page(request: Request):
    return templates.TemplateResponse("upstream.html", {
        "request": request,
        "embedded": request.query_params.get("embedded") == "1",
    })

@app.get("/data-standard", response_class=HTMLResponse)
def data_standard_page(request: Request):
    return templates.TemplateResponse("data_standard.html", {"request": request})

@app.get("/data-graph", response_class=HTMLResponse)
def data_graph_page(request: Request):
    return templates.TemplateResponse("data_graph.html", {"request": request})

@app.get("/api/industry-chain")
def industry_chain():
    if not INDUSTRY_CHAIN_PATH.exists():
        return JSONResponse({"error": "产业链数据不存在"}, status_code=404)
    with open(INDUSTRY_CHAIN_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

@app.get("/api/pinyin")
def to_pinyin(text: str = ""):
    """将中文转为拼音首字母大写驼峰（用于自动生成字段英文名）"""
    if not text.strip():
        return {"result": ""}
    parts = lazy_pinyin(text.strip(), style=Style.NORMAL)
    result = "_".join(p for p in parts if p.isalpha())
    return {"result": result}

# 首页二级菜单的分类。顺序即页面上标签页的顺序，id 与 FEATURES 的 category 对应
FEATURE_CATEGORIES = [
    {"id": "invest",  "label": "投资理财", "icon": "📈"},
    {"id": "office",  "label": "办公效率", "icon": "🗂️"},
    {"id": "data",    "label": "数据与项目", "icon": "🧭"},
    {"id": "system",  "label": "系统工具", "icon": "⚙️"},
    {"id": "game",    "label": "游戏工具", "icon": "🎮"},
]

FEATURES = [
    {"title": "投资分析",       "url": "/investment-analysis", "icon": "📈", "description": "集中管理投资研究与产业链分析", "status": "active", "category": "invest"},
    {"title": "录音转会议纪要", "url": "/audio",         "icon": "🎙️", "description": "上传录音自动生成纪要", "status": "active", "category": "office"},
    {"title": "一图一表",       "url": "/chart",         "icon": "🗂️", "description": "可编辑业务流程图",     "status": "active", "category": "office"},
    {"title": "待办管理",       "url": "/tasks",         "icon": "📝", "description": "快速登记和管理待办",   "status": "active", "category": "office"},
    {"title": "微信公众号运营", "url": "/wechat-ops",    "icon": "📰", "description": "选题、正文、标题与排版生成", "status": "active", "category": "office"},
    {"title": "需求文档工作流", "url": "/requirement-docs", "icon": "📄", "description": "描述解析、Word文档生成与版本管理", "status": "active", "category": "office"},
    {"title": "项目清单",       "url": "/project-list",   "icon": "🧭", "description": "汇总本地项目与项目简介",     "status": "active", "category": "data"},
    {"title": "数据标准",       "url": "/data-standard",  "icon": "📐", "description": "数据标准化配置与管理",   "status": "active", "category": "data"},
    {"title": "文件整理",       "url": "/file-organizer", "icon": "🗃️", "description": "按规则把下载/文稿的散落文件归类", "status": "active", "category": "system"},
    {"title": "代理网关",       "url": "/proxy",          "icon": "🌐", "description": "一键开关系统代理服务",   "status": "active", "category": "system"},
    {"title": "ComfyUI",         "url": "/comfyui",       "icon": "🎨", "description": "AI绘图：启停管理 + 一键跳转", "status": "active", "category": "system"},
    {"title": "富甲天下5修改器", "url": "/game-save",     "icon": "🎮", "description": "存档修改：金钱/士兵/资源",  "status": "active", "category": "game"},
    {"title": "三国志14修改器",  "url": "/san14-save",    "icon": "🏯", "description": "离线存档修改：城市/武将/势力", "status": "active", "category": "game"},
    {"title": "三国立志传3修改器","url": "/sango3-save",   "icon": "⚔️", "description": "存档修改：点数/天数/地图",  "status": "active", "category": "game"},
]


def grouped_features():
    """按 FEATURE_CATEGORIES 的顺序分组，空分类不出现在页面上。
    category 缺失或写错的功能兜底进「系统工具」，避免新增功能忘了标分类就从首页消失。"""
    valid = {c["id"] for c in FEATURE_CATEGORIES}
    buckets = {c["id"]: [] for c in FEATURE_CATEGORIES}
    for feature in FEATURES:
        cid = feature.get("category")
        buckets[cid if cid in valid else "system"].append(feature)
    return [
        {**cat, "features": buckets[cat["id"]]}
        for cat in FEATURE_CATEGORIES if buckets[cat["id"]]
    ]

# ── 代理配置 ──────────────────────────────────────────────
PROXY_IFACE  = "Wi-Fi"          # 网络接口名称
PROXY_HOST   = "127.0.0.1"
PROXY_PORT   = 9981


# ── 数据库 ────────────────────────────────────────────────

def get_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS tasks (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            title        TEXT    NOT NULL,
            note         TEXT,
            is_recurring INTEGER,
            task_date    TEXT,
            status       TEXT,
            done_at      TEXT
        );
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
        CREATE TABLE IF NOT EXISTS products (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id      TEXT UNIQUE NOT NULL,
            product_name    TEXT NOT NULL,
            product_desc    TEXT,
            product_manager TEXT,
            biz_contact     TEXT,
            biz_dept        TEXT,
            chart_data      TEXT,  -- JSON: {steps:[{id,name,desc}]}
            created_at      TEXT,
            updated_at      TEXT
        );
        CREATE TABLE IF NOT EXISTS charts (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            title      TEXT NOT NULL,
            data_json  TEXT NOT NULL,
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS product_changelogs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id  INTEGER NOT NULL,
            change_desc TEXT,
            changed_at  TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_events (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            stock_code   TEXT NOT NULL,
            stock_name   TEXT,
            event_date   TEXT NOT NULL,
            event_title  TEXT NOT NULL,
            event_type   TEXT DEFAULT '其他',
            event_desc   TEXT,
            status       TEXT DEFAULT 'pending',
            importance   TEXT DEFAULT 'normal',
            created_at   TEXT,
            updated_at   TEXT
        );
        CREATE TABLE IF NOT EXISTS stock_decision_notes (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            code          TEXT NOT NULL,
            name          TEXT,
            note          TEXT NOT NULL,
            target_date   TEXT,
            created_at    TEXT NOT NULL,
            resolved      INTEGER NOT NULL DEFAULT 0,
            verdict       TEXT,
            resolved_note TEXT,
            resolved_at   TEXT
        );
        CREATE TABLE IF NOT EXISTS game_save_hero_templates (
            id          TEXT PRIMARY KEY,
            name        TEXT NOT NULL,
            data_json   TEXT NOT NULL,
            hero_count  INTEGER DEFAULT 0,
            field_count INTEGER DEFAULT 0,
            created_at  TEXT,
            updated_at  TEXT
        );
    """)

    # 数据标准模块 — 兼容旧表迁移
    cols_rules = [r[1] for r in conn.execute("PRAGMA table_info(rules)").fetchall()]
    if not cols_rules:
        conn.execute("""
            CREATE TABLE rules (
                id              TEXT PRIMARY KEY,
                name            TEXT NOT NULL,
                description     TEXT,
                input_json      TEXT,
                output_json     TEXT,
                created_at      TEXT,
                updated_at      TEXT
            )
        """)
    elif "input_json" not in cols_rules:
        conn.execute("ALTER TABLE rules ADD COLUMN input_json TEXT")
        conn.execute("ALTER TABLE rules ADD COLUMN output_json TEXT")
        conn.execute("ALTER TABLE rules ADD COLUMN created_at TEXT")
        conn.execute("ALTER TABLE rules ADD COLUMN updated_at TEXT")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS data_roots (
            id              TEXT PRIMARY KEY,
            name            TEXT NOT NULL,
            meaning         TEXT,
            root_type       TEXT,
            length          INTEGER,
            code_values     TEXT,
            remark          TEXT,
            created_at      TEXT,
            updated_at      TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS data_fields (
            id              TEXT PRIMARY KEY,
            name_en         TEXT NOT NULL,
            name_cn         TEXT,
            meaning         TEXT,
            root_id         TEXT,
            root_name       TEXT,
            field_type      TEXT,
            length          INTEGER,
            code_values     TEXT,
            remark          TEXT,
            created_at      TEXT,
            updated_at      TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS interfaces (
            id              TEXT PRIMARY KEY,
            name            TEXT NOT NULL,
            description     TEXT,
            input_json      TEXT,
            output_json     TEXT,
            created_at      TEXT,
            updated_at      TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS field_rules (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            field_id        TEXT,
            rule_id         TEXT,
            created_at      TEXT
        )
    """)

    # 量化参数 — 因子权重配置
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factor_weights (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            factor_key      TEXT UNIQUE NOT NULL,
            factor_name     TEXT NOT NULL,
            weight          REAL NOT NULL DEFAULT 0,
            description     TEXT,
            is_active       INTEGER DEFAULT 1,
            created_at      TEXT,
            updated_at      TEXT
        )
    """)

    # 初始化默认因子权重（来自股神计划 ai_scorer.py）
    existing_factors = conn.execute("SELECT COUNT(*) FROM factor_weights").fetchone()[0]
    if existing_factors == 0:
        now = datetime.now().isoformat()
        defaults = [
            ("technical",   "技术面", 0.30, "技术指标评分（MA/MACD/KDJ/布林带等）"),
            ("fundamental", "基本面", 0.20, "基本面指标评分（毛利率/ROE/营收增长/PE等）"),
            ("money_flow",  "资金面", 0.20, "资金流向评分（主力净流入/流通市值）"),
            ("sentiment",   "情绪面", 0.15, "市场情绪评分（换手率/量比/涨跌幅）"),
            ("chip",        "筹码面", 0.15, "筹码分布评分（筹码密集/获利比/筹码宽度）"),
        ]
        for key, name, w, desc in defaults:
            conn.execute(
                "INSERT INTO factor_weights(factor_key,factor_name,weight,description,created_at,updated_at) VALUES(?,?,?,?,?,?)",
                (key, name, w, desc, now, now),
            )

    # 因子细项参数表
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factor_sub_params (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            factor_key  TEXT NOT NULL,
            param_key   TEXT NOT NULL,
            param_name  TEXT NOT NULL,
            param_value REAL NOT NULL,
            description TEXT,
            UNIQUE(factor_key, param_key)
        )
    """)

    # 初始化默认细项参数（对应 ai_scorer.py 硬编码值）
    existing_sub = conn.execute("SELECT COUNT(*) FROM factor_sub_params").fetchone()[0]
    if existing_sub == 0:
        sub_defaults = [
            # 技术面
            ("technical", "signal_score",       "信号单项得分",       12, "每个看多/看空信号的得分绝对值"),
            ("technical", "multi_signal_bonus",  "多信号共振加分",     10, "≥2个同向信号时额外加/减分"),
            ("technical", "ma20_far_penalty",    "MA20远距惩罚(>10%)", 15, "股价偏离MA20超过10%时扣分"),
            ("technical", "ma20_mid_penalty",    "MA20中距惩罚(>5%)",  8,  "股价偏离MA20超过5%时扣分"),
            # 基本面
            ("fundamental", "gm_high_score",      "高毛利率加分(>30%)", 15, "毛利率高于30%加分"),
            ("fundamental", "gm_mid_score",       "中毛利率加分(>20%)", 10, "毛利率高于20%加分"),
            ("fundamental", "gm_low_penalty",     "低毛利率惩罚(<5%)",  10, "毛利率低于5%扣分"),
            ("fundamental", "roe_high_score",     "高ROE加分(>15%)",    15, "ROE高于15%加分"),
            ("fundamental", "roe_mid_score",      "中ROE加分(>10%)",    10, "ROE高于10%加分"),
            ("fundamental", "roe_neg_penalty",    "负ROE惩罚",          15, "ROE为负时扣分"),
            ("fundamental", "rev_growth_score",   "营收增长加分(>20%)", 10, "营收增长超过20%加分"),
            ("fundamental", "profit_growth_score","利润增长加分(>20%)", 10, "净利润增长超过20%加分"),
            ("fundamental", "pe_good_score",      "合理PE加分(0~20)",   10, "PE处于0~20合理区间加分"),
            ("fundamental", "pe_missing_penalty", "PE缺失惩罚",         15, "PE数据缺失或为0时扣分"),
            ("fundamental", "pe_bad_penalty",     "PE异常惩罚(>100)",   10, "PE超过100或为负时扣分"),
            ("fundamental", "pb_good_score",      "合理PB加分(0~2)",    5,  "PB处于0~2合理区间加分"),
            ("fundamental", "pb_missing_penalty", "PB缺失惩罚",         10, "PB数据缺失或为0时扣分"),
            # 资金面
            ("money_flow", "flow_very_high_score","强流入加分(>5%)",    30, "主力净流入占流通市值>5%"),
            ("money_flow", "flow_high_score",     "中流入加分(>3%)",    20, "主力净流入占流通市值>3%"),
            ("money_flow", "flow_mid_score",      "低流入加分(>1%)",    10, "主力净流入占流通市值>1%"),
            ("money_flow", "flow_high_penalty",   "强流出惩罚(<-3%)",   25, "主力净流出占流通市值>3%"),
            ("money_flow", "flow_mid_penalty",    "低流出惩罚(<-1%)",   15, "主力净流出占流通市值>1%"),
            # 情绪面
            ("sentiment", "turnover_high_score",  "高换手加分(>10%)",   15, "换手率高于10%加分"),
            ("sentiment", "turnover_mid_score",   "中换手加分(>5%)",    8,  "换手率高于5%加分"),
            ("sentiment", "turnover_low_penalty", "低换手惩罚(<1%)",    10, "换手率低于1%扣分"),
            ("sentiment", "vol_high_score",       "高量比加分(>2)",     15, "量比高于2加分"),
            ("sentiment", "vol_mid_score",        "中量比加分(>1.5)",   8,  "量比高于1.5加分"),
            ("sentiment", "vol_low_penalty",      "低量比惩罚(<0.5)",   10, "量比低于0.5扣分"),
            ("sentiment", "change_high_score",    "大涨跌幅加分(>5%)",  10, "涨跌幅绝对值超过5%加分"),
            ("sentiment", "change_mid_score",     "中涨跌幅加分(>3%)",  5,  "涨跌幅绝对值超过3%加分"),
            # 筹码面
            ("chip", "converging_score",          "筹码收敛信号加分",   20, "近15天筹码持续收敛"),
            ("chip", "tight_low_profit_score",    "极紧集中低获利加分", 20, "70%筹码集中+低获利比例"),
            ("chip", "wide_low_profit_score",     "大范围低获利加分",   15, "大范围套牢盘有解套动力"),
            ("chip", "low_profit_bonus",          "超低获利奖励(<10%)", 5,  "获利比例低于10%额外加分"),
            ("chip", "narrow_width_bonus",        "极窄宽度奖励(<5%)",  5,  "筹码宽度低于5%额外加分"),
        ]
        for fk, pk, pn, pv, pd in sub_defaults:
            conn.execute(
                "INSERT INTO factor_sub_params(factor_key,param_key,param_name,param_value,description) VALUES(?,?,?,?,?)",
                (fk, pk, pn, pv, pd),
            )

    # 每只股票的因子权重覆盖表（不覆盖则继承全局）
    conn.execute("""
        CREATE TABLE IF NOT EXISTS stock_factor_overrides (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            code        TEXT NOT NULL,
            factor_key  TEXT NOT NULL,
            weight      REAL NOT NULL,
            updated_at  TEXT,
            UNIQUE(code, factor_key)
        )
    """)

    # 每只股票的专属评分模型（存储特征名+权重+阈值）
    conn.execute("""
        CREATE TABLE IF NOT EXISTS stock_custom_models (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            code        TEXT UNIQUE NOT NULL,
            name        TEXT NOT NULL,
            features    TEXT NOT NULL,   -- JSON: ["feat1","feat2",...]
            weights     TEXT NOT NULL,   -- JSON: [w1, w2, ...]
            threshold   REAL NOT NULL DEFAULT 0.25,
            ic          REAL,
            icir        REAL,
            accuracy    REAL,
            up_win_rate REAL,
            sample_days INTEGER,
            description TEXT,
            created_at  TEXT,
            updated_at  TEXT
        )
    """)

    # ── 文件整理 ──
    # 用户纳入整理范围的目录（路径唯一）
    conn.execute("""
        CREATE TABLE IF NOT EXISTS organizer_dirs (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            path       TEXT UNIQUE NOT NULL,
            created_at TEXT
        )
    """)
    # 分类规则：priority 小的先匹配，命中即停（先具体后宽泛）
    conn.execute("""
        CREATE TABLE IF NOT EXISTS organizer_rules (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            category   TEXT NOT NULL,
            match_type TEXT NOT NULL DEFAULT 'keyword',  -- keyword | ext
            pattern    TEXT NOT NULL,                    -- 多个用 | 分隔
            priority   INTEGER NOT NULL DEFAULT 500,
            enabled    INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT
        )
    """)
    # 每次整理的移动流水，供「撤销上次整理」使用
    conn.execute("""
        CREATE TABLE IF NOT EXISTS organizer_runs (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            directory  TEXT NOT NULL,
            moves      TEXT NOT NULL,   -- JSON: [{filename,category,src,dest,reason}]
            moved_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT,
            undone_at  TEXT
        )
    """)
    # 首次运行播种预置规则和候选目录
    if conn.execute("SELECT COUNT(*) FROM organizer_rules").fetchone()[0] == 0:
        now = datetime.now().isoformat(timespec="seconds")
        for rule in ORGANIZER_DEFAULT_RULES:
            conn.execute(
                "INSERT INTO organizer_rules(category,match_type,pattern,priority,enabled,updated_at)"
                " VALUES(?,?,?,?,1,?)",
                (rule["category"], rule["match_type"], rule["pattern"], rule["priority"], now),
            )
    if conn.execute("SELECT COUNT(*) FROM organizer_dirs").fetchone()[0] == 0:
        now = datetime.now().isoformat(timespec="seconds")
        for d in ORGANIZER_SUGGESTED_DIRS:
            if d.is_dir():
                conn.execute("INSERT OR IGNORE INTO organizer_dirs(path,created_at) VALUES(?,?)",
                             (str(d), now))

    conn.commit()
    conn.close()

init_db()

# ── 后台任务状态追踪 ──────────────────────────────────────
refresh_task_status = {"running": False, "progress": "", "done": False, "error": None, "dates": [], "latest": None}

def get_setting(key: str, default: str = "") -> str:
    conn = get_db()
    row  = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row["value"] if row else default

def set_setting(key: str, value: str):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, value))
    conn.commit()
    conn.close()


# ── 页面路由 ──────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {
        "request": request,
        "features": FEATURES,
        "groups": grouped_features(),
    })

@app.get("/file-organizer", response_class=HTMLResponse)
def file_organizer_page(request: Request):
    return templates.TemplateResponse("file_organizer.html", {"request": request})

@app.get("/investment-analysis", response_class=HTMLResponse)
def investment_analysis_page(request: Request):
    return templates.TemplateResponse("investment_analysis.html", {"request": request})

@app.get("/tasks", response_class=HTMLResponse)
def tasks_page(request: Request):
    return templates.TemplateResponse("tasks.html", {"request": request})

@app.get("/project-list", response_class=HTMLResponse)
def project_list_page(request: Request):
    return templates.TemplateResponse("project_list.html", {"request": request})

@app.get("/api/project-list")
def project_list_api(root: str = ""):
    try:
        data = list_project_inventory(root or DEFAULT_PROJECT_ROOT)
        return {"ok": True, **data}
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

class ProjectListOpenRequest(BaseModel):
    path: str = ""

@app.post("/api/project-list/open")
def project_list_open(body: ProjectListOpenRequest):
    try:
        target = Path(body.path).expanduser().resolve()
        root = resolve_project_root()
        if not target.exists() or not target.is_dir():
            raise ValueError("目录不存在")
        if root not in target.parents and target != root:
            raise ValueError("只能打开项目根目录下的目录")
        if sys.platform == "darwin":
            subprocess.Popen(["open", str(target)])
        elif os.name == "nt":
            subprocess.Popen(["explorer", str(target)])
        else:
            subprocess.Popen(["xdg-open", str(target)])
        return {"ok": True, "path": str(target)}
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.get("/requirement-docs", response_class=HTMLResponse)
def requirement_docs_page(request: Request):
    return templates.TemplateResponse("requirement_docs.html", {"request": request})

class RequirementParseRequest(BaseModel):
    description: str = ""

class RequirementDocsListRequest(BaseModel):
    save_dir: str = ""

class RequirementDocImportRequest(BaseModel):
    path: str = ""

class RequirementCliPromptRequest(BaseModel):
    save_dir: str = ""
    source_doc_path: str = ""
    target_model: str = "Claude"
    requirement_name: str = ""

class RequirementDocGenerateRequest(BaseModel):
    save_dir: str = ""
    model_name: str = "Deepseek"
    requirement_name: str = ""
    description: str = ""
    feature_points: list[str] = Field(default_factory=list)
    selected_sections: list[str] = Field(default_factory=list)
    author: str = "程钰"
    source_doc_path: str = ""
    source_doc_text: str = ""
    iteration_notes: str = ""

@app.get("/api/requirement-docs/options")
def requirement_docs_options(
    save_dir: str = "",
    model_name: str = "Deepseek",
    requirement_name: str = "",
    source_doc_path: str = "",
):
    try:
        directory = resolve_save_dir(save_dir)
        next_version = ""
        if requirement_name.strip():
            next_version = next_requirement_version(directory, model_name, requirement_name, source_doc_path)
        return {
            "ok": True,
            "default_dir": str(DEFAULT_REQUIREMENT_DOC_DIR),
            "save_dir": str(directory),
            "model_options": list(MODEL_OPTIONS),
            "section_options": SECTION_OPTIONS,
            "next_version": next_version,
            "documents": list_requirement_documents(directory),
        }
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.post("/api/requirement-docs/parse")
def requirement_docs_parse(body: RequirementParseRequest):
    try:
        return {"ok": True, "data": parse_requirement_description(body.description)}
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.post("/api/requirement-docs/import")
def requirement_docs_import(body: RequirementDocImportRequest):
    try:
        if not body.path.strip():
            raise ValueError("请先选择或填写要解析的 Word 文档路径")
        return {"ok": True, "data": parse_requirement_docx(body.path)}
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.post("/api/requirement-docs/cli-prompt")
def requirement_docs_cli_prompt(body: RequirementCliPromptRequest):
    try:
        if not body.source_doc_path.strip():
            raise ValueError("请先选择来源文档")
        prompt = build_cli_optimization_instruction(
            source_doc_path=body.source_doc_path,
            save_dir=body.save_dir,
            target_model=body.target_model,
            requirement_name=body.requirement_name,
        )
        return {"ok": True, "prompt": prompt}
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.get("/api/requirement-docs/list")
def requirement_docs_list(save_dir: str = ""):
    try:
        directory = resolve_save_dir(save_dir)
        return {"ok": True, "save_dir": str(directory), "documents": list_requirement_documents(directory)}
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.post("/api/requirement-docs/generate")
def requirement_docs_generate(body: RequirementDocGenerateRequest):
    try:
        result = create_requirement_doc(
            save_dir=body.save_dir,
            model_name=body.model_name,
            requirement_name=body.requirement_name,
            description=body.description,
            feature_points=body.feature_points,
            selected_sections=body.selected_sections,
            author=body.author,
            source_doc_path=body.source_doc_path,
            source_doc_text=body.source_doc_text,
            iteration_notes=body.iteration_notes,
        )
        directory = resolve_save_dir(body.save_dir)
        return {
            "ok": True,
            "document": result.__dict__,
            "documents": list_requirement_documents(directory),
            "next_version": next_requirement_version(
                directory,
                body.model_name,
                body.requirement_name or result.requirement_name,
                body.source_doc_path,
            ),
        }
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.post("/api/requirement-docs/open-dir")
def requirement_docs_open_dir(body: RequirementDocsListRequest):
    try:
        directory = resolve_save_dir(body.save_dir)
        directory.mkdir(parents=True, exist_ok=True)
        if sys.platform == "darwin":
            subprocess.Popen(["open", str(directory)])
        elif os.name == "nt":
            subprocess.Popen(["explorer", str(directory)])
        else:
            subprocess.Popen(["xdg-open", str(directory)])
        return {"ok": True, "path": str(directory)}
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

@app.get("/wechat-ops", response_class=HTMLResponse)
def wechat_ops_page(request: Request):
    return templates.TemplateResponse("wechat_ops.html", {"request": request})

class WechatTopicsRequest(BaseModel):
    keyword: str
    audience: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""

class WechatArticleRequest(BaseModel):
    topic: str = ""
    core_viewpoint: str = ""
    content_notes: str = ""
    title: str = ""
    audience: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""
    mode: str = "rich"
    style: str = "auto"

class WechatTitlesRequest(BaseModel):
    markdown: str = ""
    topic: str = ""
    core_viewpoint: str = ""
    content_notes: str = ""
    audience: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""
    style: str = "auto"
    kept_titles: list[str] = Field(default_factory=list)

class WechatConvertRequest(BaseModel):
    markdown: str
    title_hint: str = "wechat"

class WechatImagesRequest(BaseModel):
    topic: str = ""
    core_viewpoint: str = ""
    content_notes: str = ""
    markdown: str = ""
    audience: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""

class WechatFullRequest(BaseModel):
    core_viewpoint: str
    keyword: str = ""
    audience: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""

class WechatProbeRequest(BaseModel):
    provider: str = ""
    model: str = ""
    probe_all: bool = False

class WechatPublishRequest(BaseModel):
    title: str = ""
    markdown: str
    image: dict = Field(default_factory=dict)
    image_position: str = "after_intro"

class WechatRenderImagesRequest(BaseModel):
    images: list[dict] = Field(default_factory=list)
    checkpoint: str = ""
    width: int = 1024
    height: int = 576
    steps: int = 16
    cfg: float = 6.5
    force: bool = False
    speed: str = "auto"  # auto=有加速LoRA就用 / off=强制原始采样参数

def _wechat_model_for_provider(provider_key: str, provider_item: dict, override: str = "") -> str:
    manual_model = (override or "").strip()
    if manual_model:
        return manual_model
    saved_model = (get_setting(f"wechat_{provider_key}_model") or "").strip()
    return saved_model or str(provider_item.get("model") or "").strip()

def _wechat_provider_options() -> list[dict]:
    config = load_config(BASE_DIR)
    options = []
    for key, item in config.providers.items():
        setting_key = str(item.get("setting_key") or "")
        env_key = str(item.get("api_key_env") or "")
        db_key = get_setting(setting_key) if setting_key else ""
        env_value = os.getenv(env_key, "") if env_key else ""
        options.append({
            "key": key,
            "label": item.get("label") or key,
            "model": _wechat_model_for_provider(key, item),
            "setting_key": setting_key,
            "api_key_env": env_key,
            "api_key_configured": bool((db_key or "").strip() or (env_value or "").strip()),
        })
    return options

def _wechat_selected_provider(provider: str = "") -> str:
    config = load_config(BASE_DIR)
    configured = set(config.providers.keys())
    default_provider = config.provider if config.provider in configured else next(iter(configured), "deepseek")
    selected = (provider or get_setting("wechat_llm_provider", default_provider) or default_provider).strip()
    if selected not in configured:
        selected = default_provider
    return selected

def _wechat_content_assistant(provider: str = "", model: str = "") -> WechatContentAssistant:
    selected = _wechat_selected_provider(provider)
    base_config = load_config(BASE_DIR)
    provider_item = base_config.providers.get(selected) or {}
    setting_key = str(provider_item.get("setting_key") or "")
    overrides = {
        "api_key": get_setting(setting_key) if setting_key else "",
        "model": _wechat_model_for_provider(selected, provider_item, model),
    }
    return WechatContentAssistant(load_config(
        BASE_DIR,
        provider_override=selected,
        provider_overrides=overrides,
    ))

def _wechat_json(result) -> dict:
    return {
        "ok": True,
        "data": result.data,
        "output_files": result.output_files,
        "provider": result.provider,
        "model": result.model,
    }

def _wechat_error(exc: Exception):
    status = 400 if isinstance(exc, (ValueError, LLMError, FileNotFoundError, ComfyUIError)) else 500
    return JSONResponse({"ok": False, "error": str(exc)}, status_code=status)

def _output_public_url(file_path: str) -> str:
    try:
        rel = Path(file_path).resolve().relative_to(OUTPUT_DIR.resolve())
    except Exception:
        return ""
    return "/output/" + "/".join(urllib.parse.quote(part) for part in rel.parts)

def _wechat_probe_one(provider: str = "", model: str = "") -> dict:
    started = datetime.now()
    selected = _wechat_selected_provider(provider)
    assistant = _wechat_content_assistant(selected, model)
    try:
        resp = assistant.client.generate(
            system="你是模型连通性检测助手。",
            prompt="只回复 OK",
            max_tokens=3,
            temperature=0,
        )
        elapsed_ms = int((datetime.now() - started).total_seconds() * 1000)
        return {
            "provider": resp.provider,
            "model": resp.model,
            "available": True,
            "message": resp.text or "OK",
            "latency_ms": elapsed_ms,
        }
    except LLMError as exc:
        elapsed_ms = int((datetime.now() - started).total_seconds() * 1000)
        return {
            "provider": selected,
            "model": assistant.client.model,
            "available": False,
            "error": str(exc),
            "latency_ms": elapsed_ms,
        }

@app.get("/api/wechat-assistant/status")
def wechat_assistant_status(provider: str = ""):
    try:
        selected = _wechat_selected_provider(provider)
        config = load_config(BASE_DIR, provider_override=selected)
        key_env = str(config.provider_config.get("api_key_env") or "")
        setting_key = str(config.provider_config.get("setting_key") or "")
        db_key = get_setting(setting_key) if setting_key else ""
        return {
            "ok": True,
            "provider": config.provider,
            "provider_label": config.provider_config.get("label") or config.provider,
            "model": _wechat_model_for_provider(selected, config.provider_config),
            "api_key_env": key_env,
            "setting_key": setting_key,
            "api_key_configured": bool((db_key or "").strip() or (os.getenv(key_env, "") if key_env else "").strip()),
            "output_dir": str(config.output_dir),
            "providers": _wechat_provider_options(),
        }
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/probe")
def wechat_assistant_probe(body: WechatProbeRequest):
    try:
        if body.probe_all:
            results = []
            for item in _wechat_provider_options():
                results.append(_wechat_probe_one(item["key"], ""))
            return {"ok": True, "results": results}
        return {"ok": True, "results": [_wechat_probe_one(body.provider, body.model)]}
    except Exception as exc:
        return _wechat_error(exc)

@app.get("/api/wechat-assistant/hot-topics")
def wechat_assistant_hot_topics(limit: int = 50):
    try:
        return {"ok": True, "data": fetch_hot_topics(limit=max(1, min(limit, 50)))}
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/topics")
def wechat_assistant_topics(body: WechatTopicsRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).generate_topics(
            keyword=body.keyword,
            audience=body.audience,
            tone=body.tone,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/article")
def wechat_assistant_article(body: WechatArticleRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).write_article(
            topic=body.topic,
            core_viewpoint=body.core_viewpoint,
            content_notes=body.content_notes,
            title=body.title,
            audience=body.audience,
            tone=body.tone,
            mode=body.mode,
            style=body.style,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/article-stream")
def wechat_assistant_article_stream(body: WechatArticleRequest):
    def generate():
        try:
            assistant = _wechat_content_assistant(body.provider, body.model)
            for chunk in assistant.write_article_stream(
                topic=body.topic,
                core_viewpoint=body.core_viewpoint,
                content_notes=body.content_notes,
                title=body.title,
                audience=body.audience,
                tone=body.tone,
                mode=body.mode,
                style=body.style,
            ):
                if chunk.startswith("\x00"):
                    meta = chunk[1:]
                    yield f"data: {json.dumps({'done': True, 'meta': json.loads(meta)})}\n\n"
                else:
                    yield f"data: {json.dumps({'text': chunk})}\n\n"
        except Exception as exc:
            yield f"data: {json.dumps({'error': str(exc)})}\n\n"
    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )

@app.post("/api/wechat-assistant/titles")
def wechat_assistant_titles(body: WechatTitlesRequest):
    try:
        assistant = _wechat_content_assistant(body.provider, body.model)
        if (body.markdown or "").strip():
            result = assistant.optimize_titles(markdown=body.markdown)
        else:
            result = assistant.generate_titles_for_topic(
                topic=body.topic,
                core_viewpoint=body.core_viewpoint,
                content_notes=body.content_notes,
                audience=body.audience,
                tone=body.tone,
                style=body.style,
                kept_titles=body.kept_titles,
            )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/images")
def wechat_assistant_images(body: WechatImagesRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).generate_image_ideas(
            topic=body.topic,
            core_viewpoint=body.core_viewpoint,
            content_notes=body.content_notes,
            markdown=body.markdown,
            audience=body.audience,
            tone=body.tone,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

@app.get("/api/wechat-assistant/comfy-status")
async def wechat_assistant_comfy_status():
    try:
        models = list_checkpoints(_COMFY_DIR)
        default_checkpoint = choose_checkpoint(_COMFY_DIR) if models else ""
        return {
            "ok": True,
            "alive": await _comfy_alive(),
            "pid": _comfy_find_pid(),
            "url": _COMFY_FALLBACK_URLS[0],
            "default_checkpoint": default_checkpoint,
            "models": models,
            "speed_lora": find_speed_lora(_COMFY_DIR),
        }
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/render-images")
async def wechat_assistant_render_images(body: WechatRenderImagesRequest):
    try:
        if not body.images:
            raise ValueError("请先生成配图方案")
        comfy_autostarted = False
        if not await _comfy_alive():
            # 自动拉起本地 ComfyUI（冷启动约 30-60 秒），不再要求用户手动启动
            start_resp = await comfyui_start()
            if start_resp.status_code >= 400:
                raise ComfyUIError(
                    f"ComfyUI 自动启动失败：{start_resp.body.decode('utf-8', 'replace')}"
                )
            comfy_autostarted = True
            for _ in range(90):
                if await _comfy_alive():
                    break
                await asyncio.sleep(1)
            else:
                raise ComfyUIError("ComfyUI 自动启动超时（90秒），请到 ComfyUI 管理页查看日志")

        config = load_config(BASE_DIR)
        rendered = []
        errors = []
        files = []
        checkpoint = body.checkpoint.strip() or choose_checkpoint(_COMFY_DIR)

        # 生图期间阻止系统空闲睡眠（实测睡眠会把一次生图拖到 17+ 分钟并触发超时）
        caffeinate_proc = None
        try:
            caffeinate_proc = subprocess.Popen(["/usr/bin/caffeinate", "-i"])
        except Exception:
            pass
        try:
            for index, item in enumerate(body.images[:4]):
                image = dict(item or {})
                if image.get("image_url") and image.get("image_path") and not body.force:
                    rendered.append(image)
                    continue

                prompt = str(
                    image.get("prompt")
                    or image.get("scene")
                    or image.get("caption")
                    or image.get("name")
                    or ""
                ).strip()
                try:
                    result = await asyncio.to_thread(
                        generate_comfy_image,
                        prompt=prompt,
                        output_dir=config.output_dir,
                        comfy_dir=_COMFY_DIR,
                        base_url=_COMFY_FALLBACK_URLS[0],
                        checkpoint=checkpoint,
                        width=body.width,
                        height=body.height,
                        steps=body.steps,
                        cfg=body.cfg,
                        speed=body.speed,
                    )
                    image.update({
                        "image_path": result.image_path,
                        "image_url": _output_public_url(result.image_path),
                        "speed_lora": result.speed_lora,
                        "checkpoint": result.checkpoint,
                        "seed": result.seed,
                        "width": result.width,
                        "height": result.height,
                        "comfy_prompt": result.prompt,
                        "comfy_filename": result.comfy_filename,
                        "comfy_subfolder": result.comfy_subfolder,
                        "generated_at": datetime.now().isoformat(timespec="seconds"),
                    })
                    files.append(result.image_path)
                except Exception as exc:
                    image["render_error"] = str(exc)
                    errors.append({"index": index, "error": str(exc)})
                rendered.append(image)
        finally:
            if caffeinate_proc:
                caffeinate_proc.terminate()

        return {
            "ok": True,
            "data": {
                "images": rendered,
                "errors": errors,
                "checkpoint": checkpoint,
                "comfy_autostarted": comfy_autostarted,
            },
            "output_files": files,
        }
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/publish-text")
def wechat_assistant_publish_text(body: WechatPublishRequest):
    try:
        result = _wechat_content_assistant().format_publish_text(
            title=body.title,
            markdown=body.markdown,
            image=body.image,
            image_position=body.image_position,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/publish-html")
def wechat_assistant_publish_html(body: WechatPublishRequest):
    try:
        result = _wechat_content_assistant().format_publish_html(
            title=body.title,
            markdown=body.markdown,
            image=body.image,
            image_position=body.image_position,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

class WechatDiagnoseRequest(BaseModel):
    markdown: str
    core_viewpoint: str = ""

@app.post("/api/wechat-assistant/diagnose")
def wechat_assistant_diagnose(body: WechatDiagnoseRequest):
    try:
        result = _wechat_content_assistant().diagnose_article(
            markdown=body.markdown,
            core_viewpoint=body.core_viewpoint,
        )
        return {"ok": True, "data": result}
    except Exception as exc:
        return _wechat_error(exc)

class WechatHookRequest(BaseModel):
    topic: str = ""
    core_viewpoint: str = ""
    article_opening: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/hooks")
def wechat_assistant_hooks(body: WechatHookRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).generate_hooks(
            topic=body.topic, core_viewpoint=body.core_viewpoint,
            article_opening=body.article_opening, tone=body.tone,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


class WechatQuotesRequest(BaseModel):
    markdown: str
    topic: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/quotes")
def wechat_assistant_quotes(body: WechatQuotesRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).extract_quotes(
            markdown=body.markdown, topic=body.topic,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


class WechatEndingRequest(BaseModel):
    topic: str = ""
    core_viewpoint: str = ""
    markdown: str = ""
    audience: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/engagement-ending")
def wechat_assistant_ending(body: WechatEndingRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).generate_engagement_ending(
            topic=body.topic, core_viewpoint=body.core_viewpoint,
            markdown=body.markdown, audience=body.audience, tone=body.tone,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


class WechatSeriesRequest(BaseModel):
    topic: str = ""
    core_viewpoint: str = ""
    audience: str = ""
    tone: str = ""
    num_episodes: int = 5
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/series-plan")
def wechat_assistant_series(body: WechatSeriesRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).plan_content_series(
            topic=body.topic, core_viewpoint=body.core_viewpoint,
            audience=body.audience, tone=body.tone, num_episodes=body.num_episodes,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


class WechatPositioningRequest(BaseModel):
    article_topic: str
    article_summary: str = ""
    account_positioning: str
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/check-positioning")
def wechat_assistant_check_positioning(body: WechatPositioningRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).check_account_positioning(
            article_topic=body.article_topic,
            article_summary=body.article_summary,
            account_positioning=body.account_positioning,
        )
        return {"ok": True, **result}
    except Exception as exc:
        return _wechat_error(exc)


# ── multi-platform ────────────────────────────────────────────────────────────
class WechatPlatformRequest(BaseModel):
    markdown: str
    title: str = ""
    platform: str = "zhihu"
    core_viewpoint: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/platform-format")
def wechat_platform_format(body: WechatPlatformRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).format_for_platform(
            markdown=body.markdown, title=body.title,
            platform=body.platform, core_viewpoint=body.core_viewpoint,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


# ── compliance ────────────────────────────────────────────────────────────────
class WechatComplianceRequest(BaseModel):
    markdown: str
    title: str = ""
    deep: bool = False
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/compliance")
def wechat_compliance(body: WechatComplianceRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).check_compliance(
            markdown=body.markdown, title=body.title, deep=body.deep,
        )
        return {"ok": True, **result}
    except Exception as exc:
        return _wechat_error(exc)


# ── materials library ─────────────────────────────────────────────────────────
class WechatMaterialAddRequest(BaseModel):
    title: str
    content: str
    category: str = "通用"
    tags: list[str] = []

@app.get("/api/wechat-assistant/materials")
def wechat_materials_list(category: str = "", query: str = ""):
    try:
        items = _wechat_content_assistant().list_materials(category=category, query=query)
        return {"ok": True, "materials": items, "total": len(items)}
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/materials")
def wechat_materials_add(body: WechatMaterialAddRequest):
    try:
        item = _wechat_content_assistant().add_material(
            title=body.title, content=body.content,
            category=body.category, tags=body.tags,
        )
        return {"ok": True, "material": item}
    except Exception as exc:
        return _wechat_error(exc)

@app.delete("/api/wechat-assistant/materials/{material_id}")
def wechat_materials_delete(material_id: str):
    try:
        deleted = _wechat_content_assistant().delete_material(material_id)
        return {"ok": deleted, "error": "" if deleted else "未找到该素材"}
    except Exception as exc:
        return _wechat_error(exc)


# ── knowledge extraction ──────────────────────────────────────────────────────
class WechatKnowledgeExtractRequest(BaseModel):
    markdown: str
    title: str = ""
    topic: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/extract-knowledge")
def wechat_extract_knowledge(body: WechatKnowledgeExtractRequest):
    try:
        svc = _wechat_content_assistant(provider=body.provider, model=body.model)
        result = svc.extract_knowledge(
            markdown=body.markdown, title=body.title, topic=body.topic,
        )
        return {"ok": True, "data": result}
    except Exception as exc:
        return _wechat_error(exc)

# ── topic lifecycle ────────────────────────────────────────────────────────────
class WechatLifecycleRequest(BaseModel):
    topic: str
    source_count: int = 1
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/topic-lifecycle")
def wechat_topic_lifecycle(body: WechatLifecycleRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).analyze_topic_lifecycle(
            topic=body.topic, source_count=body.source_count,
        )
        return result
    except Exception as exc:
        return _wechat_error(exc)


class WechatExcerptRequest(BaseModel):
    markdown: str
    title: str = ""
    core_viewpoint: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/excerpt")
def wechat_assistant_excerpt(body: WechatExcerptRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).generate_excerpt(
            markdown=body.markdown,
            title=body.title,
            core_viewpoint=body.core_viewpoint,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


class WechatRewriteRequest(BaseModel):
    paragraph: str
    instruction: str = ""
    topic: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/rewrite-paragraph")
def wechat_assistant_rewrite(body: WechatRewriteRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).rewrite_paragraph(
            paragraph=body.paragraph,
            instruction=body.instruction,
            topic=body.topic,
            tone=body.tone,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


class WechatCoverImageRequest(BaseModel):
    topic: str = ""
    core_viewpoint: str = ""
    tone: str = ""
    provider: str = ""
    model: str = ""

@app.post("/api/wechat-assistant/cover-image-idea")
def wechat_assistant_cover_image_idea(body: WechatCoverImageRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).generate_cover_image_idea(
            topic=body.topic,
            core_viewpoint=body.core_viewpoint,
            tone=body.tone,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)


@app.get("/api/wechat-assistant/image-proxy")
async def wechat_assistant_image_proxy(keywords: str = "", width: int = 1024, height: int = 576):
    """免费 AI 生图（pollinations.ai，无需 Key），并落盘到输出目录供发布引用。

    原 source.unsplash.com 已停止服务（503），改为按完整英文提示词生成真实图片。
    """
    from fastapi.responses import Response as FastapiResponse
    import urllib.request as _ur
    import zlib

    kw = re.sub(r"[^\w\s,.\-]", " ", (keywords or "").strip())[:300].strip()
    if not kw:
        return JSONResponse({"ok": False, "error": "缺少提示词"}, status_code=400)
    seed = zlib.crc32(kw.encode("utf-8")) % 1_000_000  # 同一提示词稳定出同一张图
    url = (
        f"https://image.pollinations.ai/prompt/{urllib.parse.quote(kw)}"
        f"?width={max(256, min(int(width or 1024), 1536))}"
        f"&height={max(256, min(int(height or 576), 1536))}"
        f"&nologo=true&seed={seed}"
    )

    def _fetch() -> tuple[bytes, str]:
        from wechat_assistant._ssl import build_ssl_context
        req = _ur.Request(url, headers={"User-Agent": "investment-hub/1.0"})
        with _ur.urlopen(req, timeout=90, context=build_ssl_context()) as resp:
            return resp.read(), resp.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()

    try:
        data, ctype = await asyncio.to_thread(_fetch)
        headers = {}
        try:
            config = load_config(BASE_DIR)
            image_dir = config.output_dir / "images"
            image_dir.mkdir(parents=True, exist_ok=True)
            ext = ".png" if "png" in ctype else ".jpg"
            target = image_dir / f"free_{seed}{ext}"
            target.write_bytes(data)
            headers = {
                "X-Image-Path": str(target),
                "X-Image-Url": _output_public_url(str(target)),
            }
        except Exception:
            pass  # 落盘失败不影响图片返回
        return FastapiResponse(content=data, media_type=ctype, headers=headers)
    except Exception as exc:
        return JSONResponse({"ok": False, "error": f"免费生图失败：{exc}"}, status_code=502)

@app.post("/api/wechat-assistant/convert")
def wechat_assistant_convert(body: WechatConvertRequest):
    try:
        result = _wechat_content_assistant().convert_markdown(
            markdown=body.markdown,
            title_hint=body.title_hint,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

@app.post("/api/wechat-assistant/full")
def wechat_assistant_full(body: WechatFullRequest):
    try:
        result = _wechat_content_assistant(body.provider, body.model).generate_full(
            core_viewpoint=body.core_viewpoint,
            keyword=body.keyword,
            audience=body.audience,
            tone=body.tone,
        )
        return _wechat_json(result)
    except Exception as exc:
        return _wechat_error(exc)

@app.get("/san14-save", response_class=HTMLResponse)
def san14_save_page(request: Request):
    return templates.TemplateResponse("san14_save.html", {"request": request})

def _san14_default_remote_dir() -> Path:
    downloads = Path.home() / "Downloads"
    san14_root = downloads / "三国志14"
    candidates: list[Path] = []
    search_specs = [
        (san14_root, ("remote", "*/remote", "*/*/remote")),
        (downloads, ("三国志14*/remote", "三国志14*/*/remote", "三国志14*/*/*/remote")),
    ]
    for root, patterns in search_specs:
        if root.exists():
            for pattern in patterns:
                try:
                    candidates.extend(p for p in root.glob(pattern) if p.is_dir())
                except Exception:
                    continue
            if candidates:
                break
    if candidates:
        unique = {str(p): p for p in candidates}
        ordered = sorted(unique.values(), key=lambda p: p.stat().st_mtime, reverse=True)
        return ordered[0]
    return Path.home()

def _san14_progress_time(data: bytes) -> str:
    for off in range(0, min(256, max(0, len(data) - 12)), 2):
        vals = [int.from_bytes(data[off + i:off + i + 2], "little") for i in range(0, 12, 2)]
        year, month, marker, day, hour, minute = vals
        if (
            2000 <= year <= 2035
            and 1 <= month <= 12
            and marker == 0
            and 1 <= day <= 31
            and 0 <= hour <= 23
            and 0 <= minute <= 59
        ):
            return f"{year:04d}-{month:02d}-{day:02d} {hour:02d}:{minute:02d}"
    return ""

def _san14_lwc_info(data: bytes) -> dict | None:
    magic = b"LWC\x1a"
    off = data.find(magic)
    if off < 0 or len(data) < off + 268:
        return None

    uncompressed_size = int.from_bytes(data[off + 4:off + 8], "little")
    compressed_size = int.from_bytes(data[off + 8:off + 12], "little")
    payload_offset = off + 268
    payload_end = payload_offset + compressed_size
    return {
        "offset": off,
        "uncompressed_size": uncompressed_size,
        "compressed_size": compressed_size,
        "table_size": 256,
        "payload_offset": payload_offset,
        "payload_size": max(0, len(data) - payload_offset),
        "complete": payload_end <= len(data),
    }

class San14LwcError(Exception):
    pass

def _san14_lwc_decompress(data: bytes) -> tuple[bytes, dict]:
    info = _san14_lwc_info(data)
    if not info:
        raise San14LwcError("未发现 LWC 压缩块")
    if not info["complete"]:
        raise San14LwcError("LWC 压缩流不完整")

    off = info["offset"]
    uncompressed_size = info["uncompressed_size"]
    compressed_size = info["compressed_size"]
    table = data[off + 12:off + 268]
    payload = data[off + 268:off + 268 + compressed_size]
    out = bytearray()
    pos = 0
    bitbuf = 0
    bits_left = 0

    def read_bit() -> int:
        nonlocal pos, bitbuf, bits_left
        if bits_left == 0:
            if pos >= len(payload):
                raise San14LwcError("LWC 位流提前结束")
            bitbuf = payload[pos]
            pos += 1
            bits_left = 8
        bit = (bitbuf >> 7) & 1
        bitbuf = (bitbuf << 1) & 0xff
        bits_left -= 1
        return bit

    def read_value() -> int:
        prefix = 0
        bit_count = 0
        while True:
            bit = read_bit()
            bit_count += 1
            prefix = bit + 2 * prefix
            if bit == 0:
                break

        suffix = 0
        for _ in range(bit_count):
            suffix = read_bit() + 2 * suffix
        return prefix + suffix

    while len(out) < uncompressed_size:
        symbol = read_value()
        if symbol < 256:
            out.append(table[symbol])
            continue

        distance = symbol - 256
        length = read_value() + 3
        if distance <= 0 or distance > len(out):
            raise San14LwcError(f"LWC 回溯距离异常：{distance}")
        for _ in range(length):
            out.append(out[-distance])
            if len(out) > uncompressed_size:
                raise San14LwcError("LWC 解压结果超过声明大小")

    return bytes(out), {
        "offset": off,
        "uncompressed_size": uncompressed_size,
        "compressed_size": compressed_size,
        "consumed_payload_bytes": pos,
        "remaining_bits": bits_left,
        "sha256": hashlib.sha256(out).hexdigest(),
    }

def _san14_lwc_compress_literals(data: bytes, table: bytes | None = None) -> bytes:
    if table is None or len(table) != 256 or len(set(table)) != 256:
        table = bytes(range(256))
    inverse = {value: idx for idx, value in enumerate(table)}
    payload = bytearray()
    current = 0
    used = 0

    def write_bit(bit: int) -> None:
        nonlocal current, used
        current = ((current << 1) | (bit & 1)) & 0xff
        used += 1
        if used == 8:
            payload.append(current)
            current = 0
            used = 0

    def write_value(value: int) -> None:
        bit_count = 1
        while value >= (1 << (bit_count + 1)) - 2:
            bit_count += 1
        prefix = (1 << bit_count) - 2
        suffix = value - prefix
        for _ in range(bit_count - 1):
            write_bit(1)
        write_bit(0)
        for shift in range(bit_count - 1, -1, -1):
            write_bit((suffix >> shift) & 1)

    for byte in data:
        write_value(inverse[byte])

    if used:
        payload.append((current << (8 - used)) & 0xff)

    return b"LWC\x1a" + struct.pack("<II", len(data), len(payload)) + table + bytes(payload)

def _san14_find_offsets(data: bytes, needle: bytes, limit: int = 8) -> list[str]:
    offsets: list[str] = []
    start = 0
    while len(offsets) < limit:
        idx = data.find(needle, start)
        if idx < 0:
            break
        offsets.append(f"0x{idx:x}")
        start = idx + 1
    return offsets

def _san14_expected_note(path: Path) -> dict | None:
    note_roots = [path.parent, path.parent.parent, path.parent.parent.parent]
    candidates: list[Path] = []
    for root in note_roots:
        if root.exists():
            candidates.extend(p for p in root.glob("*.txt") if p.is_file())
    if not candidates:
        return None

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    note_path = candidates[0]
    text = note_path.read_text(encoding="utf-8-sig", errors="replace")
    resources = []
    for label in ("资金", "军粮", "士兵"):
        m = re.search(label + r"(?:是|:|：)?\s*(\d+)", text)
        if m:
            resources.append({"label": label, "value": int(m.group(1))})

    name_aliases = {
        "吕布": "呂布",
        "张辽": "張遼",
        "高顺": "高順",
        "刘备": "劉備",
        "孙坚": "孫堅",
        "袁绍": "袁紹",
    }
    people = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or not any(k in line for k in ("统率", "武力", "智", "政治", "魅力")):
            continue
        name = re.split(r"[，,]", line, maxsplit=1)[0].strip()
        stats = {}
        patterns = {
            "统率": r"统率\s*(\d+)",
            "武力": r"武力\s*(\d+)",
            "智力": r"智(?:力|,|，)?\s*(\d+)",
            "政治": r"政治\s*(\d+)",
            "魅力": r"魅力\s*(\d+)",
        }
        for key, pattern in patterns.items():
            m = re.search(pattern, line)
            if m:
                stats[key] = int(m.group(1))
        if name and stats:
            people.append({
                "name": name,
                "traditional": name_aliases.get(name, name),
                "stats": stats,
            })

    return {
        "source": str(note_path),
        "resources": resources,
        "people": people,
    }

def _san14_numeric_checks(data: bytes, expected: dict | None) -> list[dict]:
    if expected and expected.get("resources"):
        resources = expected["resources"]
    else:
        resources = [
            {"label": "资金", "value": 35000},
            {"label": "军粮", "value": 201000},
            {"label": "士兵", "value": 90000},
        ]

    checks = []
    seen: set[tuple[str, int, str]] = set()
    for item in resources:
        label = item["label"]
        value = int(item["value"])
        for form in _san14_numeric_forms(value):
            key = (label, form["stored_value"], form["type"])
            if key in seen:
                continue
            seen.add(key)
            checks.append({
                "label": label,
                "display_value": value,
                "storage": form["storage"],
                "stored_value": form["stored_value"],
                "type": form["type"],
                "offsets": _san14_find_offsets(data, form["needle"]),
            })
    return checks

def _san14_numeric_forms(value: int) -> list[dict]:
    variants = [("原值", value)]
    if value % 10 == 0:
        variants.append(("/10", value // 10))

    forms = []
    for storage_label, stored_value in variants:
        if 0 <= stored_value <= 0xffffffff:
            forms.append({
                "storage": storage_label,
                "stored_value": stored_value,
                "type": "u32_le",
                "needle": struct.pack("<I", stored_value),
            })
        if 0 <= stored_value <= 0xffff:
            forms.append({
                "storage": storage_label,
                "stored_value": stored_value,
                "type": "u16_le",
                "needle": struct.pack("<H", stored_value),
            })
        if 0 <= stored_value <= 0xff:
            forms.append({
                "storage": storage_label,
                "stored_value": stored_value,
                "type": "u8",
                "needle": bytes([stored_value]),
            })
    return forms

def _san14_probe_decompressed(data: bytes, path: Path | None = None) -> dict:
    expected = _san14_expected_note(path) if path else None
    people = [
        "呂布", "吕布", "張遼", "张辽", "高順", "高顺",
        "劉備", "刘备", "曹操", "孫堅", "孙坚", "董卓", "袁紹", "袁绍",
    ]
    if expected:
        for item in expected.get("people", []):
            for name in (item.get("name"), item.get("traditional")):
                if name and name not in people:
                    people.append(name)

    person_hits = []
    for name in people:
        hits = _san14_find_offsets(data, name.encode("utf-16le"))
        if hits:
            person_hits.append({"name": name, "encoding": "UTF-16LE", "offsets": hits})

    checks = _san14_numeric_checks(data, expected)
    sample_values = [
        {
            "value": item["stored_value"],
            "display_value": item["display_value"],
            "label": item["label"],
            "storage": item["storage"],
            "type": item["type"],
            "offsets": item["offsets"],
        }
        for item in checks
        if item["offsets"]
    ]

    return {
        "expected": expected,
        "person_hits": person_hits,
        "sample_value_checks": checks,
        "sample_value_hits": sample_values,
        "head_hex": data[:64].hex(" "),
    }

def _san14_file_info(path: Path) -> dict:
    data = path.read_bytes()
    name = path.name
    lower = name.lower()
    progress_magic = b"SN14SVEXVER0000"
    is_progress = data[4:4 + len(progress_magic)] == progress_magic
    lwc = _san14_lwc_info(data)
    is_lwc = data.startswith(b"LWC\x1a")
    language = ""
    if "sc" in lower:
        language = "简体"
    elif "tc" in lower:
        language = "繁体"

    kind = "unknown"
    kind_label = "未知文件"
    editable = False
    if is_progress and lower.startswith("svdex"):
        kind, kind_label, editable = "manual", "手动存档", True
    elif is_progress and lower.startswith("autosdex"):
        kind, kind_label, editable = "auto", "自动存档", True
    elif is_lwc and lower.startswith("configs"):
        kind, kind_label = "config", "设置文件"
    elif is_lwc and lower.startswith("prdata"):
        kind, kind_label = "profile", "玩家记录"

    slot = ""
    m = re.search(r"(\d{2})(?=\.s14$)", lower)
    if m:
        slot = m.group(1)

    stat = path.stat()
    return {
        "name": name,
        "path": str(path),
        "kind": kind,
        "kind_label": kind_label,
        "language": language,
        "slot": slot,
        "editable": editable,
        "size": stat.st_size,
        "mtime": int(stat.st_mtime),
        "save_time": _san14_progress_time(data) if is_progress else "",
        "magic": "SN14SVEXVER0000" if is_progress else ("LWC" if is_lwc else ""),
        "lwc": lwc,
    }

@app.get("/api/san14/browse")
def san14_browse(path: str = ""):
    """目录浏览器接口，返回子目录列表和 .s14 文件列表，供前端文件选择器使用。"""
    default_dir = _san14_default_remote_dir()
    target = Path(path).expanduser() if path else (default_dir.parent if default_dir else Path.home())
    if target.exists() and target.is_file():
        target = target.parent
    if not target.exists() or not target.is_dir():
        return JSONResponse({"error": "目录不存在或无法访问"}, status_code=400)

    dirs, files = [], []
    try:
        entries = sorted(target.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
    except Exception as e:
        return JSONResponse({"error": f"读取目录失败: {e}"}, status_code=400)

    for entry in entries:
        if entry.name.startswith("."):
            continue
        try:
            if entry.is_dir():
                dirs.append({"name": entry.name, "path": str(entry)})
            elif entry.is_file() and entry.suffix.lower() == ".s14":
                stat = entry.stat()
                files.append({
                    "name": entry.name,
                    "path": str(entry),
                    "size": stat.st_size,
                    "mtime": int(stat.st_mtime),
                })
        except Exception:
            continue

    parent = str(target.parent) if target.parent != target else ""

    # 快捷根目录：下载目录下的三国志14相关路径
    roots = []
    for candidate in [
        Path.home() / "Downloads" / "三国志14",
        Path.home() / "Downloads",
        Path.home() / "Documents",
        Path.home(),
    ]:
        if candidate.exists() and candidate.is_dir():
            roots.append({"name": candidate.name, "path": str(candidate)})

    return {
        "ok": True,
        "path": str(target),
        "parent": parent,
        "roots": roots,
        "dirs": dirs,
        "files": files,
    }


@app.get("/api/san14/files")
def san14_files(path: str = ""):
    target = Path(path).expanduser() if path else _san14_default_remote_dir()
    if target.exists() and target.is_file():
        target = target.parent
    if not target.exists() or not target.is_dir():
        return JSONResponse({"error": "目录不存在或无法访问"}, status_code=400)

    files = []
    for item in sorted(target.glob("*.s14")):
        try:
            files.append(_san14_file_info(item))
        except Exception:
            continue

    priority = {"manual": 0, "auto": 1, "config": 2, "profile": 3, "unknown": 4}
    files.sort(key=lambda f: (priority.get(f["kind"], 9), -f["mtime"], f["name"]))
    recommended = next((f for f in files if f["kind"] == "manual"), None)

    return {
        "ok": True,
        "path": str(target),
        "files": files,
        "recommended": recommended,
    }

@app.get("/api/san14/analyze")
def san14_analyze(file: str):
    target = Path(file).expanduser()
    if not target.exists() or not target.is_file():
        return JSONResponse({"error": "存档文件不存在或无法访问"}, status_code=400)
    if target.suffix.lower() != ".s14":
        return JSONResponse({"error": "请选择 .s14 存档文件"}, status_code=400)

    try:
        raw = target.read_bytes()
        decompressed, lwc_meta = _san14_lwc_decompress(raw)
        file_info = _san14_file_info(target)
    except San14LwcError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"error": f"解析失败：{exc}"}, status_code=500)

    return {
        "ok": True,
        "file": file_info,
        "lwc": lwc_meta,
        "probe": _san14_probe_decompressed(decompressed, target),
    }

class San14FileRequest(BaseModel):
    file: str

@app.post("/api/san14/rebuild-test")
def san14_rebuild_test(body: San14FileRequest):
    target = Path(body.file).expanduser()
    if not target.exists() or not target.is_file():
        return JSONResponse({"error": "存档文件不存在或无法访问"}, status_code=400)
    if target.suffix.lower() != ".s14":
        return JSONResponse({"error": "请选择 .s14 存档文件"}, status_code=400)

    try:
        raw = target.read_bytes()
        info = _san14_lwc_info(raw)
        if not info:
            return JSONResponse({"error": "未发现 LWC 压缩块"}, status_code=400)
        decompressed, lwc_meta = _san14_lwc_decompress(raw)
        original_table = raw[info["offset"] + 12:info["offset"] + 268]
        rebuilt_lwc = _san14_lwc_compress_literals(decompressed, original_table)
        rebuilt_raw = raw[:info["offset"]] + rebuilt_lwc
        check, check_meta = _san14_lwc_decompress(rebuilt_raw)
        if check != decompressed:
            return JSONResponse({"error": "重新编码校验失败，已取消写入"}, status_code=500)

        out_dir = target.parent / "new"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / target.name
        out_path.write_bytes(rebuilt_raw)
    except San14LwcError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"error": f"生成失败：{exc}"}, status_code=500)

    return {
        "ok": True,
        "path": str(out_path),
        "old_size": len(raw),
        "new_size": len(rebuilt_raw),
        "sha256": lwc_meta["sha256"],
        "check": {
            "uncompressed_size": check_meta["uncompressed_size"],
            "compressed_size": check_meta["compressed_size"],
            "consumed_payload_bytes": check_meta["consumed_payload_bytes"],
            "remaining_bits": check_meta["remaining_bits"],
        },
        "note": "这是未修改字段的格式测试存档，用于验证重新编码后的 LWC 能被游戏读取。",
    }

class San14DiffRequest(BaseModel):
    file_a: str
    file_b: str
    label: str = ""

def _san14_diff_regions(data_a: bytes, data_b: bytes, cluster_gap: int = 32) -> list[dict]:
    if len(data_a) != len(data_b):
        return [{"note": f"文件长度不一致：{len(data_a)} vs {len(data_b)}，仅对比较短部分"}]

    diff_offsets = [i for i in range(len(data_a)) if data_a[i] != data_b[i]]
    if not diff_offsets:
        return []

    # 聚合相邻差异为 region
    regions = []
    start = diff_offsets[0]
    end = diff_offsets[0]
    for off in diff_offsets[1:]:
        if off - end <= cluster_gap:
            end = off
        else:
            regions.append((start, end))
            start = off
            end = off
    regions.append((start, end))

    result = []
    for rstart, rend in regions:
        ctx_start = max(0, rstart - 8)
        ctx_end = min(len(data_a), rend + 9)
        seg_a = data_a[ctx_start:ctx_end]
        seg_b = data_b[ctx_start:ctx_end]
        changed_len = rend - rstart + 1

        interp = []
        for off in range(rstart, min(rend + 1, rend + 5)):
            for size, fmt, label in ((1, None, "u8"), (2, "<H", "u16_le"), (4, "<I", "u32_le")):
                if off + size <= len(data_a):
                    va = data_a[off:off + size]
                    vb = data_b[off:off + size]
                    if va == vb:
                        continue
                    if size == 1:
                        ia, ib = va[0], vb[0]
                    else:
                        ia = struct.unpack(fmt, va)[0]
                        ib = struct.unpack(fmt, vb)[0]
                    interp.append({
                        "offset": f"0x{off:x}",
                        "type": label,
                        "before": ia,
                        "after": ib,
                        "delta": ib - ia,
                    })
            if len(interp) >= 16:
                break

        # 去重：只保留最有意义的解读（优先 u32，相同偏移只保留最大类型）
        seen_offsets: dict[str, str] = {}
        deduped = []
        for item in interp:
            key = item["offset"]
            if key not in seen_offsets:
                seen_offsets[key] = item["type"]
                deduped.append(item)
            else:
                type_rank = {"u8": 0, "u16_le": 1, "u32_le": 2}
                if type_rank.get(item["type"], -1) > type_rank.get(seen_offsets[key], -1):
                    seen_offsets[key] = item["type"]
                    deduped = [x for x in deduped if x["offset"] != key]
                    deduped.append(item)

        result.append({
            "region_start": f"0x{rstart:x}",
            "region_end": f"0x{rend:x}",
            "changed_bytes": changed_len,
            "context_offset": f"0x{ctx_start:x}",
            "context_before": seg_a.hex(" "),
            "context_after": seg_b.hex(" "),
            "interpretations": deduped,
        })
    return result

@app.post("/api/san14/diff")
def san14_diff(body: San14DiffRequest):
    path_a = Path(body.file_a).expanduser()
    path_b = Path(body.file_b).expanduser()
    for p in (path_a, path_b):
        if not p.exists() or not p.is_file():
            return JSONResponse({"error": f"文件不存在：{p}"}, status_code=400)
        if p.suffix.lower() != ".s14":
            return JSONResponse({"error": "请选择 .s14 存档文件"}, status_code=400)
    try:
        raw_a = path_a.read_bytes()
        raw_b = path_b.read_bytes()
        dec_a, _ = _san14_lwc_decompress(raw_a)
        dec_b, _ = _san14_lwc_decompress(raw_b)
    except San14LwcError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"error": f"解压失败：{exc}"}, status_code=500)

    regions = _san14_diff_regions(dec_a, dec_b)
    return {
        "ok": True,
        "file_a": path_a.name,
        "file_b": path_b.name,
        "label": body.label,
        "decompressed_size_a": len(dec_a),
        "decompressed_size_b": len(dec_b),
        "region_count": len(regions),
        "regions": regions,
    }


class San14Patch(BaseModel):
    offset: int
    data_hex: str

class San14WriteRequest(BaseModel):
    file: str
    patches: list[San14Patch]
    dry_run: bool = False

@app.post("/api/san14/write")
def san14_write(body: San14WriteRequest):
    target = Path(body.file).expanduser()
    if not target.exists() or not target.is_file():
        return JSONResponse({"error": "存档文件不存在或无法访问"}, status_code=400)
    if target.suffix.lower() != ".s14":
        return JSONResponse({"error": "请选择 .s14 存档文件"}, status_code=400)

    try:
        raw = target.read_bytes()
        lwc_info = _san14_lwc_info(raw)
        if not lwc_info:
            return JSONResponse({"error": "未发现 LWC 压缩块"}, status_code=400)
        decompressed, _ = _san14_lwc_decompress(raw)
        patched = bytearray(decompressed)

        applied = []
        for patch in body.patches:
            try:
                patch_bytes = bytes.fromhex(patch.data_hex.replace(" ", ""))
            except ValueError:
                return JSONResponse({"error": f"patch data_hex 格式错误：{patch.data_hex}"}, status_code=400)
            end = patch.offset + len(patch_bytes)
            if end > len(patched):
                return JSONResponse({"error": f"patch 偏移越界：offset={patch.offset:#x} size={len(patch_bytes)}"}, status_code=400)
            original_hex = patched[patch.offset:end].hex(" ")
            patched[patch.offset:end] = patch_bytes
            applied.append({
                "offset": f"0x{patch.offset:x}",
                "before": original_hex,
                "after": patch_bytes.hex(" "),
            })

        original_table = raw[lwc_info["offset"] + 12:lwc_info["offset"] + 268]
        rebuilt_lwc = _san14_lwc_compress_literals(bytes(patched), original_table)
        rebuilt_raw = raw[:lwc_info["offset"]] + rebuilt_lwc

        check, _ = _san14_lwc_decompress(rebuilt_raw)
        if check != bytes(patched):
            return JSONResponse({"error": "重编码回校验失败，已取消写入"}, status_code=500)

        if body.dry_run:
            return {
                "ok": True,
                "dry_run": True,
                "patches_applied": applied,
                "new_size": len(rebuilt_raw),
                "note": "dry_run=true，未写入磁盘",
            }

        out_dir = target.parent / "new"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / target.name
        out_path.write_bytes(rebuilt_raw)
    except San14LwcError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"error": f"写入失败：{exc}"}, status_code=500)

    return {
        "ok": True,
        "dry_run": False,
        "path": str(out_path),
        "old_size": len(raw),
        "new_size": len(rebuilt_raw),
        "patches_applied": applied,
    }


# ── 城市资源定位 ──────────────────────────────────────────
# 已知命名城市的相对偏移（相对于寿春锚点）
_SAN14_CITY_NAMES_PATH = Path.home() / ".baibao" / "san14_city_names.json"
_SAN14_DEFAULT_CITY_NAMES: dict[int, str] = {
    -3 * 0xBC: "下邳",
    -2 * 0xBC: "小沛",
    0:          "寿春",
    1 * 0xBC:  "庐江",
}

def _san14_load_city_names() -> dict[int, str]:
    """读取用户自定义城市名映射（相对寿春锚点偏移 → 名称）。"""
    if _SAN14_CITY_NAMES_PATH.exists():
        try:
            raw = json.loads(_SAN14_CITY_NAMES_PATH.read_text())
            return {int(k): v for k, v in raw.items()}
        except Exception:
            pass
    return dict(_SAN14_DEFAULT_CITY_NAMES)

def _san14_save_city_names(mapping: dict[int, str]) -> None:
    _SAN14_CITY_NAMES_PATH.parent.mkdir(parents=True, exist_ok=True)
    _SAN14_CITY_NAMES_PATH.write_text(
        json.dumps({str(k): v for k, v in mapping.items()}, ensure_ascii=False, indent=2)
    )

def _san14_locate_cities(data: bytes, city_names: dict[int, str] | None = None) -> dict | None:
    """
    扫描解压数据，定位全部城市资源块。

    算法：
    1. 寻找锚点（寿春）：找 pre==0x06 且相邻4城（下邳/小沛/寿春/庐江）均合法的位置。
    2. 从锚点向后走到网格起点（首个全零记录前停止）。
    3. 从网格起点向前扫描所有步长为 0xBC 的槽位：
       - pre < 16（势力编号）且三项 u32 均 ≤ 9_999_999 → 城市记录
       - pre ≥ 16 → 数据区结束
       - 全零记录 → 空城，仍收录（pre==0 且三值为零）
    4. 返回所有收录记录，标记空城（empty=True）。
    """
    if city_names is None:
        city_names = _san14_load_city_names()

    STEP     = 0xBC
    MAX_RES  = 9_999_999
    MAX_PRE  = 15

    def read_vals(pos: int) -> tuple[int, int, int] | None:
        if pos < 1 or pos + 12 > len(data):
            return None
        a = int.from_bytes(data[pos:pos+4],   "little")
        b = int.from_bytes(data[pos+4:pos+8], "little")
        c = int.from_bytes(data[pos+8:pos+12], "little")
        if a > MAX_RES or b > MAX_RES or c > MAX_RES:
            return None
        return a, b, c

    # 步骤1：找锚点（寿春）
    # 条件：以锚点为0，(-3,-2,0,+1)×STEP 处的4城均有效，
    #        且4城的 pre 字节完全相同（同一势力的连续城市群）。
    # 这样可区分 寿春（4城 pre 均=6）和 下邳（pre 混合 4/6）。
    MIN_Z, MIN_J, MIN_S = 1000, 1000, 100
    anchor: int | None = None
    for i, byte in enumerate(data):
        if byte != 0x06:
            continue
        cand = i + 1
        ok = True
        pre_vals: list[int] = []
        for rel in (-3 * STEP, -2 * STEP, 0, STEP):
            pos = cand + rel
            vals = read_vals(pos)
            if vals is None:
                ok = False
                break
            pre = data[pos - 1]
            if pre > MAX_PRE:
                ok = False
                break
            z, j, s = vals
            if z < MIN_Z or j < MIN_J or s < MIN_S:
                ok = False
                break
            pre_vals.append(pre)
        # 4城 pre 字节必须全部相同（同一势力群）
        if ok and len(set(pre_vals)) != 1:
            ok = False
        if ok:
            anchor = cand
            break

    if anchor is None:
        return None

    # 步骤2：向后走找网格起点
    grid_start = anchor
    for n in range(1, 60):
        pos = anchor - n * STEP
        if pos < 1:
            break
        pre  = data[pos - 1]
        vals = read_vals(pos)
        if vals is None or pre > MAX_PRE:
            break
        z, j, s = vals
        if z == 0 and j == 0 and s == 0:
            break
        grid_start = pos

    # 步骤3：从网格起点向前全量扫描
    cities: list[dict] = []
    slot = 0
    for n in range(80):
        pos = grid_start + n * STEP
        if pos + 12 > len(data):
            break
        pre  = data[pos - 1] if pos > 0 else 0
        if pre > MAX_PRE:
            break
        vals = read_vals(pos)
        if vals is None:
            break
        z, j, s = vals
        slot += 1
        rel   = pos - anchor
        name  = city_names.get(rel, f"城市{slot}")
        empty = (z == 0 and j == 0 and s == 0)
        cities.append({
            "name":     name,
            "slot":     slot,
            "abs":      pos,
            "pre":      pre,
            "empty":    empty,
            "zijin":    z,
            "junliang": j,
            "shidao":   s,
        })

    if not cities:
        return None
    return {"anchor": anchor, "cities": cities}


@app.get("/api/san14/city-names")
def san14_get_city_names():
    """返回当前城市名映射（rel偏移 → 名称）。"""
    names = _san14_load_city_names()
    return {str(k): v for k, v in names.items()}

class San14CityNamesPatch(BaseModel):
    updates: dict[str, str]  # {str(rel): name}

@app.post("/api/san14/city-names")
def san14_update_city_names(body: San14CityNamesPatch):
    """更新并持久化城市名映射。"""
    current = _san14_load_city_names()
    for k, v in body.updates.items():
        rel = int(k)
        if v.strip():
            current[rel] = v.strip()
        elif rel in current:
            del current[rel]
    _san14_save_city_names(current)
    return {"ok": True, "count": len(current)}

@app.get("/api/san14/cities")
def san14_cities(file: str):
    target = Path(file).expanduser()
    if not target.exists():
        return JSONResponse({"error": "文件不存在"}, status_code=400)
    try:
        raw = target.read_bytes()
        decompressed, _ = _san14_lwc_decompress(raw)
    except San14LwcError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)

    result = _san14_locate_cities(decompressed)
    if result is None:
        return JSONResponse({"error": "未找到城市资源块，请确认是进度存档"}, status_code=400)
    return result


@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request):
    return templates.TemplateResponse("settings.html", {
        "request": request,
        "qwen_key": get_setting("qwen_api_key"),
        "qwen_model": get_setting("qwen_audio_model", "qwen-audio-turbo"),
        "deepseek_key": get_setting("deepseek_api_key"),
        "wechat_provider": _wechat_selected_provider(),
        "wechat_qwen_model": get_setting("wechat_qwen_model", "qwen-plus"),
        "wechat_deepseek_model": get_setting("wechat_deepseek_model", "deepseek-v4-flash"),
    })

@app.post("/settings")
async def save_settings(
    qwen_api_key:    str = Form(""),
    qwen_audio_model: str = Form("qwen-audio-turbo"),
    deepseek_api_key: str = Form(""),
    wechat_llm_provider: str = Form("deepseek"),
    wechat_qwen_model: str = Form("qwen-plus"),
    wechat_deepseek_model: str = Form("deepseek-v4-flash"),
):
    set_setting("qwen_api_key", qwen_api_key)
    set_setting("qwen_audio_model", qwen_audio_model)
    set_setting("deepseek_api_key", deepseek_api_key)
    set_setting("wechat_llm_provider", _wechat_selected_provider(wechat_llm_provider))
    set_setting("wechat_qwen_model", wechat_qwen_model)
    set_setting("wechat_deepseek_model", wechat_deepseek_model)
    return JSONResponse({"ok": True})

@app.get("/stock", response_class=HTMLResponse)
def stock_page(request: Request):
    import re as _re
    reports = sorted(REPORT_DIR.glob("*_report.md"), reverse=True) if REPORT_DIR.exists() else []
    dates   = [r.stem.replace("_report", "") for r in reports if _re.match(r"\d{4}-\d{2}-\d{2}_report", r.name)]
    return templates.TemplateResponse("stock.html", {
        "request": request,
        "dates": dates,
        "embedded": request.query_params.get("embedded") == "1",
    })

@app.get("/audio", response_class=HTMLResponse)
def audio_page(request: Request):
    return templates.TemplateResponse("audio.html", {"request": request})

@app.get("/chart", response_class=HTMLResponse)
def chart_page(request: Request):
    conn   = get_db()
    charts = [dict(r) for r in conn.execute("SELECT id,title,updated_at FROM charts ORDER BY updated_at DESC").fetchall()]
    conn.close()
    return templates.TemplateResponse("chart.html", {"request": request, "charts": charts})

@app.get("/quant-params", response_class=HTMLResponse)
def quant_params_page(request: Request):
    return templates.TemplateResponse("quant_params.html", {
        "request": request,
        "embedded": request.query_params.get("embedded") == "1",
    })


# ── 股票分析 API ──────────────────────────────────────────

QUANT_DIR  = Path.home() / "project" / "quant_trading"
QUANT_VENV = QUANT_DIR / ".venv" / "bin" / "python3.14"
REFRESH_TIMEOUT_SECONDS = 600
REFRESH_LOG_PATH = BASE_DIR / ".stock_refresh.log"

def _do_refresh():
    """后台执行股神计划，更新全局状态"""
    global refresh_task_status
    refresh_task_status = {"running": True, "progress": "启动分析引擎…", "done": False, "error": None, "dates": [], "latest": None}
    output_lines: list[str] = []

    def finish(error: str | None = None):
        reports = sorted(REPORT_DIR.glob("*_report.md"), reverse=True) if REPORT_DIR.exists() else []
        dates = [r.stem.replace("_report", "") for r in reports if re.match(r"\d{4}-\d{2}-\d{2}_report", r.name)]
        refresh_task_status.update({
            "running": False,
            "progress": "",
            "done": True,
            "error": error,
            "dates": dates,
            "latest": dates[0] if dates else None,
        })

    def write_refresh_log():
        try:
            REFRESH_LOG_PATH.write_text("\n".join(output_lines[-300:]), encoding="utf-8")
        except Exception:
            pass

    try:
        env = os.environ.copy()
        env["QUANT_MAX_WORKERS"] = "1"
        proc = subprocess.Popen(
            [str(QUANT_VENV), "-u", "main.py"],
            cwd=str(QUANT_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            errors="replace",
            bufsize=1,
            env=env,
        )

        def read_output():
            if not proc.stdout:
                return
            for raw in proc.stdout:
                line = re.sub(r"\x1b\[[0-9;]*m", "", raw).strip()
                if not line:
                    continue
                output_lines.append(line)
                if len(output_lines) > 300:
                    del output_lines[:-300]
                refresh_task_status["progress"] = line[-180:]

        reader = threading.Thread(target=read_output, daemon=True)
        reader.start()
        try:
            returncode = proc.wait(timeout=REFRESH_TIMEOUT_SECONDS)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            reader.join(timeout=2)
            write_refresh_log()
            latest = output_lines[-1] if output_lines else "启动分析引擎…"
            finish(f"分析超时（>{REFRESH_TIMEOUT_SECONDS // 60}分钟）。最近进度：{latest}")
            return

        reader.join(timeout=2)
        write_refresh_log()
        if returncode != 0:
            fatal = next((line for line in output_lines if "FATAL:" in line or "Traceback" in line), "")
            tail = "\n".join(output_lines[-12:])
            error_text = "\n".join(part for part in (fatal, tail) if part) or "运行失败"
            finish(error_text[-800:])
            return
        finish()
    except Exception as e:
        finish(str(e))

@app.post("/api/stock/refresh")
async def start_refresh():
    """启动后台刷新（非阻塞）"""
    global refresh_task_status
    if refresh_task_status["running"]:
        return JSONResponse({"error": "已有刷新任务在运行，请稍后"}, status_code=409)
    t = threading.Thread(target=_do_refresh, daemon=True)
    t.start()
    return {"ok": True, "message": "后台刷新已启动"}

@app.get("/api/stock/refresh/status")
def get_refresh_status():
    """查询刷新进度"""
    return refresh_task_status


analyze_task_status: dict = {"running": False, "code": None, "name": None, "done": False, "error": None, "result": None}

def _do_analyze_single(code: str, name: str):
    global analyze_task_status
    analyze_task_status = {"running": True, "code": code, "name": name, "done": False, "error": None, "result": None}
    try:
        result = subprocess.run(
            [str(QUANT_VENV), "main.py", "--stock", code],
            cwd=str(QUANT_DIR),
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            analyze_task_status.update({"running": False, "done": True, "error": result.stderr[-500:] or "分析失败"})
            return
        # 从最新 JSON 报告里取该股票的结果
        json_files = sorted(REPORT_DIR.glob("*_report.json"), reverse=True) if REPORT_DIR.exists() else []
        stock_data = None
        if json_files:
            with open(json_files[0], "r", encoding="utf-8") as f:
                stocks = json.load(f)
            stock_data = next((s for s in stocks if s.get("code") == code), None)
        analyze_task_status.update({"running": False, "done": True, "result": stock_data})
    except subprocess.TimeoutExpired:
        analyze_task_status.update({"running": False, "done": True, "error": "分析超时（>120s）"})
    except Exception as e:
        analyze_task_status.update({"running": False, "done": True, "error": str(e)})

@app.post("/api/stock/analyze")
async def start_analyze_single(request: Request):
    global analyze_task_status
    if analyze_task_status["running"]:
        return JSONResponse({"error": "已有分析任务在运行，请稍后"}, status_code=409)
    body = await request.json()
    code = body.get("code", "").strip()
    name = body.get("name", code)
    if not code:
        return JSONResponse({"error": "code 不能为空"}, status_code=400)
    import threading
    threading.Thread(target=_do_analyze_single, args=(code, name), daemon=True).start()
    return {"ok": True, "message": f"开始分析 {name}({code})"}

@app.get("/api/stock/analyze/status")
def get_analyze_status():
    return analyze_task_status


# ── 历史数据下载 ──────────────────────────────────────────────────────────────
download_task_status: dict = {
    "running": False, "done": False, "error": None,
    "current_code": None, "current_name": None,
    "progress": [],          # [{code, name, msg}] 滚动日志
    "results": [],           # 最终每只股票的结果
    "summary": [],           # hist_daily 数据库统计（各股行数范围）
}

def _do_download_history(cmd_args: list):
    """
    用 quant_trading venv 的 Python 运行 download_history.py，
    逐行读取 JSON 进度输出，实时写入 download_task_status。
    与 _do_analyze_single 的模式一致，避免 investment_hub 直接依赖 quant_trading 的包。
    """
    global download_task_status
    download_task_status = {
        "running": True, "done": False, "error": None,
        "current_code": None, "progress": [], "results": [], "summary": [],
    }
    try:
        import subprocess
        proc = subprocess.Popen(
            [str(QUANT_VENV), "download_history.py"] + cmd_args,
            cwd=str(QUANT_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        for raw_line in proc.stdout:
            line = raw_line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue

            if obj.get("__RESULT__"):
                # 最后一行：汇总结果
                download_task_status["results"] = obj.get("results", [])
                download_task_status["summary"] = obj.get("summary", [])
            else:
                # 进度行
                download_task_status["current_code"] = obj.get("code")
                download_task_status["progress"].append(obj)
                if len(download_task_status["progress"]) > 200:
                    download_task_status["progress"] = download_task_status["progress"][-200:]

        proc.wait()
        if proc.returncode != 0:
            stderr = proc.stderr.read()[-400:]
            download_task_status.update({
                "running": False, "done": True,
                "error": f"进程退出码 {proc.returncode}：{stderr}",
            })
        else:
            download_task_status.update({"running": False, "done": True})
    except Exception as e:
        import traceback
        download_task_status.update({
            "running": False, "done": True,
            "error": str(e) + "\n" + traceback.format_exc()[-300:],
        })

@app.post("/api/stock/download-history")
async def start_download_history(request: Request):
    """触发历史数据增量下载（整个 watchlist）"""
    global download_task_status
    if download_task_status["running"]:
        return JSONResponse({"error": "下载任务正在进行，请稍后"}, status_code=409)
    stocks = _read_watchlist()
    if not stocks:
        return JSONResponse({"error": "监控列表为空"}, status_code=400)
    import threading
    threading.Thread(target=_do_download_history, args=([],), daemon=True).start()
    return {"ok": True, "message": f"开始下载 {len(stocks)} 只股票近3年历史数据"}

@app.post("/api/stock/download-history/single")
async def start_download_history_single(request: Request):
    """触发单只股票历史数据增量下载"""
    global download_task_status
    if download_task_status["running"]:
        return JSONResponse({"error": "下载任务正在进行，请稍后"}, status_code=409)
    body = await request.json()
    code = body.get("code", "").strip()
    name = body.get("name", code)
    if not code:
        return JSONResponse({"error": "code 不能为空"}, status_code=400)
    import threading
    threading.Thread(
        target=_do_download_history,
        args=(["--code", code, "--name", name],),
        daemon=True,
    ).start()
    return {"ok": True, "message": f"开始下载 {name}({code}) 近3年历史数据"}

@app.get("/api/stock/download-history/status")
def get_download_status():
    return download_task_status

# ── 回测 ──────────────────────────────────────────────────────────────────────

backtest_task_status: dict = {
    "running": False, "done": False, "error": None,
    "current_code": None, "progress": [], "results": [],
}

def _do_backtest(cmd_args: list):
    global backtest_task_status
    backtest_task_status = {
        "running": True, "done": False, "error": None,
        "current_code": None, "progress": [], "results": [],
    }
    try:
        import subprocess
        proc = subprocess.Popen(
            [str(QUANT_VENV), "backtest_web.py"] + cmd_args,
            cwd=str(QUANT_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        for raw_line in proc.stdout:
            line = raw_line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            if obj.get("__RESULT__"):
                backtest_task_status["results"] = obj.get("results", [])
            else:
                backtest_task_status["current_code"] = obj.get("code")
                backtest_task_status["progress"].append(obj)
                if len(backtest_task_status["progress"]) > 300:
                    backtest_task_status["progress"] = backtest_task_status["progress"][-300:]
        proc.wait()
        if proc.returncode != 0:
            stderr = proc.stderr.read()[-400:]
            backtest_task_status.update({"running": False, "done": True, "error": f"进程退出码 {proc.returncode}：{stderr}"})
        else:
            backtest_task_status.update({"running": False, "done": True})
    except Exception as e:
        import traceback
        backtest_task_status.update({"running": False, "done": True, "error": str(e) + "\n" + traceback.format_exc()[-300:]})

VALID_BACKTEST_DAYS = {0, 60, 180, 360, 720}

@app.post("/api/stock/backtest")
async def start_backtest(request: Request):
    global backtest_task_status
    if backtest_task_status["running"]:
        return JSONResponse({"error": "回测任务正在进行，请稍后"}, status_code=409)
    stocks = _read_watchlist()
    if not stocks:
        return JSONResponse({"error": "监控列表为空"}, status_code=400)
    body = await request.json()
    days = int(body.get("days", 0))
    if days not in VALID_BACKTEST_DAYS:
        return JSONResponse({"error": f"days 只能是 {sorted(VALID_BACKTEST_DAYS)}"}, status_code=400)
    cmd_args = ["--days", str(days)] if days > 0 else []
    import threading
    threading.Thread(target=_do_backtest, args=(cmd_args,), daemon=True).start()
    return {"ok": True, "message": f"开始回测 {len(stocks)} 只股票"}

@app.post("/api/stock/backtest/single")
async def start_backtest_single(request: Request):
    global backtest_task_status
    if backtest_task_status["running"]:
        return JSONResponse({"error": "回测任务正在进行，请稍后"}, status_code=409)
    body = await request.json()
    code = body.get("code", "").strip()
    if not code:
        return JSONResponse({"error": "code 不能为空"}, status_code=400)
    days = int(body.get("days", 0))
    if days not in VALID_BACKTEST_DAYS:
        return JSONResponse({"error": f"days 只能是 {sorted(VALID_BACKTEST_DAYS)}"}, status_code=400)
    cmd_args = ["--code", code] + (["--days", str(days)] if days > 0 else [])
    import threading
    threading.Thread(target=_do_backtest, args=(cmd_args,), daemon=True).start()
    return {"ok": True, "message": f"开始回测 {code}"}

@app.get("/api/stock/backtest/status")
def get_backtest_status():
    return backtest_task_status

@app.get("/api/stock/backtest/results")
def get_backtest_results():
    return {"results": backtest_task_status.get("results", [])}

# ── 每只股票因子权重覆盖 ───────────────────────────────────────────────────────

FACTOR_KEYS = ["technical", "fundamental", "money_flow", "sentiment", "chip"]
FACTOR_NAMES = {"technical": "技术面", "fundamental": "基本面",
                "money_flow": "资金面", "sentiment": "情绪面", "chip": "筹码面"}
DEFAULT_WEIGHTS = {"technical": 0.30, "fundamental": 0.20, "money_flow": 0.20, "sentiment": 0.15, "chip": 0.15}

@app.get("/api/stock-factors/{code}")
def get_stock_factors(code: str):
    """获取某只股票的因子权重（覆盖值 + 全局默认值）"""
    conn = get_db()
    rows = conn.execute(
        "SELECT factor_key, weight FROM stock_factor_overrides WHERE code=?", (code,)
    ).fetchall()
    conn.close()
    overrides = {r["factor_key"]: r["weight"] for r in rows}
    result = []
    for fk in FACTOR_KEYS:
        result.append({
            "factor_key":   fk,
            "factor_name":  FACTOR_NAMES[fk],
            "weight":       overrides.get(fk, DEFAULT_WEIGHTS[fk]),
            "is_override":  fk in overrides,
            "default":      DEFAULT_WEIGHTS[fk],
        })
    return result

@app.get("/api/investment-analysis/factor-weights")
def get_investment_analysis_factor_weights():
    """
    批量版 /api/stock-factors/{code}：一次性返回 watchlist 里每只股票的因子权重（覆盖值+默认值），
    供投资分析详情页做因子归因展示，避免逐只股票单独请求。
    """
    conn = get_db()
    rows = conn.execute("SELECT code, factor_key, weight FROM stock_factor_overrides").fetchall()
    conn.close()
    overrides = {}
    for r in rows:
        overrides.setdefault(r["code"], {})[r["factor_key"]] = r["weight"]
    watch_codes = [c for c, _n in _read_watchlist()]
    out = {}
    for code in watch_codes:
        ov = overrides.get(code, {})
        out[code] = [
            {
                "factor_key": fk,
                "factor_name": FACTOR_NAMES[fk],
                "weight": ov.get(fk, DEFAULT_WEIGHTS[fk]),
                "is_override": fk in ov,
            }
            for fk in FACTOR_KEYS
        ]
    return {"weights": out}

@app.put("/api/stock-factors/{code}")
async def update_stock_factors(code: str, request: Request):
    """批量更新某只股票的因子权重覆盖"""
    body = await request.json()  # {"technical": 0.4, "fundamental": 0.15, ...}
    weights = {k: float(v) for k, v in body.items() if k in FACTOR_KEYS}
    if not weights:
        return JSONResponse({"error": "无有效因子数据"}, status_code=400)
    total = sum(weights.values())
    if abs(total - 1.0) > 0.01:
        return JSONResponse({"error": f"权重之和必须等于1，当前为 {total:.2f}"}, status_code=400)
    now = datetime.now().isoformat()
    conn = get_db()
    try:
        for fk, w in weights.items():
            conn.execute(
                "INSERT INTO stock_factor_overrides(code,factor_key,weight,updated_at) VALUES(?,?,?,?) "
                "ON CONFLICT(code,factor_key) DO UPDATE SET weight=excluded.weight, updated_at=excluded.updated_at",
                (code, fk, w, now),
            )
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.delete("/api/stock-factors/{code}")
def reset_stock_factors(code: str):
    """重置某只股票的因子权重（删除所有覆盖，恢复全局默认）"""
    conn = get_db()
    try:
        conn.execute("DELETE FROM stock_factor_overrides WHERE code=?", (code,))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.get("/api/stock/download-history/summary")
def get_download_summary():
    """返回各股票在本地数据库中的数据量统计（直接读 SQLite，无需 quant 包）"""
    import sqlite3
    db_path = QUANT_DIR / "data" / "hist_daily.db"
    if not db_path.exists():
        return {"summary": []}
    try:
        conn = sqlite3.connect(str(db_path))
        rows = conn.execute(
            "SELECT code, COUNT(*) as cnt, MIN(date), MAX(date) FROM hist_daily GROUP BY code"
        ).fetchall()
        conn.close()
        return {"summary": [{"code": r[0], "count": r[1], "from": r[2], "to": r[3]} for r in rows]}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/stock/results")
def get_stock_results():
    """读取最新的股神计划分析结果 JSON（供前端展示用）"""
    json_files = sorted(REPORT_DIR.glob("*_report.json"), reverse=True) if REPORT_DIR.exists() else []
    if not json_files:
        return JSONResponse({"error": "暂无分析结果，请先点击「刷新分析」"}, status_code=404)
    latest = json_files[0]
    date_str = latest.stem.replace("_report", "")
    date_fmt = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    with open(latest, "r", encoding="utf-8") as f:
        stocks = json.load(f)
    import os
    updated_ts = int(os.path.getmtime(latest))
    return {"date": date_fmt, "stocks": stocks, "updated_at": updated_ts}

@app.get("/api/stock/results/history")
def get_stock_results_history(limit: int = 5):
    """读取最近多期股神计划分析结果，用于趋势和边际变化分析"""
    json_files = sorted(REPORT_DIR.glob("*_report.json"), reverse=True) if REPORT_DIR.exists() else []
    reports = []
    for path in json_files[:max(1, min(limit, 20))]:
        date_str = path.stem.replace("_report", "")
        date_fmt = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        try:
            with open(path, "r", encoding="utf-8") as f:
                stocks = json.load(f)
        except Exception:
            stocks = []
        reports.append({
            "date": date_fmt,
            "stocks": stocks,
            "updated_at": int(os.path.getmtime(path)),
        })
    return {"reports": reports}


@app.get("/api/investment-analysis/dimensions")
def get_investment_analysis_dimensions():
    """
    供给/需求/盈利三维评分——唯一权威计算来源（investment_snapshot.stock_dimensions）。
    前端 investment_analysis.js 和每日定时快照脚本都从这个接口/函数取数，避免两边各算一份导致结果漂移。
    """
    json_files = sorted(REPORT_DIR.glob("*_report.json"), reverse=True) if REPORT_DIR.exists() else []
    if not json_files:
        return {"date": None, "dimensions": {}}
    latest = json_files[0]
    date_str = latest.stem.replace("_report", "")
    date_fmt = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    with open(latest, "r", encoding="utf-8") as f:
        stocks = json.load(f)

    watchlist = [{"code": c, "name": n} for c, n in _read_watchlist()]
    stock_universe = watchlist + [{"code": s.get("code"), "name": s.get("name")} for s in stocks]
    exposures_by_code = _snap_build_exposures(_snap_load_industry_data(), stock_universe)

    dimensions = {}
    for stock in stocks:
        dims = _snap_stock_dimensions(stock, exposures_by_code)
        for key in _snap_code_keys(stock.get("code")):
            dimensions[key] = dims
    return {"date": date_fmt, "dimensions": dimensions}

@app.get("/api/investment-analysis/history")
def get_investment_analysis_history(limit: int = 30):
    """
    读取 stock_dimension_snapshots 的历史快照（launchd 每日任务写入），按股票代码分组返回，
    供前端画三维评分趋势。limit 是每只股票最多返回的天数。
    """
    conn = get_db()
    try:
        conn.execute(
            "CREATE TABLE IF NOT EXISTS stock_dimension_snapshots ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, snapshot_date TEXT NOT NULL, code TEXT NOT NULL, "
            "name TEXT, probability REAL, score REAL, price REAL, risk_label TEXT, "
            "supply REAL, demand REAL, profit REAL, spread REAL, divergent INTEGER, "
            "exposures_count INTEGER, created_at TEXT NOT NULL, UNIQUE(snapshot_date, code))"
        )
        rows = conn.execute(
            "SELECT snapshot_date, code, name, probability, price, supply, demand, profit, spread, divergent "
            "FROM stock_dimension_snapshots ORDER BY code, snapshot_date DESC"
        ).fetchall()
    finally:
        conn.close()

    by_code = {}
    for row in rows:
        bucket = by_code.setdefault(row["code"], [])
        if len(bucket) >= max(1, min(limit, 365)):
            continue
        bucket.append({
            "date": row["snapshot_date"],
            "probability": row["probability"],
            "price": row["price"],
            "supply": row["supply"],
            "demand": row["demand"],
            "profit": row["profit"],
            "spread": row["spread"],
            "divergent": bool(row["divergent"]),
        })
    for code in by_code:
        by_code[code].reverse()  # 按日期升序，方便前端画趋势线
    return {"history": by_code}

@app.get("/api/investment-analysis/notes")
def list_decision_notes(code: str = ""):
    """决策复盘记录：当初为什么关注这只股票，以及后续是否验证。"""
    conn = get_db()
    if code:
        rows = conn.execute(
            "SELECT * FROM stock_decision_notes WHERE code=? ORDER BY created_at DESC", (code,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM stock_decision_notes ORDER BY created_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/investment-analysis/notes")
async def create_decision_note(request: Request):
    body = await request.json()
    code = str(body.get("code", "")).strip()
    name = str(body.get("name", "")).strip()
    note = str(body.get("note", "")).strip()
    target_date = str(body.get("target_date") or "").strip() or None
    if not code or not note:
        return JSONResponse({"error": "code 和 note 不能为空"}, status_code=400)
    conn = get_db()
    now = datetime.now().isoformat()
    cur = conn.execute(
        "INSERT INTO stock_decision_notes(code,name,note,target_date,created_at,resolved) VALUES (?,?,?,?,?,0)",
        (code, name, note, target_date, now),
    )
    conn.commit()
    note_id = cur.lastrowid
    conn.close()
    return {"ok": True, "id": note_id}

@app.patch("/api/investment-analysis/notes/{note_id}")
async def resolve_decision_note(note_id: int, request: Request):
    body = await request.json()
    verdict = str(body.get("verdict", "")).strip()
    resolved_note = str(body.get("resolved_note") or "").strip() or None
    if verdict not in ("兑现", "部分兑现", "未兑现"):
        return JSONResponse({"error": "verdict 必须是 兑现/部分兑现/未兑现"}, status_code=400)
    conn = get_db()
    now = datetime.now().isoformat()
    conn.execute(
        "UPDATE stock_decision_notes SET resolved=1, verdict=?, resolved_note=?, resolved_at=? WHERE id=?",
        (verdict, resolved_note, now, note_id),
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/investment-analysis/notes/{note_id}")
def delete_decision_note(note_id: int):
    conn = get_db()
    conn.execute("DELETE FROM stock_decision_notes WHERE id=?", (note_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/investment-analysis/factor-backtest")
def get_investment_analysis_factor_backtest():
    """
    读取 quant_trading/backtest_all_stocks.py 的回测结果（launchd 每日16:10自动跑），
    每只股票在“当前权重配置”和历史真实K线回测出的“最优权重方案”之间的准确率对比。
    这是权重方案级别的回测，不是单因子独立剥离验证——见 investment_analysis.js 里的说明。
    """
    csv_path = QUANT_DIR / "reports" / "all_stocks_weight_optimization.csv"
    if not csv_path.exists():
        return {"generated_at": None, "backtest": {}}
    import csv as csv_module
    out = {}
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        for row in csv_module.DictReader(f):
            out[row["code"]] = {
                "days": int(float(row["days"])),
                "volatility": float(row["volatility"]),
                "current_accuracy": float(row["current_accuracy"]),
                "best_scheme": row["best_scheme"],
                "best_accuracy": float(row["best_accuracy"]),
                "improvement": float(row["improvement"]),
                "rec_weights": {
                    "technical": float(row["rec_tech"]),
                    "fundamental": float(row["rec_fund"]),
                    "money_flow": float(row["rec_money"]),
                    "sentiment": float(row["rec_sentiment"]),
                    "chip": float(row["rec_chip"]),
                },
            }
    generated_at = int(os.path.getmtime(csv_path))
    return {"generated_at": generated_at, "backtest": out}

@app.get("/api/investment-analysis/kline")
def get_investment_analysis_kline(code: str, days: int = 120):
    """
    读取 quant_trading/data/hist_daily.db 的日线 OHLCV（下载历史数据模块产出，已存3年），
    供投资分析详情页画 K 线。days 控制返回最近多少个交易日。
    """
    db_path = QUANT_DIR / "data" / "hist_daily.db"
    if not db_path.exists():
        return {"code": code, "kline": []}
    # hist_daily 里代码可能带/不带后缀，用纯数字键匹配
    digits = re.sub(r"\D", "", code)
    candidates = [code, digits] if digits else [code]
    limit = max(5, min(days, 1000))
    conn = sqlite3.connect(str(db_path))
    try:
        rows = []
        for cand in candidates:
            rows = conn.execute(
                "SELECT date, open, high, low, close, volume FROM hist_daily "
                "WHERE code=? ORDER BY date DESC LIMIT ?",
                (cand, limit),
            ).fetchall()
            if rows:
                break
    finally:
        conn.close()
    kline = [
        {"date": r[0], "open": r[1], "high": r[2], "low": r[3], "close": r[4], "volume": r[5]}
        for r in reversed(rows)
    ]
    return {"code": code, "kline": kline}

@app.get("/api/investment-analysis/thresholds")
def get_investment_analysis_thresholds():
    """
    信号/决策阈值的唯一权威来源（config/investment_thresholds.json）。
    前端 investment_analysis.js 从这里取，Python(investment_snapshot) 直接读同一文件，
    避免两端各自硬编码 55%/40%/背离阈值/危险关键词导致页面徽章与推送不一致。
    """
    return _snap_load_thresholds()

@app.get("/api/investment-analysis/alerts")
def get_investment_analysis_alerts(limit: int = 50):
    """读取 stock_alerts 表（investment_snapshot 每日快照落库），跨股票的近期预警历史。"""
    conn = get_db()
    try:
        conn.execute(
            "CREATE TABLE IF NOT EXISTS stock_alerts ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, alert_date TEXT NOT NULL, code TEXT NOT NULL, "
            "name TEXT, alert_type TEXT, message TEXT NOT NULL, created_at TEXT NOT NULL)"
        )
        rows = conn.execute(
            "SELECT alert_date, code, name, alert_type, message, created_at "
            "FROM stock_alerts ORDER BY alert_date DESC, id DESC LIMIT ?",
            (max(1, min(limit, 500)),),
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]

@app.get("/api/stock/report")
def get_stock_report(date: str):
    path = REPORT_DIR / f"{date}_report.md"
    if not path.exists():
        return JSONResponse({"error": "报告不存在"}, status_code=404)
    md   = path.read_text(encoding="utf-8")
    return JSONResponse({"date": date, "stocks": _parse_report(md)})

def _parse_report(md: str) -> list:
    stocks = []
    # 先从排行榜表格提取 reason（名称 → reason 映射）
    reason_map = {}
    for rm in re.finditer(
        r"\|\s*\d+\s*\|\s*\*\*(.+?)\*\*\([^)]+\)\s*\|\s*[🟢🟡🔴][^|]+\|[^|]+\|\s*(.+?)\s*\|", md
    ):
        reason_map[rm.group(1).strip()] = rm.group(2).strip()

    # 按股票分段（### N. 名称 (代码) emoji）
    blocks = re.split(r"\n(?=### \d+\.)", md)
    for block in blocks:
        m = re.match(r"### \d+\.\s+(.+?)\s+\((\d+)\)\s*([🟢🟡🔴]?)", block)
        if not m:
            continue
        name, code, emoji = m.group(1), m.group(2), m.group(3)

        prob_m   = re.search(r"明日上涨概率.*?\*\*(\d+\.?\d*)%\*\*", block)
        price_m  = re.search(r"收盘价\*\*:\s*([\d.]+)元", block)
        change_m = re.search(r"涨跌\*\*:\s*([▲▼])([\d.]+)%", block)
        prob     = float(prob_m.group(1))  if prob_m  else 0
        price    = float(price_m.group(1)) if price_m else 0
        change_dir = change_m.group(1)    if change_m else ""
        change_val = float(change_m.group(2)) if change_m else 0
        change_pct = change_val if change_dir == "▲" else -change_val

        # 指标表格
        indicators = {}
        for row in re.finditer(r"\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|", block):
            key, val, sig = row.group(1), row.group(2), row.group(3)
            if key not in ("指标", "---", "数值"):
                indicators[key] = {"value": val.strip(), "signal": sig.strip()}

        # 综合评分
        scores = {}
        for sm in re.finditer(r"- (.+?): [★☆]+\s*\((\d+)\)", block):
            scores[sm.group(1)] = int(sm.group(2))

        # 技术信号
        sig_m  = re.search(r"\*\*技术信号\*\*:\s*(.+)", block)
        signals = [s.strip() for s in sig_m.group(1).split(",")] if sig_m else []

        # 筹码
        chip = {}
        for cm in re.finditer(r"\|\s*(.+?)\s*\|\s*(.+?)\s*\|", block):
            if cm.group(1) in ("70%筹码区间","筹码宽度","获利比例","15日收敛趋势","筹码信号"):
                chip[cm.group(1)] = cm.group(2).strip()
        bar_m = re.search(r"> 15日宽度变化趋势：(.+)", block)
        chip["trend_bar"] = bar_m.group(1).strip() if bar_m else ""

        stocks.append({
            "name": name, "code": code, "emoji": emoji,
            "probability": prob, "price": price, "change_pct": change_pct,
            "indicators": indicators, "scores": scores,
            "signals": [s for s in signals if s and s != "无明显信号"],
            "chip": chip,
            "reason": reason_map.get(name, ""),
        })
    return stocks


# ── Watchlist 管理 API ───────────────────────────────────

WATCHLIST_PATH = Path.home() / "project" / "quant_trading" / "config" / "watchlist.json"

def _read_watchlist() -> list:
    if WATCHLIST_PATH.exists():
        with open(WATCHLIST_PATH, "r", encoding="utf-8") as f:
            return json.load(f).get("stocks", [])
    return []

def _write_watchlist(stocks: list):
    data = {"stocks": stocks}
    WATCHLIST_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(WATCHLIST_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _stock_code_keys(code: str) -> set[str]:
    raw = str(code or "").strip()
    if not raw:
        return set()
    digits = re.sub(r"\D", "", raw)
    return {v for v in {
        raw,
        raw.upper(),
        re.sub(r"^HK:", "", raw, flags=re.I),
        raw.split(".")[0],
        digits,
    } if v}

def _stock_search_matches(query: str, code: str, name: str) -> bool:
    needle = str(query or "").strip().lower()
    if not needle:
        return False
    return needle in f"{code} {name}".lower()

def _append_stock_search_hit(hits: list, seen: set, code, name, query: str = ""):
    code = str(code or "").strip()
    name = str(name or "").replace(" ", "").strip()
    if not code or not name or code in seen:
        return
    if query and not _stock_search_matches(query, code, name):
        return
    seen.add(code)
    hits.append({"code": code, "name": name})

def _search_stock_from_quant(query: str) -> list:
    """优先复用 quant_trading 的 akshare 搜索；失败时交给后续兜底。"""
    import sys
    quant_src = Path.home() / "project" / "quant_trading" / "src"
    if str(quant_src) not in sys.path:
        sys.path.insert(0, str(quant_src))
    try:
        from data_collector import StockDataCollector
        return StockDataCollector.search_stock(query) or []
    except Exception:
        return []

def _search_stock_from_eastmoney(query: str) -> list:
    """不依赖 pandas/akshare 的轻量在线搜索兜底。"""
    from urllib.parse import urlencode
    from urllib.request import Request, urlopen

    params = {
        "input": query,
        "type": "14",
        "token": "44c9d251add88e27b65ed86506f6e5da",
        "count": "20",
    }
    url = "http://searchapi.eastmoney.com/api/suggest/get?" + urlencode(params)
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(req, timeout=5) as resp:
            payload = json.loads(resp.read().decode("utf-8", "ignore"))
    except Exception:
        return []

    rows = ((payload.get("QuotationCodeTable") or {}).get("Data") or [])
    hits, seen = [], set()
    for row in rows:
        code = str(row.get("Code") or row.get("UnifiedCode") or "").strip()
        name = str(row.get("Name") or "").replace(" ", "").strip()
        if not re.fullmatch(r"\d{6}", code):
            continue
        _append_stock_search_hit(hits, seen, code, name)
    return hits

def _search_stock_from_local_context(query: str) -> list:
    """本地兜底：从监控列表、分析报告和事件表中找已有股票。"""
    hits, seen = [], set()

    for code, name in _read_watchlist():
        _append_stock_search_hit(hits, seen, code, name, query)

    json_files = sorted(REPORT_DIR.glob("*_report.json"), reverse=True) if REPORT_DIR.exists() else []
    for path in json_files[:20]:
        try:
            with open(path, "r", encoding="utf-8") as f:
                stocks = json.load(f)
        except Exception:
            continue
        if not isinstance(stocks, list):
            continue
        for stock in stocks:
            if isinstance(stock, dict):
                _append_stock_search_hit(hits, seen, stock.get("code"), stock.get("name"), query)

    try:
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT stock_code, stock_name FROM stock_events WHERE stock_code LIKE ? OR stock_name LIKE ? LIMIT 20",
            (f"%{query}%", f"%{query}%")
        ).fetchall()
        conn.close()
        for row in rows:
            _append_stock_search_hit(hits, seen, row["stock_code"], row["stock_name"], query)
    except Exception:
        pass

    return hits

@app.get("/api/watchlist")
def get_watchlist():
    return _read_watchlist()

@app.post("/api/watchlist")
async def add_watchlist(request: Request):
    body = await request.json()
    code = body.get("code", "").strip()
    name = body.get("name", "").strip()
    if not code or not name:
        return JSONResponse({"error": "code 和 name 不能为空"}, status_code=400)
    stocks = _read_watchlist()
    code_keys = _stock_code_keys(code)
    if any(_stock_code_keys(s[0]) & code_keys for s in stocks):
        return JSONResponse({"error": f"{code} 已在监控列表中"}, status_code=409)
    stocks.append([code, name])
    _write_watchlist(stocks)
    return {"ok": True, "stocks": stocks}

@app.delete("/api/watchlist/{code}")
def remove_watchlist(code: str):
    stocks = _read_watchlist()
    code_keys = _stock_code_keys(code)
    new_stocks = [s for s in stocks if not (_stock_code_keys(s[0]) & code_keys)]
    if len(new_stocks) == len(stocks):
        return JSONResponse({"error": f"{code} 不在监控列表中"}, status_code=404)
    _write_watchlist(new_stocks)
    return {"ok": True, "stocks": new_stocks}

# ── 重要时间节点 页面 ─────────────────────────────────────
@app.get("/stock-events", response_class=HTMLResponse)
def stock_events_page(request: Request):
    return templates.TemplateResponse("stock_events.html", {
        "request": request,
        "embedded": request.query_params.get("embedded") == "1",
    })

# ── 重要时间节点 API ──────────────────────────────────────

@app.get("/api/stock-events")
def list_stock_events(code: str = "", status: str = "", type: str = ""):
    conn = get_db()
    sql = "SELECT * FROM stock_events WHERE 1=1"
    params: list = []
    if code:
        sql += " AND stock_code=?"
        params.append(code)
    if status:
        sql += " AND status=?"
        params.append(status)
    if type:
        sql += " AND event_type=?"
        params.append(type)
    sql += " ORDER BY event_date ASC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/stock-events")
def create_stock_event(data: dict):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO stock_events(stock_code,stock_name,event_date,event_title,event_type,event_desc,status,importance,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (data["stock_code"], data.get("stock_name",""), data["event_date"],
             data["event_title"], data.get("event_type","其他"),
             data.get("event_desc",""), data.get("status","pending"),
             data.get("importance","normal"), now, now)
        )
        conn.commit()
        return {"ok": True, "id": cur.lastrowid}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.put("/api/stock-events/{eid}")
def update_stock_event(eid: int, data: dict):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    conn.execute(
        "UPDATE stock_events SET stock_code=?,stock_name=?,event_date=?,event_title=?,event_type=?,event_desc=?,status=?,importance=?,updated_at=? WHERE id=?",
        (data["stock_code"], data.get("stock_name",""), data["event_date"],
         data["event_title"], data.get("event_type","其他"),
         data.get("event_desc",""), data.get("status","pending"),
         data.get("importance","normal"), now, eid)
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.patch("/api/stock-events/{eid}/status")
def update_event_status(eid: int, data: dict):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    conn.execute("UPDATE stock_events SET status=?,updated_at=? WHERE id=?",
                 (data["status"], now, eid))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/stock-events/{eid}")
def delete_stock_event(eid: int):
    conn = get_db()
    conn.execute("DELETE FROM stock_events WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/stock-events/import")
async def import_stock_events(request: Request):
    import io, csv
    body = await request.body()
    text = body.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    count = 0
    for row in reader:
        code  = (row.get("stock_code") or "").strip()
        date  = (row.get("event_date") or "").strip()
        title = (row.get("event_title") or "").strip()
        if not code or not date or not title:
            continue
        conn.execute(
            "INSERT INTO stock_events(stock_code,stock_name,event_date,event_title,event_type,event_desc,status,importance,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (code, (row.get("stock_name") or "").strip(), date, title,
             (row.get("event_type") or "其他").strip(),
             (row.get("event_desc") or "").strip(),
             (row.get("status") or "pending").strip(),
             (row.get("importance") or "normal").strip(), now, now)
        )
        count += 1
    conn.commit()
    conn.close()
    return {"ok": True, "count": count}

@app.get("/api/stock-events/export")
def export_stock_events():
    conn = get_db()
    rows = conn.execute("SELECT * FROM stock_events ORDER BY event_date").fetchall()
    conn.close()
    header = "stock_code,stock_name,event_date,event_title,event_type,event_desc,status,importance\n"
    lines = []
    for r in rows:
        d = dict(r)
        def q(v): return f'"{v}"' if v and (',' in str(v) or '"' in str(v)) else (v or "")
        lines.append(f"{q(d['stock_code'])},{q(d['stock_name'])},{q(d['event_date'])},{q(d['event_title'])},{q(d['event_type'])},{q(d['event_desc'])},{q(d['status'])},{q(d['importance'])}")
    return _csv_response(header + "\n".join(lines), f"重要时间节点_{datetime.now().strftime('%Y%m%d')}.csv")

@app.get("/api/stock-events/template")
def get_events_template():
    content = "stock_code,stock_name,event_date,event_title,event_type,event_desc,status,importance\n"
    content += "300244,迪安诊断,2026-07-15,半年报披露,业绩,预计7月披露半年报,pending,high\n"
    content += "603881,数据港,2026-08-01,机构调研,调研,,pending,normal\n"
    return _csv_response(content, "重要时间节点导入模板.csv")

@app.get("/api/stock/search")
def search_stock(q: str = ""):
    """股票搜索：精准代码匹配 或 名称模糊搜索"""
    query = q.strip()
    if not query:
        return []

    hits, seen = [], set()
    for source in (
        _search_stock_from_quant,
        _search_stock_from_eastmoney,
        _search_stock_from_local_context,
    ):
        for item in source(query):
            _append_stock_search_hit(hits, seen, item.get("code"), item.get("name"))
        if hits:
            return hits[:20]
    return []


# ── 录音转会议纪要 API ────────────────────────────────────

@app.post("/api/audio/transcribe")
async def transcribe_audio(file: UploadFile = File(...)):
    api_key = get_setting("qwen_api_key")
    if not api_key:
        return JSONResponse({"error": "请先在设置页配置 Qwen API Key"}, status_code=400)

    # 保存上传文件
    suffix  = Path(file.filename).suffix
    save_path = UPLOAD_DIR / f"audio_{datetime.now().strftime('%Y%m%d%H%M%S')}{suffix}"
    content = await file.read()
    save_path.write_bytes(content)

    try:
        import base64
        audio_b64  = base64.b64encode(content).decode()
        audio_url  = f"data:audio/mpeg;base64,{audio_b64}"
        model      = get_setting("qwen_audio_model", "qwen-audio-turbo")

        import urllib.request, json as _json
        payload = _json.dumps({
            "model": model,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "audio_url", "audio_url": {"url": audio_url}},
                    {"type": "text", "text": (
                        "请将这段录音转写并整理成会议纪要，格式如下：\n"
                        "## 会议纪要\n"
                        "**日期**：（从内容推断，无法确认请留空）\n"
                        "**参会人**：（从内容推断）\n"
                        "**主题**：（一句话概括）\n\n"
                        "### 一、主要讨论内容\n"
                        "（分条列出关键议题和讨论结果）\n\n"
                        "### 二、决议事项\n"
                        "（列出明确的决定和行动项，注明负责人和时间节点）\n\n"
                        "### 三、待跟进事项\n"
                        "（列出尚未确定、需要后续跟进的内容）"
                    )}
                ]
            }]
        }).encode()

        req = urllib.request.Request(
            "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions",
            data=payload,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            }
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = _json.loads(resp.read())

        minutes = result["choices"][0]["message"]["content"]
        return JSONResponse({"minutes": minutes, "filename": file.filename})

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        save_path.unlink(missing_ok=True)


# ── 代理网关 ─────────────────────────────────────────────

@app.get("/proxy", response_class=HTMLResponse)
def proxy_page(request: Request):
    return templates.TemplateResponse("proxy.html", {"request": request})

def _get_proxy_status() -> dict:
    """读取当前系统代理状态，返回 {enabled, host, port}"""
    try:
        out = subprocess.check_output(
            ["networksetup", "-getwebproxy", PROXY_IFACE],
            text=True, timeout=5
        )
        enabled = "Enabled: Yes" in out
        return {"enabled": enabled, "host": PROXY_HOST, "port": PROXY_PORT, "iface": PROXY_IFACE}
    except Exception as e:
        return {"enabled": False, "host": PROXY_HOST, "port": PROXY_PORT, "iface": PROXY_IFACE, "error": str(e)}

@app.get("/api/proxy/status")
def proxy_status():
    return JSONResponse(_get_proxy_status())

@app.post("/api/proxy/toggle")
def proxy_toggle():
    status = _get_proxy_status()
    target = "off" if status["enabled"] else "on"
    try:
        for cmd in [
            ["networksetup", "-setwebproxystate",          PROXY_IFACE, target],
            ["networksetup", "-setsecurewebproxystate",    PROXY_IFACE, target],
            ["networksetup", "-setsocksfirewallproxystate",PROXY_IFACE, target],
        ]:
            subprocess.run(cmd, check=True, timeout=5)
        return JSONResponse({"enabled": target == "on", "host": PROXY_HOST, "port": PROXY_PORT})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# ── 一图一表 API ──────────────────────────────────────────

class ChartData(BaseModel):
    title:     str
    data_json: str  # JSON string of nodes+edges

@app.get("/api/charts")
def list_charts():
    conn   = get_db()
    charts = [dict(r) for r in conn.execute("SELECT id,title,updated_at FROM charts ORDER BY updated_at DESC").fetchall()]
    conn.close()
    return charts

@app.get("/api/charts/{chart_id}")
def get_chart(chart_id: int):
    conn = get_db()
    row  = conn.execute("SELECT * FROM charts WHERE id=?", (chart_id,)).fetchone()
    conn.close()
    if not row:
        return JSONResponse({"error": "not found"}, status_code=404)
    return dict(row)

@app.post("/api/charts")
def create_chart(data: ChartData):
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    cur  = conn.execute(
        "INSERT INTO charts(title,data_json,created_at,updated_at) VALUES(?,?,?,?)",
        (data.title, data.data_json, now, now)
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return {"id": new_id, "title": data.title, "updated_at": now}

@app.put("/api/charts/{chart_id}")
def update_chart(chart_id: int, data: ChartData):
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    conn.execute(
        "UPDATE charts SET title=?, data_json=?, updated_at=? WHERE id=?",
        (data.title, data.data_json, now, chart_id)
    )
    conn.commit()
    conn.close()
    return {"id": chart_id, "title": data.title, "updated_at": now}

@app.delete("/api/charts/{chart_id}")
def delete_chart(chart_id: int):
    conn = get_db()
    conn.execute("DELETE FROM charts WHERE id=?", (chart_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── 任务 API ──────────────────────────────────────────────

class TaskBase(BaseModel):
    title:        str
    note:         Optional[str] = None
    is_recurring: int
    task_date:    str

class Task(TaskBase):
    id:      int
    status:  str
    done_at: Optional[str] = None

@app.get("/api/tasks", response_model=list[Task])
def read_tasks(date: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT id,title,note,is_recurring,task_date,status,done_at FROM tasks WHERE task_date=?", (date,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/tasks/dates")
def read_task_dates():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT task_date FROM tasks ORDER BY task_date DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/tasks", response_model=Task)
def create_task(task: TaskBase):
    conn = get_db()
    cur  = conn.execute(
        "INSERT INTO tasks(title,note,is_recurring,task_date,status,done_at) VALUES(?,?,?,?,?,?)",
        (task.title, task.note, task.is_recurring, task.task_date, "todo", None)
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return {**task.dict(), "id": new_id, "status": "todo", "done_at": None}

@app.put("/api/tasks/{task_id}/status")
def update_task_status(task_id: int, body: dict):
    status  = body["status"]
    done_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if status == "done" else None
    conn    = get_db()
    conn.execute("UPDATE tasks SET status=?, done_at=? WHERE id=?", (status, done_at, task_id))
    conn.commit()
    conn.close()
    return {"id": task_id, "status": status, "done_at": done_at}

@app.delete("/api/tasks/{task_id}")
def delete_task(task_id: int):
    conn = get_db()
    conn.execute("DELETE FROM tasks WHERE id=?", (task_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── 数据标准 导入/导出 API ─────────────────────────────

def _csv_response(content, filename):
    from fastapi.responses import Response
    # Use ASCII filename to avoid latin-1 encoding error in headers
    safe_name = filename.encode("ascii", "ignore").decode("ascii")
    return Response(
        content=content.encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'}
    )

# --- 字根 导出模板 / 导出 / 导入 ---
@app.get("/api/data-roots/template")
def download_root_template():
    header = "字根ID,字根名,字根含义,字根类型,字根长度,字根码值,字根备注\n"
    example = 'ROOT_EXAMPLE,示例字根,这是一个示例,字符型,10,"[""A"",""B"",""C""]",示例备注\n'
    return _csv_response(header + example, "字根导入模板.csv")

@app.get("/api/data-roots/export")
def export_roots():
    conn = get_db()
    rows = conn.execute("SELECT id,name,meaning,root_type,length,code_values,remark FROM data_roots ORDER BY id").fetchall()
    conn.close()
    header = "字根ID,字根名,字根含义,字根类型,字根长度,字根码值,字根备注\n"
    lines = ""
    for r in rows:
        cv = (r[5] or "").replace('"', '""')
        lines += f'{r[0]},{r[1]},{r[2] or ""},{r[3] or ""},{r[4] or ""},"{cv}",{r[6] or ""}\n'
    return _csv_response(header + lines, f"字根导出_{datetime.now().strftime('%Y%m%d')}.csv")

@app.post("/api/data-roots/import")
async def import_roots(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode("utf-8-sig")
    import csv, io
    reader = csv.reader(io.StringIO(text))
    header = next(reader, None)
    if not header or len(header) < 4:
        return JSONResponse({"error": "文件格式不正确，请下载使用模板文件"}, status_code=400)
    success, errors = 0, 0
    conn = get_db()
    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, row in enumerate(reader, 2):
        try:
            rid = (row[0] or "").strip()
            rname = (row[1] or "").strip()
            if not rid or not rname:
                errors += 1; continue
            rmean = (row[2] or "").strip()
            rtype = (row[3] or "字符型").strip()
            rlen = int(row[4]) if len(row) > 4 and (row[4] or "").strip() else None
            rcode = (row[5] or "").strip() if len(row) > 5 else None
            rremark = (row[6] or "").strip() if len(row) > 6 else None
            existing = conn.execute("SELECT id FROM data_roots WHERE id=?", (rid,)).fetchone()
            if existing:
                conn.execute(
                    "UPDATE data_roots SET name=?,meaning=?,root_type=?,length=?,code_values=?,remark=?,updated_at=? WHERE id=?",
                    (rname, rmean, rtype, rlen, rcode, rremark, now_ts, rid)
                )
            else:
                conn.execute(
                    "INSERT INTO data_roots(id,name,meaning,root_type,length,code_values,remark,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
                    (rid, rname, rmean, rtype, rlen, rcode, rremark, now_ts, now_ts)
                )
            success += 1
        except Exception:
            errors += 1
    conn.commit()
    conn.close()
    return {"ok": True, "success": success, "errors": errors}

# --- 字段 导出模板 / 导出 / 导入 ---
@app.get("/api/data-fields/template")
def download_field_template():
    header = "字段ID,字段英文名,字段中文名,字段含义,引用字根ID,引用字根名,字段类型,字段长度,字段码值,字段备注\n"
    example = 'FIELD_EXAMPLE,exampleField,示例字段,用于示例,ROOT_EXAMPLE,示例字根,字符型,10,"[""A"",""B""]",示例备注\n'
    return _csv_response(header + example, "字段导入模板.csv")

@app.get("/api/data-fields/export")
def export_fields():
    conn = get_db()
    rows = conn.execute("SELECT id,name_en,name_cn,meaning,root_id,root_name,field_type,length,code_values,remark FROM data_fields ORDER BY id").fetchall()
    conn.close()
    header = "字段ID,字段英文名,字段中文名,字段含义,引用字根ID,引用字根名,字段类型,字段长度,字段码值,字段备注\n"
    lines = ""
    for r in rows:
        cv = (r[8] or "").replace('"', '""')
        lines += f'{r[0]},{r[1]},{r[2] or ""},{r[3] or ""},{r[4] or ""},{r[5] or ""},{r[6] or ""},{r[7] or ""},"{cv}",{r[9] or ""}\n'
    return _csv_response(header + lines, f"字段导出_{datetime.now().strftime('%Y%m%d')}.csv")

@app.post("/api/data-fields/import")
async def import_fields(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode("utf-8-sig")
    import csv, io
    reader = csv.reader(io.StringIO(text))
    header = next(reader, None)
    if not header or len(header) < 2:
        return JSONResponse({"error": "文件格式不正确，请下载使用模板文件"}, status_code=400)
    success, errors = 0, 0
    conn = get_db()
    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, row in enumerate(reader, 2):
        try:
            fid = (row[0] or "").strip()
            fname_en = (row[1] or "").strip()
            if not fid or not fname_en:
                errors += 1; continue
            fname_cn = (row[2] or "").strip() if len(row) > 2 else ""
            fmean = (row[3] or "").strip() if len(row) > 3 else ""
            froot_id = (row[4] or "").strip() if len(row) > 4 else None
            froot_name = (row[5] or "").strip() if len(row) > 5 else None
            ftype = (row[6] or "").strip() if len(row) > 6 else None
            flen = int(row[7]) if len(row) > 7 and (row[7] or "").strip() else None
            fcode = (row[8] or "").strip() if len(row) > 8 else None
            fremark = (row[9] or "").strip() if len(row) > 9 else ""
            existing = conn.execute("SELECT id FROM data_fields WHERE id=?", (fid,)).fetchone()
            if existing:
                conn.execute(
                    "UPDATE data_fields SET name_en=?,name_cn=?,meaning=?,root_id=?,root_name=?,field_type=?,length=?,code_values=?,remark=?,updated_at=? WHERE id=?",
                    (fname_en, fname_cn, fmean, froot_id, froot_name, ftype, flen, fcode, fremark, now_ts, fid)
                )
            else:
                conn.execute(
                    "INSERT INTO data_fields(id,name_en,name_cn,meaning,root_id,root_name,field_type,length,code_values,remark,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                    (fid, fname_en, fname_cn, fmean, froot_id, froot_name, ftype, flen, fcode, fremark, now_ts, now_ts)
                )
            success += 1
        except Exception:
            errors += 1
    conn.commit()
    conn.close()
    return {"ok": True, "success": success, "errors": errors}


# ── 数据标准 API ──────────────────────────────────────────

# 字根 CRUD
@app.get("/api/data-roots")
def list_roots():
    conn = get_db()
    rows = conn.execute("SELECT * FROM data_roots ORDER BY updated_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/data-roots")
def create_root(data: dict):
    conn = get_db()
    conn.execute(
        "INSERT INTO data_roots(id,name,meaning,root_type,length,code_values,remark,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (data["id"], data["name"], data.get("meaning"), data.get("root_type"),
         data.get("length"), data.get("code_values"), data.get("remark"),
         data.get("created_at"), data.get("updated_at"))
    )
    conn.commit()
    conn.close()
    return {"ok": True, "id": data["id"]}

@app.put("/api/data-roots/{root_id}")
def update_root(root_id: str, data: dict):
    conn = get_db()
    conn.execute(
        "UPDATE data_roots SET name=?,meaning=?,root_type=?,length=?,code_values=?,remark=?,updated_at=? WHERE id=?",
        (data["name"], data.get("meaning"), data.get("root_type"),
         data.get("length"), data.get("code_values"), data.get("remark"),
         data.get("updated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"), root_id)
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/data-roots/{root_id}")
def delete_root(root_id: str):
    conn = get_db()
    conn.execute("DELETE FROM data_roots WHERE id=?", (root_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# 字段 CRUD
@app.get("/api/data-fields")
def list_fields():
    conn = get_db()
    rows = conn.execute("SELECT * FROM data_fields ORDER BY updated_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/data-fields")
def create_field(data: dict):
    conn = get_db()
    conn.execute(
        "INSERT INTO data_fields(id,name_en,name_cn,meaning,root_id,root_name,field_type,length,code_values,remark,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
        (data["id"], data["name_en"], data.get("name_cn"), data.get("meaning"),
         data.get("root_id"), data.get("root_name"), data.get("field_type"),
         data.get("length"), data.get("code_values"), data.get("remark"),
         data.get("created_at"), data.get("updated_at"))
    )
    conn.commit()
    conn.close()
    return {"ok": True, "id": data["id"]}

@app.put("/api/data-fields/{field_id}")
def update_field(field_id: str, data: dict):
    conn = get_db()
    conn.execute(
        "UPDATE data_fields SET name_en=?,name_cn=?,meaning=?,root_id=?,root_name=?,field_type=?,length=?,code_values=?,remark=?,updated_at=? WHERE id=?",
        (data["name_en"], data.get("name_cn"), data.get("meaning"),
         data.get("root_id"), data.get("root_name"), data.get("field_type"),
         data.get("length"), data.get("code_values"), data.get("remark"),
         data.get("updated_at") or datetime.now().strftime('%Y-%m-%d %H:%M:%S'), field_id)
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/data-fields/{field_id}")
def delete_field(field_id: str):
    conn = get_db()
    conn.execute("DELETE FROM data_fields WHERE id=?", (field_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# 接口 CRUD
@app.get("/api/interfaces")
def list_interfaces():
    conn = get_db()
    rows = conn.execute("SELECT * FROM interfaces ORDER BY updated_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/interfaces")
def create_interface(data: dict):
    conn = get_db()
    conn.execute(
        "INSERT INTO interfaces(id,name,description,input_json,output_json,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
        (data["id"], data["name"], data.get("description"),
         data.get("input_json", "[]"), data.get("output_json", "[]"),
         data.get("created_at"), data.get("updated_at"))
    )
    conn.commit()
    conn.close()
    return {"ok": True, "id": data["id"]}

@app.put("/api/interfaces/{iface_id}")
def update_interface(iface_id: str, data: dict):
    conn = get_db()
    conn.execute(
        "UPDATE interfaces SET name=?,description=?,input_json=?,output_json=?,updated_at=? WHERE id=?",
        (data["name"], data.get("description"),
         data.get("input_json", "[]"), data.get("output_json", "[]"),
         data.get("updated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"), iface_id)
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/interfaces/{iface_id}")
def delete_interface(iface_id: str):
    conn = get_db()
    conn.execute("DELETE FROM interfaces WHERE id=?", (iface_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# 规则 CRUD
@app.get("/api/rules")
def list_rules():
    conn = get_db()
    rows = conn.execute("SELECT * FROM rules ORDER BY updated_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/rules")
def create_rule(data: dict):
    conn = get_db()
    conn.execute(
        "INSERT INTO rules(id,name,description,input_json,output_json,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
        (data["id"], data["name"], data.get("description"),
         data.get("input_json", "[]"), data.get("output_json", "[]"),
         data.get("created_at"), data.get("updated_at"))
    )
    conn.commit()
    conn.close()
    return {"ok": True, "id": data["id"]}

@app.put("/api/rules/{rule_id}")
def update_rule(rule_id: str, data: dict):
    conn = get_db()
    conn.execute(
        "UPDATE rules SET name=?,description=?,input_json=?,output_json=?,updated_at=? WHERE id=?",
        (data["name"], data.get("description"),
         data.get("input_json", "[]"), data.get("output_json", "[]"),
         data.get("updated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"), rule_id)
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/rules/{rule_id}")
def delete_rule(rule_id: str):
    conn = get_db()
    conn.execute("DELETE FROM rules WHERE id=?", (rule_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── 产品管理 API（一图一表入口）────────────────────────────

class ProductBase(BaseModel):
    product_id:      str
    product_name:    str
    product_desc:    Optional[str] = None
    product_manager: Optional[str] = None
    biz_contact:     Optional[str] = None
    biz_dept:        Optional[str] = None

class ProductCreate(ProductBase):
    pass

class ProductUpdate(BaseModel):
    product_name:    Optional[str] = None
    product_desc:    Optional[str] = None
    product_manager: Optional[str] = None
    biz_contact:     Optional[str] = None
    biz_dept:        Optional[str] = None

@app.get("/api/products", response_model=list[dict])
def list_products():
    conn = get_db()
    rows = conn.execute("SELECT * FROM products ORDER BY updated_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/products")
def create_product(p: ProductCreate):
    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO products(product_id,product_name,product_desc,product_manager,biz_contact,biz_dept,chart_data,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (p.product_id, p.product_name, p.product_desc, p.product_manager, p.biz_contact, p.biz_dept, None, now_ts, now_ts)
        )
        conn.commit()
        return {"ok": True, "id": cur.lastrowid}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.put("/api/products/{pid}")
def update_product(pid: int, p: ProductUpdate):
    conn = get_db()
    conn.execute(
        "UPDATE products SET product_name=?,product_desc=?,product_manager=?,biz_contact=?,biz_dept=?,updated_at=? WHERE id=?",
        (p.product_name, p.product_desc, p.product_manager, p.biz_contact, p.biz_dept,
         datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pid)
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/products/{pid}")
def delete_product(pid: int):
    conn = get_db()
    conn.execute("DELETE FROM products WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return {"ok": True}

# 获取产品的图表数据
@app.get("/api/products/{pid}/chart")
def get_product_chart(pid: int):
    conn = get_db()
    row = conn.execute("SELECT id,product_name,chart_data FROM products WHERE id=?", (pid,)).fetchone()
    conn.close()
    if not row:
        return JSONResponse({"error": "产品不存在"}, status_code=404)
    return {"id": row[0], "product_name": row[1], "chart_data": row[2]}

# 保存产品的图表数据（步骤信息）
@app.put("/api/products/{pid}/chart")
def save_product_chart(pid: int, body: dict):
    chart_data = body.get("chart_data")
    title = body.get("title")
    conn = get_db()
    if title:
        conn.execute("UPDATE products SET product_name=?,chart_data=?,updated_at=? WHERE id=?",
                     (title, chart_data, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pid))
    else:
        conn.execute("UPDATE products SET chart_data=?,updated_at=? WHERE id=?",
                     (chart_data, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── 步骤列表导出 Excel ──────────────────────────────────────

def _compute_step_codes(steps: list) -> dict:
    """与前端 computeStepCodes 保持一致的算法"""
    codes = {}
    pad = lambda n: str(n).zfill(4)

    l1i = 0
    for s in steps:
        if (s.get("level") or 1) == 1:
            l1i += 1
            codes[s["id"]] = f"P-{pad(l1i)}"

    l2cnt: dict = {}
    for s in steps:
        if (s.get("level") or 1) != 2:
            continue
        pid = s.get("parent_id")
        l2cnt[pid] = l2cnt.get(pid, 0) + 1
        codes[s["id"]] = f"{codes.get(pid, 'P-0000')}-{pad(l2cnt[pid])}"

    l3cnt: dict = {}
    for s in steps:
        if (s.get("level") or 1) != 3:
            continue
        pid = s.get("parent_id")
        l3cnt[pid] = l3cnt.get(pid, 0) + 1
        codes[s["id"]] = f"{codes.get(pid, 'P-0000-0000')}-{pad(l3cnt[pid])}"

    return codes


@app.get("/api/products/{pid}/steps/export")
def export_steps_excel(pid: int):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    conn = get_db()
    row = conn.execute("SELECT product_name, chart_data FROM products WHERE id=?", (pid,)).fetchone()
    conn.close()
    if not row:
        return JSONResponse({"error": "产品不存在"}, status_code=404)

    product_name = row[0] or "未命名产品"
    steps = []
    if row[1]:
        try:
            steps = json.loads(row[1]).get("steps", [])
        except Exception:
            steps = []

    codes = _compute_step_codes(steps)

    # 计算前序/后序（同层级相对顺序）
    same_level: dict = {}
    for s in steps:
        lv = s.get("level") or 1
        same_level.setdefault(lv, []).append(s)

    id_map = {s["id"]: s for s in steps}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "步骤列表"

    # ── 样式定义 ──
    header_font  = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor="3B5BDB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align   = Alignment(vertical="center", wrap_text=True)
    code_font    = Font(name="Courier New", size=9, color="475569")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    lv_colors = {1: "EEF4FF", 2: "FFF4E6", 3: "EDFBF1"}

    HEADERS = ["编号", "步骤名称", "步骤层级", "操作角色", "操作渠道",
               "上级步骤", "前序步骤", "后序步骤", "关联规则/接口", "步骤类型", "步骤描述"]
    COL_WIDTHS = [18, 28, 10, 14, 16, 20, 20, 20, 24, 10, 30]

    # 首行：产品名称合并
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1, value=f"产品：{product_name}  步骤清单")
    title_cell.font = Font(name="微软雅黑", bold=True, size=12, color="1E293B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="F1F5F9")
    ws.row_dimensions[1].height = 28

    # 表头行
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # 数据行
    for ri, s in enumerate(steps, 3):
        lv   = s.get("level") or 1
        code = codes.get(s["id"], "")
        peers = same_level.get(lv, [])
        pidx  = next((i for i, x in enumerate(peers) if x["id"] == s["id"]), -1)
        prev_step = peers[pidx - 1]["name"] if pidx > 0 else ""
        next_step = peers[pidx + 1]["name"] if 0 <= pidx < len(peers) - 1 else ""
        parent_name = id_map.get(s.get("parent_id"), {}).get("name", "") if s.get("parent_id") else ""
        channels = "、".join(s.get("channels") or [])
        linked = s.get("linked_name") or s.get("linked_id") or ""

        row_vals = [
            code,
            s.get("name", ""),
            f"L{lv}（{'一二三'[lv-1]}级）",
            s.get("role", ""),
            channels,
            parent_name,
            prev_step,
            next_step,
            linked,
            s.get("step_type", "步骤"),
            s.get("desc", ""),
        ]

        row_fill = PatternFill("solid", fgColor=lv_colors.get(lv, "FFFFFF"))
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = cell_align
            c.border = border
            c.fill = row_fill
            if ci == 1:
                c.font = code_font
            else:
                c.font = Font(name="微软雅黑", size=10)

        ws.row_dimensions[ri].height = 18

    # 列宽
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    safe_name = f"{product_name}_步骤清单_{datetime.now().strftime('%Y%m%d')}.xlsx"
    encoded   = quote(safe_name)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ── 产品变更日志 API ────────────────────────────────────────

@app.get("/api/products/{pid}/changelogs")
def get_product_changelogs(pid: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT id,product_id,change_desc,changed_at FROM product_changelogs WHERE product_id=? ORDER BY changed_at DESC",
        (pid,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/products/{pid}/changelogs")
def add_product_changelog(pid: int, body: dict):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO product_changelogs(product_id,change_desc,changed_at) VALUES(?,?,?)",
        (pid, body.get("change_desc", ""), now)
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return {"id": new_id, "product_id": pid, "change_desc": body.get("change_desc", ""), "changed_at": now}


# ── 股票专属评分模型 API ─────────────────────────────────────────────────────

@app.get("/api/stock-models")
def list_stock_models():
    """列出所有已注册的专属模型"""
    conn = get_db()
    rows = conn.execute(
        "SELECT code,name,threshold,ic,icir,accuracy,up_win_rate,sample_days,description,updated_at "
        "FROM stock_custom_models ORDER BY updated_at DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/stock-models/{code}")
def get_stock_model(code: str):
    """获取某只股票的专属模型（含权重）"""
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM stock_custom_models WHERE code=?", (code,)
    ).fetchone()
    conn.close()
    if not row:
        return JSONResponse({"error": f"{code} 暂无专属模型"}, status_code=404)
    d = dict(row)
    d["features"] = json.loads(d["features"])
    d["weights"]  = json.loads(d["weights"])
    return d

@app.put("/api/stock-models/{code}")
async def upsert_stock_model(code: str, request: Request):
    """注册或更新某只股票的专属评分模型"""
    body = await request.json()
    required = ["name","features","weights","threshold","ic","icir","accuracy","up_win_rate"]
    for f in required:
        if f not in body:
            return JSONResponse({"error": f"缺少字段 {f}"}, status_code=400)
    now = datetime.now().isoformat()
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO stock_custom_models
                (code,name,features,weights,threshold,ic,icir,accuracy,up_win_rate,
                 sample_days,description,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(code) DO UPDATE SET
                name=excluded.name, features=excluded.features, weights=excluded.weights,
                threshold=excluded.threshold, ic=excluded.ic, icir=excluded.icir,
                accuracy=excluded.accuracy, up_win_rate=excluded.up_win_rate,
                sample_days=excluded.sample_days, description=excluded.description,
                updated_at=excluded.updated_at
        """, (
            code,
            body["name"],
            json.dumps(body["features"], ensure_ascii=False),
            json.dumps(body["weights"]),
            float(body["threshold"]),
            float(body["ic"]),
            float(body["icir"]),
            float(body["accuracy"]),
            float(body["up_win_rate"]),
            body.get("sample_days"),
            body.get("description",""),
            now, now,
        ))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.delete("/api/stock-models/{code}")
def delete_stock_model(code: str):
    """删除某只股票的专属模型"""
    conn = get_db()
    try:
        conn.execute("DELETE FROM stock_custom_models WHERE code=?", (code,))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.post("/api/stock-models/{code}/export")
def export_model_to_quant(code: str):
    """将专属模型导出为 quant_trading/config/custom_models.json"""
    conn = get_db()
    rows = conn.execute("SELECT * FROM stock_custom_models").fetchall()
    conn.close()
    models = {}
    for r in rows:
        d = dict(r)
        models[d["code"]] = {
            "name":      d["name"],
            "features":  json.loads(d["features"]),
            "weights":   json.loads(d["weights"]),
            "threshold": d["threshold"],
            "metrics":   {"ic": d["ic"], "icir": d["icir"],
                          "accuracy": d["accuracy"], "up_win_rate": d["up_win_rate"]},
        }
    out = QUANT_DIR / "config" / "custom_models.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(models, f, ensure_ascii=False, indent=2)
    return {"ok": True, "path": str(out), "count": len(models)}

# ── 量化参数 API（因子权重配置）────────────────────────────

class FactorWeightCreate(BaseModel):
    factor_key:   str
    factor_name:  str
    weight:       float
    description:  Optional[str] = None
    is_active:    Optional[int] = 1

class FactorWeightUpdate(BaseModel):
    factor_name:  Optional[str] = None
    weight:       Optional[float] = None
    description:  Optional[str] = None
    is_active:    Optional[int] = None

@app.get("/api/factor-weights")
def list_factor_weights():
    """获取所有因子权重列表"""
    conn = get_db()
    rows = conn.execute("SELECT * FROM factor_weights ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/factor-weights")
def create_factor_weight(f: FactorWeightCreate):
    """新建因子权重"""
    now_ts = datetime.now().isoformat()
    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO factor_weights(factor_key,factor_name,weight,description,is_active,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
            (f.factor_key, f.factor_name, f.weight, f.description, f.is_active, now_ts, now_ts)
        )
        conn.commit()
        return {"ok": True, "id": cur.lastrowid}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.put("/api/factor-weights/{fid}")
def update_factor_weight(fid: int, f: FactorWeightUpdate):
    """更新因子权重"""
    conn = get_db()
    try:
        fields = []
        values = []
        if f.factor_name is not None:
            fields.append("factor_name=?"); values.append(f.factor_name)
        if f.weight is not None:
            fields.append("weight=?"); values.append(f.weight)
        if f.description is not None:
            fields.append("description=?"); values.append(f.description)
        if f.is_active is not None:
            fields.append("is_active=?"); values.append(f.is_active)
        fields.append("updated_at=?"); values.append(datetime.now().isoformat())
        values.append(fid)
        conn.execute(f"UPDATE factor_weights SET {','.join(fields)} WHERE id=?", values)
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.delete("/api/factor-weights/{fid}")
def delete_factor_weight(fid: int):
    """删除因子权重"""
    conn = get_db()
    try:
        conn.execute("DELETE FROM factor_weights WHERE id=?", (fid,))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

# ── 因子细项参数 API ──────────────────────────────────────

class SubParamUpdate(BaseModel):
    param_value: float

@app.get("/api/factor-sub-params")
def list_sub_params():
    """获取所有细项参数（按 factor_key 分组）"""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM factor_sub_params ORDER BY factor_key, id"
    ).fetchall()
    conn.close()
    result: dict = {}
    for r in rows:
        d = dict(r)
        fk = d["factor_key"]
        result.setdefault(fk, []).append(d)
    return result

@app.put("/api/factor-sub-params/{pid}")
def update_sub_param(pid: int, body: SubParamUpdate):
    """更新单个细项参数值"""
    conn = get_db()
    try:
        conn.execute("UPDATE factor_sub_params SET param_value=? WHERE id=?",
                     (body.param_value, pid))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        conn.close()

@app.post("/api/factor-sub-params/sync")
def sync_sub_params_to_quant():
    """将细项参数导出为 JSON 供股神计划使用"""
    import json
    conn = get_db()
    rows = conn.execute("SELECT * FROM factor_sub_params ORDER BY factor_key, id").fetchall()
    conn.close()

    config: dict = {}
    for r in rows:
        d = dict(r)
        config.setdefault(d["factor_key"], {})[d["param_key"]] = d["param_value"]

    # 同时写入因子主权重
    conn2 = get_db()
    fw_rows = conn2.execute("SELECT factor_key, weight, is_active FROM factor_weights").fetchall()
    conn2.close()
    weights = {r["factor_key"]: {"weight": r["weight"], "is_active": bool(r["is_active"])} for r in fw_rows}
    config["_weights"] = weights

    out_path = Path.home() / "project" / "quant_trading" / "config" / "scorer_params.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "path": str(out_path)}


# ── 富甲天下5 存档修改器 ──────────────────────────────────

import zlib as _zlib
import shutil as _shutil
try:
    import opencc as _opencc
    _t2s = _opencc.OpenCC('t2s')
    _s2t = _opencc.OpenCC('s2t')
except Exception:
    _t2s = None
    _s2t = None

# 存档字段定义：(显示名, 主公记录内相对偏移)
_SAV_FIELDS = [
    ("金额",   0x0C4),
    ("士兵",   0x0E0),
    ("金属",   0x0E4),
    ("木材",   0x0E8),
    ("石头",   0x0EC),
    ("棉花",   0x0F0),
    ("皮革",   0x0F4),
]

_PET_RECORD_SIZE = 0x248
_PET_LEGACY_CURRENT_BASE = 0x248
_PET_NAME_OFF = 0x40
_PET_LEVEL_OFF = 0x54
_PET_UPGRADE_POINTS_OFF = 0x88
_LORD_NAME_OFF = 0x04
_LORD_NAME_LEN = 0x08
_LORD_HERO_IDX_OFF = 0x08

# 武将数据：每武将252字节。自由/剧情模式块号不同，需要按结构定位。
_HERO_UNIT_SIZE  = 252
_HERO_PRIMARY_TOTAL = 130
_HERO_SECONDARY_TOTAL = 80
_HERO_PRIMARY_NAME_OFF = 12
_HERO_SECONDARY_NAME_OFF = 4
_LORD_MEMBER_RANGES = (
    (0x130, 0x148),
    (0x150, 0x164),
    (0x170, 0x18C),
    (0x190, 0x1A4),
)
# 维度字段：(显示名, 块内偏移)
_HERO_DIMS = [
    ("体力上限", 44),
    ("武力",    60),
    ("智力",    68),
    ("陆战",    76),
    ("水战",    84),
    ("林战",    92),
    ("防御力",  108),
]

def _sav_clean_big5_name(raw: bytes) -> str:
    try:
        name = bytes(raw).rstrip(b'\x20\x00').decode('big5')
    except Exception:
        return ""
    return name.replace(' ', '')

def _sav_encode_big5_name(name: str, width: int = 8) -> bytes:
    name = str(name or "").strip().replace(" ", "")
    if _s2t:
        name = _s2t.convert(name)
    if not name:
        raise ValueError("名称不能为空")

    char_bytes = []
    for ch in name:
        try:
            char_bytes.append(ch.encode('big5'))
        except Exception as e:
            raise ValueError(f"名称包含无法写入 Big5 的字符：{ch}") from e

    raw_len = sum(len(b) for b in char_bytes)
    if raw_len > width:
        raise ValueError(f"名称过长，最多 {width} 字节 Big5")
    if len(char_bytes) <= 1:
        return b''.join(char_bytes).ljust(width, b'\x00')

    spaces = width - raw_len
    gaps = len(char_bytes) - 1
    base_spaces, extra = divmod(spaces, gaps)
    out = bytearray()
    for idx, part in enumerate(char_bytes):
        out += part
        if idx < gaps:
            out += b' ' * (base_spaces + (1 if idx < extra else 0))
    return bytes(out).ljust(width, b'\x00')

def _sav_is_cjk_name(name: str, min_len: int = 2) -> bool:
    if len(name) < min_len:
        return False
    return all(
        ('\u3400' <= ch <= '\u9fff') or ('\uf900' <= ch <= '\ufaff')
        for ch in name
    )

def _sav_read_pet_at(data: bytearray | bytes, base: int) -> dict | None:
    need_len = base + _PET_UPGRADE_POINTS_OFF + 4
    if len(data) < need_len:
        return None
    name = _sav_clean_big5_name(data[base + _PET_NAME_OFF: base + _PET_NAME_OFF + 8])
    if not _sav_is_cjk_name(name):
        return None
    name_clean = name.replace(' ', '')
    name_s = (_t2s.convert(name_clean) if _t2s else name_clean)
    return {
        "base": base,
        "name": name_clean,
        "name_s": name_s,
        "level": struct.unpack_from('<I', data, base + _PET_LEVEL_OFF)[0],
        "upgrade_points": struct.unpack_from('<I', data, base + _PET_UPGRADE_POINTS_OFF)[0],
    }

def _sav_scan_pets(data: bytearray | bytes) -> list[dict]:
    pets = []
    for base in range(0, len(data), _PET_RECORD_SIZE):
        pet = _sav_read_pet_at(data, base)
        if pet:
            pet["slot"] = base // _PET_RECORD_SIZE
            pets.append(pet)
    return pets

def _sav_find_state_block_idx(blocks: list[dict]) -> int | None:
    min_len = max(off for _, off in _SAV_FIELDS) + 4
    for idx, block in enumerate(blocks):
        data = block['data']
        if len(data) < min_len:
            continue
        if len(data) % _PET_RECORD_SIZE == 0 and _sav_scan_pets(data):
            return idx
    return None

def _sav_pick_current_pet(data: bytearray | bytes) -> dict | None:
    legacy_pet = _sav_read_pet_at(data, _PET_LEGACY_CURRENT_BASE)
    if legacy_pet:
        legacy_pet["slot"] = _PET_LEGACY_CURRENT_BASE // _PET_RECORD_SIZE
        return legacy_pet
    pets = _sav_scan_pets(data)
    return pets[0] if pets else None

def _sav_current_lord_name(blocks: list[dict]) -> str:
    if not blocks:
        return ""
    data = blocks[0]['data']
    if len(data) < _LORD_NAME_OFF + _LORD_NAME_LEN:
        return ""
    return _sav_clean_big5_name(data[_LORD_NAME_OFF: _LORD_NAME_OFF + _LORD_NAME_LEN])

def _sav_current_lord_idx(blocks: list[dict]) -> int | None:
    lord_name = _sav_current_lord_name(blocks)
    if not lord_name:
        return None
    lord_name_s = (_t2s.convert(lord_name) if _t2s else lord_name)
    for hero in _sav_read_heroes(blocks):
        if hero["name"] == lord_name or hero["name_s"] == lord_name_s:
            return hero["idx"]
    return None

def _sav_find_current_lord_record(blocks: list[dict]) -> tuple[int, int]:
    state_idx = _sav_find_state_block_idx(blocks)
    if state_idx is None:
        raise ValueError("未识别到玩家资源块")
    data = blocks[state_idx]['data']
    lord_idx = _sav_current_lord_idx(blocks)
    if lord_idx is not None:
        for base in range(0, len(data), _PET_RECORD_SIZE):
            if base + _LORD_HERO_IDX_OFF + 4 > len(data):
                break
            record_lord_idx = struct.unpack_from('<I', data, base + _LORD_HERO_IDX_OFF)[0]
            if record_lord_idx == lord_idx:
                return state_idx, base

    pet = _sav_pick_current_pet(data)
    if pet:
        return state_idx, pet["base"]
    raise ValueError("未识别到当前主公记录")

def _sav_read_lord_fields_at(data: bytearray | bytes, base: int) -> dict:
    return {name: struct.unpack_from('<I', data, base + off)[0] for name, off in _SAV_FIELDS}

def _sav_read_lords(blocks: list[dict]) -> list[dict]:
    state_idx = _sav_find_state_block_idx(blocks)
    if state_idx is None:
        return []
    data = blocks[state_idx]['data']
    heroes = {h["idx"]: h for h in _sav_read_heroes(blocks)}
    current_idx = _sav_current_lord_idx(blocks)
    lords = []
    for slot, base in enumerate(range(0, len(data), _PET_RECORD_SIZE)):
        if base + max(off for _, off in _SAV_FIELDS) + 4 > len(data):
            break
        hero_idx = struct.unpack_from('<I', data, base + _LORD_HERO_IDX_OFF)[0]
        hero = heroes.get(hero_idx)
        if not hero:
            continue
        pet = _sav_read_pet_at(data, base)
        if pet:
            pet["slot"] = slot
        lords.append({
            "slot": slot,
            "base": base,
            "hero_idx": hero_idx,
            "name": hero["name"],
            "name_s": hero["name_s"],
            "is_current": current_idx is not None and hero_idx == current_idx,
            "fields": _sav_read_lord_fields_at(data, base),
            "pet": pet,
        })
    return lords

def _sav_write_lords(blocks: list[dict], lords_patch) -> None:
    if not lords_patch:
        return
    state_idx = _sav_find_state_block_idx(blocks)
    if state_idx is None:
        raise ValueError("未识别到玩家资源块")
    data = blocks[state_idx]['data']
    field_off = {name: off for name, off in _SAV_FIELDS}

    if isinstance(lords_patch, list):
        items = [(entry.get("slot"), entry) for entry in lords_patch if isinstance(entry, dict)]
    else:
        items = list((lords_patch or {}).items())

    for slot_key, patch in items:
        if not isinstance(patch, dict):
            continue
        try:
            slot = int(slot_key)
        except Exception:
            continue
        base = slot * _PET_RECORD_SIZE
        if base + max(off for _, off in _SAV_FIELDS) + 4 > len(data):
            continue

        for name, value in (patch.get("fields") or {}).items():
            if name not in field_off:
                continue
            struct.pack_into('<I', data, base + field_off[name], max(0, int(value)))

        pet_patch = patch.get("pet") or {}
        if pet_patch and _sav_read_pet_at(data, base):
            if "name" in pet_patch:
                data[base + _PET_NAME_OFF: base + _PET_NAME_OFF + 8] = _sav_encode_big5_name(pet_patch.get("name"))
            if "level" in pet_patch:
                struct.pack_into('<I', data, base + _PET_LEVEL_OFF, max(0, int(pet_patch.get("level") or 0)))
            if "upgrade_points" in pet_patch:
                value = max(0, int(pet_patch.get("upgrade_points") or 0))
                struct.pack_into('<I', data, base + _PET_UPGRADE_POINTS_OFF, value)

def _sav_read_fields(blocks: list[dict]) -> dict:
    state_idx, base = _sav_find_current_lord_record(blocks)
    data = blocks[state_idx]['data']
    return _sav_read_lord_fields_at(data, base)

def _sav_write_fields(blocks: list[dict], fields_patch: dict) -> None:
    state_idx, base = _sav_find_current_lord_record(blocks)
    data = blocks[state_idx]['data']
    field_off = {name: off for name, off in _SAV_FIELDS}
    for name, value in fields_patch.items():
        if name not in field_off:
            continue
        struct.pack_into('<I', data, base + field_off[name], max(0, int(value)))

def _sav_read_current_pet(blocks: list[dict]) -> dict | None:
    try:
        state_idx, base = _sav_find_current_lord_record(blocks)
    except Exception:
        return None
    pet = _sav_read_pet_at(blocks[state_idx]['data'], base)
    if pet:
        pet["slot"] = base // _PET_RECORD_SIZE
    return pet

def _sav_write_current_pet(blocks: list[dict], pet_patch: dict):
    if not any(key in pet_patch for key in ("name", "level", "upgrade_points")):
        return
    state_idx, base = _sav_find_current_lord_record(blocks)
    data = blocks[state_idx]['data']
    pet = _sav_read_pet_at(data, base)
    if not pet:
        raise ValueError("当前宠物数据不存在")
    if "name" in pet_patch:
        data[base + _PET_NAME_OFF: base + _PET_NAME_OFF + 8] = _sav_encode_big5_name(pet_patch.get("name"))
    if "level" in pet_patch:
        struct.pack_into('<I', data, base + _PET_LEVEL_OFF, max(0, int(pet_patch.get("level") or 0)))
    if "upgrade_points" in pet_patch:
        value = max(0, int(pet_patch.get("upgrade_points") or 0))
        struct.pack_into('<I', data, base + _PET_UPGRADE_POINTS_OFF, value)

def _sav_valid_hero_name(data: bytearray | bytes, base: int, name_off: int) -> bool:
    if base + name_off + 8 > len(data):
        return False
    name = _sav_clean_big5_name(data[base + name_off: base + name_off + 8])
    return _sav_is_cjk_name(name)

def _sav_find_hero_tables(blocks: list[dict]) -> list[dict]:
    tables = []
    primary_need = (_HERO_PRIMARY_TOTAL - 1) * _HERO_UNIT_SIZE + max(off for _, off in _HERO_DIMS) + 4
    secondary_dim_offsets = {name: off - 8 for name, off in _HERO_DIMS}
    secondary_need = (
        (_HERO_SECONDARY_TOTAL - 1) * _HERO_UNIT_SIZE
        + max(secondary_dim_offsets.values()) + 4
    )

    for block_idx, block in enumerate(blocks):
        data = block['data']
        if len(data) >= primary_need:
            checks = [0, 1, 10, 50, 100, 129]
            valid = sum(
                1 for row in checks
                if _sav_valid_hero_name(data, row * _HERO_UNIT_SIZE, _HERO_PRIMARY_NAME_OFF)
            )
            if valid >= 4:
                tables.append({
                    "block_idx": block_idx,
                    "total": _HERO_PRIMARY_TOTAL,
                    "name_off": _HERO_PRIMARY_NAME_OFF,
                    "dim_offsets": dict(_HERO_DIMS),
                })
                continue

        if len(data) >= secondary_need:
            checks = [0, 1, 10, 30, 50, 79]
            valid = sum(
                1 for row in checks
                if _sav_valid_hero_name(data, row * _HERO_UNIT_SIZE, _HERO_SECONDARY_NAME_OFF)
            )
            if valid >= 4:
                tables.append({
                    "block_idx": block_idx,
                    "total": _HERO_SECONDARY_TOTAL,
                    "name_off": _HERO_SECONDARY_NAME_OFF,
                    "dim_offsets": secondary_dim_offsets,
                })
    return tables

def _sav_read_heroes(blocks: list[dict]) -> list[dict]:
    heroes = []
    hero_idx = 0
    for table in _sav_find_hero_tables(blocks):
        data = blocks[table["block_idx"]]['data']
        for row in range(table["total"]):
            base = row * _HERO_UNIT_SIZE
            name_clean = _sav_clean_big5_name(data[base + table["name_off"]: base + table["name_off"] + 8])
            if not name_clean:
                name_clean = f"Hero#{hero_idx}"
            name_s = (_t2s.convert(name_clean) if _t2s else name_clean)
            dims = {}
            for dim_name, off in table["dim_offsets"].items():
                dims[dim_name] = struct.unpack_from('<I', data, base + off)[0]
            heroes.append({"idx": hero_idx, "name": name_clean, "name_s": name_s, "dims": dims})
            hero_idx += 1
    owner_map = _sav_hero_lord_map(blocks, heroes)
    for hero in heroes:
        hero["lord_name"] = owner_map.get(hero["idx"], "")
    return heroes

def _sav_hero_lord_map(blocks: list[dict], heroes: list[dict]) -> dict[int, str]:
    state_idx = _sav_find_state_block_idx(blocks)
    if state_idx is None:
        return {}
    data = blocks[state_idx]['data']
    hero_by_idx = {h["idx"]: h for h in heroes}
    owner_names: dict[int, list[str]] = {}

    def add_owner(hero_idx: int, lord_name: str):
        if hero_idx not in hero_by_idx or not lord_name:
            return
        names = owner_names.setdefault(hero_idx, [])
        if lord_name not in names:
            names.append(lord_name)

    for base in range(0, len(data), _PET_RECORD_SIZE):
        if base + _LORD_HERO_IDX_OFF + 4 > len(data):
            break
        lord_idx = struct.unpack_from('<I', data, base + _LORD_HERO_IDX_OFF)[0]
        lord = hero_by_idx.get(lord_idx)
        if not lord:
            continue
        lord_name = lord.get("name_s") or lord.get("name") or f"#{lord_idx}"
        add_owner(lord_idx, lord_name)

        for start, end in _LORD_MEMBER_RANGES:
            for rel_off in range(start, end, 4):
                if base + rel_off + 4 > len(data):
                    continue
                hero_idx = struct.unpack_from('<i', data, base + rel_off)[0]
                if hero_idx <= 0:
                    continue
                add_owner(hero_idx, lord_name)

    return {idx: "、".join(names) for idx, names in owner_names.items()}

def _sav_write_heroes(blocks: list[dict], heroes_patch: dict):
    """heroes_patch: {idx: {dim_name: value, ...}}"""
    hero_map = {}
    hero_idx = 0
    for table in _sav_find_hero_tables(blocks):
        data = blocks[table["block_idx"]]['data']
        for row in range(table["total"]):
            hero_map[hero_idx] = (data, row * _HERO_UNIT_SIZE, table["dim_offsets"])
            hero_idx += 1

    for idx_str, dim_map in heroes_patch.items():
        try:
            idx = int(idx_str)
        except Exception:
            continue
        if idx not in hero_map:
            continue
        data, base, dim_off = hero_map[idx]
        for dim_name, value in dim_map.items():
            if dim_name not in dim_off:
                continue
            struct.pack_into('<I', data, base + dim_off[dim_name], max(0, int(value)))

def _sav_try_full_block(raw: bytes, offset: int) -> dict | None:
    if offset + 10 > len(raw):
        return None
    flag = struct.unpack_from('<I', raw, offset)[0]
    size = struct.unpack_from('<I', raw, offset + 4)[0]
    if size == 0 or offset + 8 + size > len(raw):
        return None
    compressed = raw[offset + 8: offset + 8 + size]
    if not compressed or compressed[0] != 0x78:
        return None
    try:
        decompressed = _zlib.decompress(compressed)
    except Exception:
        return None
    return {"header": "full", "flag": flag, "size": size, "data": bytearray(decompressed)}

def _sav_try_short_block(raw: bytes, offset: int) -> dict | None:
    if offset + 6 > len(raw):
        return None
    size = struct.unpack_from('<I', raw, offset)[0]
    if size == 0 or offset + 4 + size > len(raw):
        return None
    compressed = raw[offset + 4: offset + 4 + size]
    if not compressed or compressed[0] != 0x78:
        return None
    try:
        decompressed = _zlib.decompress(compressed)
    except Exception:
        return None
    return {"header": "short", "flag": None, "size": size, "data": bytearray(decompressed)}

def _sav_try_block(raw: bytes, offset: int) -> dict | None:
    return _sav_try_full_block(raw, offset) or _sav_try_short_block(raw, offset)

def _sav_parse_blocks(raw: bytes) -> tuple[list[dict], bytes]:
    """
    将 .sav 文件拆分为 zlib 压缩块列表。
    支持自由模式的连续 full 块，也支持剧情模式中夹杂的 4 字节裸值
    以及「4字节长度 + zlib」短头块。
    """
    offset = 0
    blocks = []
    pending_prefix = bytearray()
    while offset < len(raw):
        parsed = _sav_try_block(raw, offset)
        if parsed:
            parsed["prefix"] = bytes(pending_prefix)
            pending_prefix.clear()
            blocks.append(parsed)
            offset += (8 if parsed["header"] == "full" else 4) + parsed["size"]
            continue

        if offset + 4 <= len(raw) and _sav_try_block(raw, offset + 4):
            pending_prefix += raw[offset: offset + 4]
            offset += 4
            continue

        return blocks, bytes(pending_prefix) + raw[offset:]
    return blocks, bytes(pending_prefix)

def _sav_build(blocks: list[dict], tail: bytes) -> bytes:
    """将块列表重新压缩拼合，并追加原始尾部数据。"""
    out = bytearray()
    for b in blocks:
        out += b.get("prefix", b"")
        compressed = _zlib.compress(bytes(b['data']), level=1)
        if b.get("header") == "short":
            out += struct.pack('<I', len(compressed))
            out += compressed
        else:
            out += struct.pack('<I', int(b.get('flag') or 0))
            out += struct.pack('<I', len(compressed))
            out += compressed
    out += tail
    return bytes(out)

def _sav_original_filename(filename: str | None, default: str = "0.sav") -> str:
    raw = str(filename or "").strip()
    name = re.split(r"[\\/]+", raw)[-1].strip()
    if not name or name in {".", ".."}:
        return default
    return name

def _sav_sha256_bytes(raw: bytes) -> str:
    import hashlib as _hashlib
    return _hashlib.sha256(raw).hexdigest()

def _sav_sha256_file(path: Path) -> str:
    import hashlib as _hashlib
    h = _hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def _sav_file_matches_raw(path: Path, raw: bytes) -> bool:
    try:
        if not path.is_file() or path.stat().st_size != len(raw):
            return False
        with path.open("rb") as f:
            for off in range(0, len(raw), 1024 * 1024):
                if f.read(1024 * 1024) != raw[off:off + 1024 * 1024]:
                    return False
        return True
    except Exception:
        return False

def _sav_matching_source_paths(filename: str | None, raw: bytes) -> list[str]:
    name = _sav_original_filename(filename, "")
    if not name:
        return []
    roots = [
        Path.home() / "Downloads" / "富甲天下5",
        Path.home() / "Downloads",
        Path.home() / "Desktop",
    ]
    matches: list[Path] = []
    seen: set[Path] = set()
    for root in roots:
        if not root.exists():
            continue
        try:
            iterator = root.rglob(name)
            for path in iterator:
                try:
                    resolved = path.resolve()
                except Exception:
                    resolved = path
                if resolved in seen:
                    continue
                seen.add(resolved)
                if _sav_file_matches_raw(path, raw):
                    matches.append(path)
        except Exception:
            continue
    matches.sort(key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)
    return [str(p) for p in matches]

def _sav_normalize_source_path(
    source_path: str | None,
    original_filename: str | None = None,
    original_size: int | None = None,
    original_sha256: str | None = None,
) -> str:
    if not source_path:
        return ""
    path = Path(str(source_path).strip()).expanduser()
    if not path.is_absolute() or not path.exists() or not path.is_file():
        raise ValueError("原始存档路径不存在，请填写完整的本机文件路径")
    expected_name = _sav_original_filename(original_filename, "")
    if expected_name and path.name != expected_name:
        raise ValueError(f"路径文件名 {path.name} 与本次上传文件 {expected_name} 不一致")
    if original_size is not None and path.stat().st_size != int(original_size):
        raise ValueError("路径对应文件大小与本次上传存档不一致")
    if original_sha256:
        expected_sha = str(original_sha256).strip().lower()
        if _sav_sha256_file(path).lower() != expected_sha:
            raise ValueError("路径对应文件内容与本次上传存档不一致")
    return str(path)

def _sav_resolve_parse_source_path(
    source_path: str | None,
    original_filename: str,
    raw: bytes,
    original_sha256: str,
) -> tuple[str, str, str, list[str]]:
    explicit_error = ""
    if source_path:
        try:
            verified = _sav_normalize_source_path(
                source_path,
                original_filename,
                len(raw),
                original_sha256,
            )
            return verified, "verified", "原始路径已按文件内容校验通过", []
        except Exception as exc:
            explicit_error = str(exc)

    matches = _sav_matching_source_paths(original_filename, raw)
    if len(matches) == 1:
        return matches[0], "inferred", "已通过文件名和文件内容自动定位到原始路径", []
    if len(matches) > 1:
        return "", "ambiguous", "找到多个内容完全相同的存档，请手动确认原始完整路径", matches[:5]
    if explicit_error:
        return "", "manual_required", f"{explicit_error}；请手动填写本次存档的完整路径", []
    return "", "manual_required", "未获得原始存档的真实完整路径，请通过目录选择功能重新选择存档；留空时会生成到应用上传目录的 new 目录", []

def _sav_output_path(tmp_path: Path, source_path: str | None, original_filename: str | None = None) -> Path:
    if source_path:
        src = Path(source_path).expanduser()
        if not src.exists() or not src.is_file():
            raise ValueError("原始存档路径不存在，请填写完整的本机文件路径")
        new_dir = src.parent / "new"
        file_name = src.name
    else:
        new_dir = tmp_path.parent / "new"
        file_name = _sav_original_filename(original_filename, tmp_path.name)
    new_dir.mkdir(parents=True, exist_ok=True)
    return new_dir / file_name

def _sav_download_url_for(path: Path) -> str | None:
    try:
        rel = path.resolve().relative_to(UPLOAD_DIR.resolve())
        return "/uploads/" + rel.as_posix()
    except Exception:
        return None

def _sav_default_pick_dir() -> Path:
    for path in (
        Path.home() / "Downloads" / "富甲天下5",
        Path.home() / "Downloads",
        Path.home() / "Desktop",
        Path.home(),
    ):
        if path.exists() and path.is_dir():
            return path
    return Path.home()

def _sav_picker_roots() -> list[dict]:
    candidates = [
        ("富甲天下5", Path.home() / "Downloads" / "富甲天下5"),
        ("下载", Path.home() / "Downloads"),
        ("桌面", Path.home() / "Desktop"),
        ("用户目录", Path.home()),
    ]
    roots = []
    seen: set[Path] = set()
    for name, path in candidates:
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        if resolved in seen or not path.exists() or not path.is_dir():
            continue
        seen.add(resolved)
        roots.append({"name": name, "path": str(path)})
    return roots

def _sav_template_now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _sav_clean_template_text(value, max_len: int = 40) -> str:
    return str(value or "").strip()[:max_len]

def _sav_sanitize_hero_template_heroes(heroes) -> dict:
    valid_dims = {name for name, _ in _HERO_DIMS}
    cleaned = {}
    if not isinstance(heroes, dict):
        return cleaned

    for key, item in heroes.items():
        if not isinstance(item, dict):
            continue
        try:
            idx = int(item.get("idx", key))
        except Exception:
            continue

        raw_dims = item.get("dims")
        if not isinstance(raw_dims, dict):
            raw_dims = item.get("changes") if isinstance(item.get("changes"), dict) else {}
        dims = {}
        for dim_name, value in raw_dims.items():
            if dim_name not in valid_dims:
                continue
            try:
                dims[dim_name] = max(0, min(9999, int(value)))
            except Exception:
                continue
        if not dims:
            continue

        cleaned[str(idx)] = {
            "idx": idx,
            "name": _sav_clean_template_text(item.get("name"), 16),
            "name_s": _sav_clean_template_text(item.get("name_s"), 16),
            "lord_name": _sav_clean_template_text(item.get("lord_name"), 60),
            "dims": dims,
        }
    return cleaned

def _sav_hero_template_counts(heroes: dict) -> tuple[int, int]:
    hero_count = len(heroes)
    field_count = sum(len((item or {}).get("dims") or {}) for item in heroes.values())
    return hero_count, field_count

def _sav_template_from_row(row) -> dict:
    try:
        data = json.loads(row["data_json"] or "{}")
    except Exception:
        data = {}
    heroes = data.get("heroes") if isinstance(data.get("heroes"), dict) else {}
    return {
        "id": row["id"],
        "name": row["name"],
        "heroes": heroes,
        "hero_count": row["hero_count"],
        "field_count": row["field_count"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }

def _sav_parse_payload(
    raw: bytes,
    original_filename: str,
    source_path_norm: str = "",
    source_path_status: str = "",
    source_path_message: str = "",
    source_path_candidates: list[str] | None = None,
) -> JSONResponse:
    original_sha256 = _sav_sha256_bytes(raw)
    try:
        blocks, tail = _sav_parse_blocks(raw)
    except Exception as e:
        return JSONResponse({"error": f"解析失败: {e}"}, status_code=400)
    if not blocks or _sav_find_state_block_idx(blocks) is None:
        return JSONResponse({"error": "存档格式不符，请确认是富甲天下5的 .sav 文件"}, status_code=400)

    try:
        fields = _sav_read_fields(blocks)
    except Exception as e:
        return JSONResponse({"error": f"读取资源失败: {e}"}, status_code=400)

    pet = None
    try:
        pet = _sav_read_current_pet(blocks)
    except Exception:
        pet = None

    lords = []
    try:
        lords = _sav_read_lords(blocks)
    except Exception:
        lords = []

    heroes = []
    try:
        heroes = _sav_read_heroes(blocks)
    except Exception:
        pass

    tmp_path = UPLOAD_DIR / f"gamesav_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.sav"
    tmp_path.write_bytes(raw)

    return JSONResponse({
        "ok": True,
        "fields": fields,
        "pet": pet,
        "lords": lords,
        "heroes": heroes,
        "tmp": str(tmp_path),
        "source_path": source_path_norm,
        "source_path_status": source_path_status,
        "source_path_message": source_path_message,
        "source_path_candidates": source_path_candidates or [],
        "original_filename": original_filename,
        "original_size": len(raw),
        "original_sha256": original_sha256,
    })

@app.get("/game-save", response_class=HTMLResponse)
def game_save_page(request: Request):
    return templates.TemplateResponse("game_save.html", {"request": request})


@app.get("/api/game-save/files")
def game_save_files(path: str = ""):
    target = Path(path).expanduser() if path else _sav_default_pick_dir()
    if target.exists() and target.is_file():
        target = target.parent
    if not target.exists() or not target.is_dir():
        return JSONResponse({"error": "目录不存在或无法访问"}, status_code=400)

    dirs, files = [], []
    try:
        entries = sorted(target.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
    except Exception as e:
        return JSONResponse({"error": f"读取目录失败: {e}"}, status_code=400)

    for entry in entries:
        if entry.name.startswith("."):
            continue
        try:
            if entry.is_dir():
                dirs.append({"name": entry.name, "path": str(entry)})
            elif entry.is_file() and entry.suffix.lower() == ".sav":
                stat = entry.stat()
                files.append({
                    "name": entry.name,
                    "path": str(entry),
                    "size": stat.st_size,
                    "mtime": int(stat.st_mtime),
                })
        except Exception:
            continue

    parent = ""
    if target.parent != target:
        parent = str(target.parent)

    return JSONResponse({
        "ok": True,
        "path": str(target),
        "parent": parent,
        "roots": _sav_picker_roots(),
        "dirs": dirs,
        "files": files,
    })


class GameSavePathParse(BaseModel):
    path: str

@app.post("/api/game-save/parse-path")
def game_save_parse_path(body: GameSavePathParse):
    path = Path(body.path).expanduser()
    if not path.is_absolute() or not path.exists() or not path.is_file():
        return JSONResponse({"error": "存档文件不存在，请重新选择"}, status_code=400)
    if path.suffix.lower() != ".sav":
        return JSONResponse({"error": "请选择 .sav 存档文件"}, status_code=400)
    try:
        raw = path.read_bytes()
    except Exception as e:
        return JSONResponse({"error": f"读取存档失败: {e}"}, status_code=400)
    return _sav_parse_payload(
        raw,
        _sav_original_filename(path.name),
        str(path),
        "selected",
        "已从本机目录选择存档，原始路径准确",
        [],
    )


@app.post("/api/game-save/parse")
async def game_save_parse(file: UploadFile = File(...), source_path: str = Form("")):
    raw = await file.read()
    original_filename = _sav_original_filename(file.filename)
    original_sha256 = _sav_sha256_bytes(raw)
    source_path_norm, source_path_status, source_path_message, source_path_candidates = _sav_resolve_parse_source_path(
        source_path,
        original_filename,
        raw,
        original_sha256,
    )
    return _sav_parse_payload(
        raw,
        original_filename,
        source_path_norm,
        source_path_status,
        source_path_message,
        source_path_candidates,
    )


class GameSaveHeroTemplateSave(BaseModel):
    id: Optional[str] = ""
    name: str
    heroes: dict

@app.get("/api/game-save/hero-templates")
def game_save_hero_templates():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT id, name, data_json, hero_count, field_count, created_at, updated_at
            FROM game_save_hero_templates
            ORDER BY updated_at DESC, created_at DESC
        """).fetchall()
        return {"ok": True, "templates": [_sav_template_from_row(row) for row in rows]}
    finally:
        conn.close()

@app.post("/api/game-save/hero-templates")
def game_save_save_hero_template(body: GameSaveHeroTemplateSave):
    name = _sav_clean_template_text(body.name, 60)
    if not name:
        return JSONResponse({"error": "请填写模板名称"}, status_code=400)

    heroes = _sav_sanitize_hero_template_heroes(body.heroes)
    if not heroes:
        return JSONResponse({"error": "当前没有可保存的武将属性修改"}, status_code=400)

    hero_count, field_count = _sav_hero_template_counts(heroes)
    now = _sav_template_now()
    template_id = _sav_clean_template_text(body.id, 64)
    data_json = json.dumps({"heroes": heroes}, ensure_ascii=False, separators=(",", ":"))

    conn = get_db()
    try:
        existing = None
        if template_id:
            existing = conn.execute("""
                SELECT id, created_at FROM game_save_hero_templates WHERE id=?
            """, (template_id,)).fetchone()
        if not existing:
            template_id = uuid4().hex
            conn.execute("""
                INSERT INTO game_save_hero_templates
                    (id, name, data_json, hero_count, field_count, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (template_id, name, data_json, hero_count, field_count, now, now))
        else:
            conn.execute("""
                UPDATE game_save_hero_templates
                SET name=?, data_json=?, hero_count=?, field_count=?, updated_at=?
                WHERE id=?
            """, (name, data_json, hero_count, field_count, now, template_id))
        conn.commit()
        row = conn.execute("""
            SELECT id, name, data_json, hero_count, field_count, created_at, updated_at
            FROM game_save_hero_templates
            WHERE id=?
        """, (template_id,)).fetchone()
        return {"ok": True, "template": _sav_template_from_row(row)}
    finally:
        conn.close()

@app.delete("/api/game-save/hero-templates/{template_id}")
def game_save_delete_hero_template(template_id: str):
    template_id = _sav_clean_template_text(template_id, 64)
    conn = get_db()
    try:
        cur = conn.execute("DELETE FROM game_save_hero_templates WHERE id=?", (template_id,))
        conn.commit()
        if cur.rowcount <= 0:
            return JSONResponse({"error": "模板不存在或已删除"}, status_code=404)
        return {"ok": True}
    finally:
        conn.close()


class GameSaveSave(BaseModel):
    tmp: str
    source_path: Optional[str] = ""
    original_filename: Optional[str] = ""
    original_size: Optional[int] = None
    original_sha256: Optional[str] = ""
    fields: dict                  # {"金额": 99999, ...}
    pet: Optional[dict] = None    # {"upgrade_points": 9}
    lords: Optional[dict] = None  # {"0": {"fields": {...}, "pet": {...}}, ...}
    heroes: Optional[dict] = None # {"70": {"武力": 62, ...}, ...}  稀疏，只含改动武将

@app.post("/api/game-save/save")
def game_save_save(body: GameSaveSave):
    tmp_path = Path(body.tmp)
    if not tmp_path.exists() or not str(tmp_path).startswith(str(UPLOAD_DIR)):
        return JSONResponse({"error": "临时文件不存在或路径非法"}, status_code=400)
    try:
        source_path_norm = _sav_normalize_source_path(
            body.source_path,
            body.original_filename,
            body.original_size,
            body.original_sha256,
        )
    except Exception as e:
        return JSONResponse({"error": f"原始存档路径校验失败: {e}"}, status_code=400)

    raw = tmp_path.read_bytes()
    try:
        blocks, tail = _sav_parse_blocks(raw)
    except Exception as e:
        return JSONResponse({"error": f"解析失败: {e}"}, status_code=400)

    if body.lords:
        try:
            _sav_write_lords(blocks, body.lords)
        except Exception as e:
            return JSONResponse({"error": f"写入主公数据失败: {e}"}, status_code=400)
    else:
        # 兼容旧前端：只写当前主公资源和当前宠物升级点数
        try:
            _sav_write_fields(blocks, body.fields)
        except Exception as e:
            return JSONResponse({"error": f"写入资源失败: {e}"}, status_code=400)

        if body.pet:
            try:
                _sav_write_current_pet(blocks, body.pet)
            except Exception as e:
                return JSONResponse({"error": f"写入宠物数据失败: {e}"}, status_code=400)

    # 写入武将维度
    if body.heroes:
        try:
            _sav_write_heroes(blocks, body.heroes)
        except Exception as e:
            return JSONResponse({"error": f"写入武将数据失败: {e}"}, status_code=400)

    new_raw = _sav_build(blocks, tail)
    try:
        output_path = _sav_output_path(tmp_path, source_path_norm, body.original_filename)
    except Exception as e:
        return JSONResponse({"error": f"生成新存档路径失败: {e}"}, status_code=400)

    # 覆写临时文件，并在原存档同级 new 目录生成新存档。
    tmp_path.write_bytes(new_raw)
    output_path.write_bytes(new_raw)
    download_url = _sav_download_url_for(output_path)

    return JSONResponse({
        "ok": True,
        "download": download_url,
        "output": str(output_path),
        "filename": output_path.name,
        "backup": "",
        "size": len(new_raw),
    })


# ── 三国立志传3 存档修改器 ──────────────────────────────────
#
# 存档格式：原始 C++ 堆内存转储，Big5 繁体中文
# 基地址：0xCA520000，哨兵：0xCA521FAB（ab 1f 52 ca）
#
# 已验证字段（v2/v3/v4/v5 四存档交叉验证）：
#   场景天数  save0:0x1578      4字节LE
#   地图 X    save0:0x4D4B0     4字节LE
#   地图 Y    save0:0x4D4B4     4字节LE
#   关羽/张飞 等级  save0:0x42490   4字节LE
#   皇甫嵩 等级     save0:0x43990   4字节LE
#
# 点数编码（指针低字节，主+副本均须更新）：
#   A组(刘备/皇甫嵩) C=8: 主@0xB8, 副@0x17BD8
#     未花费 = 8 - (ptr_low - 0xAB)
#     设N点: ptr_low = 0xB3 - N
#   B组(关羽/张飞)   C=0: 主@0x15C8, 副@0x2AD8
#     未花费 = 0xAB - ptr_low
#     设N点: ptr_low = 0xAB - N
#
# 指针结构：4字节LE = [低字节, 0x1F, 0x52, 0xCA]，只改低字节

_S3_MAGIC    = b'\xb6\xc0\xa4\x79'  # Big5「黃巾」前两字
_S3_PTR_BASE = 0xAB                  # 0xCA521FAB 低字节 = 无花费基准
_S3_EXTRA_HEROES_PATH = Path(__file__).parent / "sango3_extra_heroes.json"

# 武将信息表：黃巾之亂场景已验证的 4 位（v2/v3/v4/v5 四存档交叉确认）
# ptr_C：点数公式常数；ptr_off：指针低字节文件偏移
# 点数 = ptr_C - (raw[ptr_off] - 0xAB)；设 N 点：raw[ptr_off] = (ptr_C - N + 0xAB) & 0xFF
_S3_HEROES = [
    {
        "id": "liu",
        "name": "刘备",
        "title": "汉室宗亲",
        "faction": "桃园义勇",
        "faction_tag": "蜀",
        "ptr_off": 0x000B8,
        "ptr_C": 8,
        "level_off": None,          # 等级存储未完全定位，暂不支持
        "desc": "仁义之君，三兄弟之首",
    },
    {
        "id": "guan",
        "name": "关羽",
        "title": "武圣",
        "faction": "桃园义勇",
        "faction_tag": "蜀",
        "ptr_off": 0x15C8,
        "ptr_C": 0,
        "level_off": 0x42490,
        "desc": "义绝，千里走单骑",
    },
    {
        "id": "zhang",
        "name": "张飞",
        "title": "万人敌",
        "faction": "桃园义勇",
        "faction_tag": "蜀",
        "ptr_off": 0x2AD8,
        "ptr_C": 0,
        "level_off": 0x42490,       # 与关羽共用同一字段（二人同级时）
        "desc": "豪勇无双，声震长坂",
    },
    {
        "id": "huangfu",
        "name": "皇甫嵩",
        "title": "汉廷名将",
        "faction": "汉廷官军",
        "faction_tag": "汉",
        "ptr_off": 0x17BD8,
        "ptr_C": 8,
        "level_off": 0x43990,
        "desc": "平定黄巾第一功臣",
    },
]

# 直接字段：(name, offset, min, max)
_S3_DIRECT_FIELDS = [
    ("scene_days", 0x01578, 0, 9999),
    ("map_x",      0x4D4B0, 0, 9999),
    ("map_y",      0x4D4B4, 0, 9999),
]

def _s3_validate(raw: bytes) -> bool:
    if len(raw) < 0x4D4C0:
        return False
    return raw[0x10:0x14] == _S3_MAGIC

def _s3_read_fields(raw: bytes) -> dict:
    fields = {}
    for name, off, lo, hi in _S3_DIRECT_FIELDS:
        fields[name] = struct.unpack_from('<I', raw, off)[0]
    return fields

def _s3_load_extra_heroes() -> list[dict]:
    """从本地 JSON 读取用户自行发现的额外武将槽位。"""
    import json as _json
    if _S3_EXTRA_HEROES_PATH.exists():
        try:
            return _json.loads(_S3_EXTRA_HEROES_PATH.read_text("utf-8"))
        except Exception:
            pass
    return []

def _s3_save_extra_heroes(heroes: list[dict]) -> None:
    import json as _json
    _S3_EXTRA_HEROES_PATH.write_text(
        _json.dumps(heroes, ensure_ascii=False, indent=2), "utf-8"
    )

def _s3_all_heroes() -> list[dict]:
    """内置 4 位 + 用户发现的额外武将，去重。"""
    extra = _s3_load_extra_heroes()
    known_offs = {h["ptr_off"] for h in _S3_HEROES}
    merged = list(_S3_HEROES)
    for eh in extra:
        if eh.get("ptr_off") not in known_offs:
            merged.append(eh)
            known_offs.add(eh["ptr_off"])
    return merged

def _s3_read_heroes(raw: bytes) -> list[dict]:
    """返回所有武将（内置+用户发现）的当前点数，过滤掉不在队中的。"""
    result = []
    for h in _s3_all_heroes():
        ptr_off = h["ptr_off"]
        if ptr_off >= len(raw):
            continue
        ptr_low = raw[ptr_off]
        # 哨兵值 = 0xAB → 该槽位未激活（武将不在队中）
        if ptr_low == _S3_PTR_BASE:
            continue
        points = max(0, h["ptr_C"] - (ptr_low - _S3_PTR_BASE))
        level  = None
        if h.get("level_off") and h["level_off"] < len(raw):
            lv = struct.unpack_from('<I', raw, h["level_off"])[0]
            if 1 <= lv <= 50:
                level = lv
        result.append({
            "id":          h["id"],
            "name":        h["name"],
            "title":       h.get("title", ""),
            "faction":     h.get("faction", "未知"),
            "faction_tag": h.get("faction_tag", "?"),
            "desc":        h.get("desc", ""),
            "points":      points,
            "level":       level,
            "discovered":  h.get("discovered", False),   # 是否用户发现
        })
    return result

def _s3_discover_heroes(raw_before: bytes, raw_after: bytes,
                         points_after: int) -> list[dict]:
    """
    差分分析：比较两份存档，找出新武将的指针槽位。

    核心规则：
    1. 两份存档在该偏移的值都是 0xCA521Fxx 格式的合法指针
    2. 值发生了变化（武将状态改变）
    3. 「后」存档的值不是哨兵（0xAB），且 C 常数合理
    4. 偏移在已知武将指针分布范围内（过滤掉大量游戏世界数据）
    5. 不在已注册武将的偏移列表中（排除重复）
    """
    SENTINEL    = 0xCA521FAB
    PTR_HI_MASK = 0xFFFFFF00
    PTR_HI_WANT = 0xCA521F00   # 合法武将指针高24位

    # 已知武将偏移（排除）
    known_offs = {h["ptr_off"] for h in _s3_all_heroes()}

    # 可能包含武将指针的偏移范围（基于 4 个已知槽位的观察）
    # 覆盖头部区域 + 已知扩展区域附近
    SEARCH_RANGES = [
        (0x000070, 0x000200),    # 主头部指针数组
        (0x000200, 0x000600),    # 扩展头部
        (0x001000, 0x005000),    # 关羽/皇甫嵩 槽位附近
        (0x010000, 0x030000),    # 张飞槽位附近（0x2AD8）
    ]

    results = []

    for rng_start, rng_end in SEARCH_RANGES:
        end = min(rng_end, min(len(raw_before), len(raw_after)) - 3)
        for off in range(rng_start, end, 4):
            if off in known_offs:
                continue

            val_b = struct.unpack_from('<I', raw_before, off)[0]
            val_a = struct.unpack_from('<I', raw_after,  off)[0]

            # 两个值都必须是合法的 CA521Fxx 指针
            if (val_b & PTR_HI_MASK) != PTR_HI_WANT:
                continue
            if (val_a & PTR_HI_MASK) != PTR_HI_WANT:
                continue

            # 值必须发生变化
            if val_b == val_a:
                continue

            # 「后」存档的指针低字节必须不是哨兵
            ptr_low_a = val_a & 0xFF
            if ptr_low_a == _S3_PTR_BASE:
                continue

            # 推算 C 常数
            C_calc = points_after + ptr_low_a - _S3_PTR_BASE
            if not (0 <= C_calc <= 50):
                continue

            # 「前」存档的值（可能是哨兵，也可能是同一武将的不同状态）
            ptr_low_b = val_b & 0xFF

            results.append({
                "ptr_off":   off,
                "ptr_low":   ptr_low_a,
                "ptr_C":     C_calc,
                "points":    points_after,
                "ptr_before": ptr_low_b,
            })

    return results

def _s3_write_fields(raw: bytearray, patch: dict) -> None:
    direct_map = {name: (off, lo, hi) for name, off, lo, hi in _S3_DIRECT_FIELDS}
    hero_map   = {h["id"]: h for h in _s3_all_heroes()}
    for name, val in patch.items():
        if name in direct_map:
            off, lo, hi = direct_map[name]
            struct.pack_into('<I', raw, off, max(lo, min(hi, int(val))))
        elif name.endswith("_points") and name[:-7] in hero_map:
            h = hero_map[name[:-7]]
            n = max(0, min(200, int(val)))
            raw[h["ptr_off"]] = (h["ptr_C"] - n + _S3_PTR_BASE) & 0xFF


@app.get("/sango3-save", response_class=HTMLResponse)
def sango3_page(request: Request):
    return templates.TemplateResponse("sango3_save.html", {"request": request})


@app.post("/api/sango3-save/discover")
async def sango3_discover(
    before: UploadFile = File(...),
    after:  UploadFile = File(...),
    points: int = Form(0),
):
    """差分分析：上传加入前/后的 save0.dat，自动发现新武将槽位。"""
    raw_b = await before.read()
    raw_a = await after.read()
    if not (_s3_validate(raw_b) and _s3_validate(raw_a)):
        return JSONResponse({"error": "文件格式不符"}, status_code=400)
    candidates = _s3_discover_heroes(raw_b, raw_a, max(0, points))
    return JSONResponse({"ok": True, "candidates": candidates})


class Sango3Hero(BaseModel):
    id:          str
    name:        str
    title:       str = ""
    faction:     str = "未知"
    faction_tag: str = "?"
    desc:        str = ""
    ptr_off:     int
    ptr_C:       int
    level_off:   Optional[int] = None
    discovered:  bool = True

@app.post("/api/sango3-save/add-hero")
def sango3_add_hero(hero: Sango3Hero):
    """保存用户发现的新武将槽位到本地 JSON。"""
    extra = _s3_load_extra_heroes()
    # 去重（按 ptr_off）
    extra = [h for h in extra if h.get("ptr_off") != hero.ptr_off]
    extra.append(hero.model_dump())
    _s3_save_extra_heroes(extra)
    return JSONResponse({"ok": True, "total": len(extra)})


@app.get("/api/sango3-save/heroes-registry")
def sango3_heroes_registry():
    """返回当前已注册的所有武将（内置 + 用户发现）。"""
    return JSONResponse({
        "builtin": [h["id"] for h in _S3_HEROES],
        "extra":   _s3_load_extra_heroes(),
    })


@app.post("/api/sango3-save/parse")
async def sango3_parse(file: UploadFile = File(...)):
    raw = await file.read()
    if not _s3_validate(raw):
        return JSONResponse({"error": "文件格式不符，请上传 save0.dat（三国立志传3）"}, status_code=400)
    fields = _s3_read_fields(raw)
    heroes = _s3_read_heroes(raw)
    tmp = UPLOAD_DIR / f"sango3_{datetime.now().strftime('%Y%m%d%H%M%S')}.dat"
    tmp.write_bytes(raw)
    return JSONResponse({"ok": True, "fields": fields, "heroes": heroes, "tmp": str(tmp)})


class Sango3Save(BaseModel):
    tmp: str
    fields: dict

@app.post("/api/sango3-save/save")
def sango3_save(body: Sango3Save):
    tmp_path = Path(body.tmp)
    if not tmp_path.exists() or not str(tmp_path).startswith(str(UPLOAD_DIR)):
        return JSONResponse({"error": "临时文件不存在"}, status_code=400)
    raw = bytearray(tmp_path.read_bytes())
    if not _s3_validate(raw):
        return JSONResponse({"error": "文件已损坏"}, status_code=400)
    _s3_write_fields(raw, body.fields)
    bak = tmp_path.with_suffix(f".bak_{datetime.now().strftime('%Y%m%d%H%M%S')}")
    import shutil as _sh; _sh.copy2(tmp_path, bak)
    tmp_path.write_bytes(raw)
    return JSONResponse({"ok": True, "download": f"/uploads/{tmp_path.name}", "size": len(raw)})


def _argv_value(flag: str) -> str | None:
    """读取当前 uvicorn 启动参数中的简单 flag 值。"""
    try:
        idx = sys.argv.index(flag)
        return sys.argv[idx + 1]
    except (ValueError, IndexError):
        return None


def _restart_bind_host(request: Request) -> str:
    """推断重启后的监听地址，优先复用当前 uvicorn --host。"""
    host = _argv_value("--host") or request.url.hostname or "127.0.0.1"
    if host == "localhost":
        host = "127.0.0.1"
    if host not in {"127.0.0.1", "0.0.0.0", "::"} and re.match(r"^\d{1,3}(\.\d{1,3}){3}$", host):
        return "0.0.0.0"
    return host


def _restart_port(request: Request) -> int:
    raw = _argv_value("--port")
    if raw and raw.isdigit():
        return int(raw)
    return request.url.port or (443 if request.url.scheme == "https" else 8080)


def _restart_command(request: Request) -> list[str]:
    """构造干净的单进程 uvicorn 启动命令，避免 reload/supervisor 多进程残留。"""
    host = _restart_bind_host(request)
    port = _restart_port(request)
    return [sys.executable, "-m", "uvicorn", "main:app", "--host", host, "--port", str(port)]


def _spawn_restart_helper(request: Request) -> dict:
    """
    启动 detached helper 完成真正重启。

    之前的实现只重启当前处理请求的进程；当同一端口存在 reload supervisor、
    worker 或历史残留进程时，页面会表现为“重启了但没变化”。helper 会在响应
    返回后按端口清理本项目监听进程，再启动一份干净服务。
    """
    port = _restart_port(request)
    command = _restart_command(request)
    log_path = BASE_DIR / ".restart.log"
    helper_code = r'''
import json, os, signal, subprocess, sys, time

cfg = json.loads(sys.argv[1])
port = int(cfg["port"])
cwd = cfg["cwd"]
command = cfg["command"]
log_path = cfg["log_path"]
helper_pid = os.getpid()

def write_log(message):
    stamp = time.strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{stamp}] {message}\n")

def run_text(args):
    try:
        return subprocess.check_output(args, text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return ""

def listener_pids():
    out = run_text(["lsof", "-tiTCP:%s" % port, "-sTCP:LISTEN"])
    return {int(line) for line in out.splitlines() if line.strip().isdigit()}

def command_line(pid):
    return run_text(["ps", "-p", str(pid), "-o", "command="])

def is_target(pid):
    if pid == helper_pid:
        return False
    cmd = command_line(pid)
    return bool(cmd) and (cwd in cmd or "main:app" in cmd)

def pid_alive(pid):
    try:
        os.kill(pid, 0)
        return True
    except ProcessLookupError:
        return False
    except PermissionError:
        return True

time.sleep(1.0)
targets = {pid for pid in listener_pids() if is_target(pid)}
write_log("restart requested, targets=%s, command=%s" % (sorted(targets), command))
for pid in sorted(targets):
    try:
        os.kill(pid, signal.SIGTERM)
    except ProcessLookupError:
        pass
    except Exception as exc:
        write_log("SIGTERM failed pid=%s error=%s" % (pid, exc))

deadline = time.time() + 4.0
while time.time() < deadline:
    remaining = {pid for pid in targets if pid_alive(pid)}
    fresh = {pid for pid in listener_pids() if is_target(pid) and pid not in targets}
    if fresh:
        write_log("fresh listener detected after SIGTERM, skip duplicate start: %s" % sorted(fresh))
        sys.exit(0)
    if not remaining:
        break
    time.sleep(0.2)

remaining = {pid for pid in listener_pids() if pid in targets and is_target(pid)}
for pid in sorted(remaining):
    try:
        os.kill(pid, signal.SIGKILL)
        write_log("SIGKILL pid=%s" % pid)
    except ProcessLookupError:
        pass
    except Exception as exc:
        write_log("SIGKILL failed pid=%s error=%s" % (pid, exc))

deadline = time.time() + 4.0
while time.time() < deadline and any(pid in targets and is_target(pid) for pid in listener_pids()):
    time.sleep(0.2)

fresh = {pid for pid in listener_pids() if is_target(pid) and pid not in targets}
if fresh:
    write_log("fresh listener detected before start, skip duplicate start: %s" % sorted(fresh))
    sys.exit(0)

# Some local launchers immediately bring the service back after SIGTERM. Give them
# a short stable window so the helper does not race them for the same port.
deadline = time.time() + 2.0
while time.time() < deadline:
    fresh = {pid for pid in listener_pids() if is_target(pid) and pid not in targets}
    if fresh:
        write_log("fresh listener detected during grace window, skip duplicate start: %s" % sorted(fresh))
        sys.exit(0)
    time.sleep(0.2)

with open(log_path, "ab") as log:
    subprocess.Popen(
        command,
        cwd=cwd,
        stdin=subprocess.DEVNULL,
        stdout=log,
        stderr=subprocess.STDOUT,
        start_new_session=True,
        close_fds=True,
    )
write_log("started new process")
'''
    payload = {
        "port": port,
        "cwd": str(BASE_DIR),
        "command": command,
        "log_path": str(log_path),
    }
    subprocess.Popen(
        [sys.executable, "-c", helper_code, json.dumps(payload)],
        cwd=str(BASE_DIR),
        stdin=subprocess.DEVNULL,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
        close_fds=True,
    )
    return {"port": port, "command": command}


# ── ComfyUI 管理 ──────────────────────────────────────────
_COMFY_DIR  = Path("/Users/chengyu/project/ComfyUI")
_COMFY_PORT = 8188
_COMFY_URL  = f"http://localhost:{_COMFY_PORT}"
_COMFY_FALLBACK_URLS = (f"http://127.0.0.1:{_COMFY_PORT}", f"http://localhost:{_COMFY_PORT}")
_COMFY_LOG_PATH = BASE_DIR / ".comfyui.log"
_comfy_proc: subprocess.Popen | None = None   # 由本进程启动的子进程句柄


def _comfy_pid_running(pid: int) -> bool:
    try:
        os.kill(pid, 0)
        return True
    except ProcessLookupError:
        return False
    except PermissionError:
        return True


def _comfy_find_pid() -> int | None:
    """通过监听端口 8188 找 ComfyUI 进程 PID。"""
    try:
        # 只返回监听 8188 端口的 PID，不依赖进程名。
        out = subprocess.check_output(
            ["lsof", "-nP", f"-iTCP:{_COMFY_PORT}", "-sTCP:LISTEN", "-t"],
            text=True, stderr=subprocess.DEVNULL,
        ).strip()
        pids = [int(p) for p in out.splitlines() if p.strip().isdigit()]
        return pids[0] if pids else None
    except Exception:
        return None


async def _comfy_alive() -> bool:
    """HTTP 探活 ComfyUI。禁用代理，并同时尝试 127.0.0.1 / localhost。"""
    import urllib.request

    def probe(url: str) -> bool:
        try:
            opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
            req = urllib.request.Request(url, headers={"Cache-Control": "no-cache"})
            with opener.open(req, timeout=2) as resp:
                return resp.status < 500
        except Exception:
            return False

    for url in _COMFY_FALLBACK_URLS:
        if await asyncio.to_thread(probe, url):
            return True
    return False

def _comfy_tail_log(limit: int = 20) -> str:
    if not _COMFY_LOG_PATH.exists():
        return ""
    try:
        return "\n".join(_COMFY_LOG_PATH.read_text(errors="replace").splitlines()[-limit:])
    except Exception:
        return ""


@app.get("/comfyui", response_class=HTMLResponse)
def comfyui_page(request: Request):
    return templates.TemplateResponse("comfyui.html", {"request": request})


@app.get("/api/comfyui/status")
async def comfyui_status():
    global _comfy_proc
    pid = None
    if _comfy_proc and _comfy_proc.poll() is None:
        pid = _comfy_proc.pid
    else:
        pid = _comfy_find_pid()
    http_alive = await _comfy_alive()
    alive = http_alive or bool(pid)
    models, downloading = [], []
    ckpt_dir = _COMFY_DIR / "models" / "checkpoints"
    if ckpt_dir.exists():
        for f in sorted(ckpt_dir.iterdir()):
            sz = f.stat().st_size
            if f.suffix in (".safetensors", ".ckpt", ".pt") and sz > 1_000_000:
                models.append({"name": f.name, "size_mb": round(sz / 1024 / 1024)})
            elif f.name.endswith(".part") and sz > 0:
                total = 6952523776   # Animagine / BP XL ≈ 6.5 GB
                downloading.append({
                    "name": f.name.removesuffix(".part"),
                    "done_mb": round(sz / 1024 / 1024),
                    "pct": min(99, round(sz * 100 / total)),
                })
    return JSONResponse({
        "pid":         pid,
        "alive":       alive,
        "http_alive":  http_alive,
        "url":         _COMFY_URL,
        "models":      models,
        "downloading": downloading,
    })


@app.post("/api/comfyui/start")
async def comfyui_start():
    global _comfy_proc
    existing_pid = _comfy_find_pid()
    existing_alive = await _comfy_alive()
    if existing_pid:
        return JSONResponse({
            "ok": True,
            "pid": existing_pid,
            "alive": True,
            "http_alive": existing_alive,
            "url": _COMFY_URL,
            "msg": "ComfyUI 已在运行",
        })
    if existing_alive:
        return JSONResponse({"ok": True, "alive": True, "http_alive": True, "url": _COMFY_URL, "msg": "已在运行"})
    python = _COMFY_DIR / ".venv" / "bin" / "python"
    if not python.exists():
        return JSONResponse({"error": "ComfyUI venv 未找到"}, status_code=400)
    _COMFY_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    log_file = _COMFY_LOG_PATH.open("a", encoding="utf-8")
    log_file.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] starting ComfyUI\n")
    log_file.flush()
    _comfy_proc = subprocess.Popen(
        [str(python), "main.py",
         "--listen", "127.0.0.1",
         "--port", str(_COMFY_PORT),
         "--force-fp16"],
        cwd=str(_COMFY_DIR),
        stdout=log_file,
        stderr=subprocess.STDOUT,
        start_new_session=True,
    )
    # 等待进程出现（最多 5 秒），然后立即返回
    # 注意：ComfyUI 冷启动需要 30-60 秒，前端会轮询状态
    for _ in range(5):
        await asyncio.sleep(1)
        if _comfy_proc.poll() is not None:
            return JSONResponse({
                "error": "进程立即退出",
                "log": _comfy_tail_log(),
            }, status_code=500)
    pid = _comfy_proc.pid
    # 尝试快速探活（最多再等 10 秒）
    for _ in range(10):
        await asyncio.sleep(1)
        if await _comfy_alive():
            return JSONResponse({"ok": True, "pid": pid, "alive": True})
    return JSONResponse({"ok": True, "pid": pid, "alive": False, "msg": "正在启动中，模型加载需要约 30-60 秒，页面会自动刷新"})


@app.post("/api/comfyui/stop")
async def comfyui_stop():
    global _comfy_proc
    pid = None
    if _comfy_proc and _comfy_proc.poll() is None:
        pid = _comfy_proc.pid
        _comfy_proc.terminate()
        try:
            _comfy_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            _comfy_proc.kill()
        _comfy_proc = None
    else:
        pid = _comfy_find_pid()
        if pid:
            os.kill(pid, 15)   # SIGTERM
            await asyncio.sleep(2)
            if _comfy_pid_running(pid):
                os.kill(pid, 9)
    return JSONResponse({"ok": True, "stopped_pid": pid})


@app.post("/api/restart")
async def restart_server(request: Request):
    """
    重启当前服务。响应返回后由 detached helper 清理当前端口的本项目监听进程，
    再启动一份干净的 uvicorn 服务，避免多进程/残留监听导致“重启没有用”。
    """
    info = _spawn_restart_helper(request)
    return JSONResponse({"ok": True, "mode": "detached", **info})


# ── 文件整理 ──────────────────────────────────────────────

def _organizer_rules() -> list[dict]:
    conn = get_db()
    rows = conn.execute(
        "SELECT id,category,match_type,pattern,priority,enabled FROM organizer_rules ORDER BY priority, id"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


class OrganizerDirIn(BaseModel):
    path: str


class OrganizerRuleIn(BaseModel):
    category: str
    match_type: str = "keyword"
    pattern: str
    priority: int = 500
    enabled: int = 1


class OrganizerRunIn(BaseModel):
    path: str
    include_hidden: bool = False


@app.get("/api/file-organizer/dirs")
def organizer_list_dirs():
    """列出纳管目录及各自的散落文件数。目录被用户手工删掉时不报错，标 missing 让前端提示。"""
    conn = get_db()
    rows = conn.execute("SELECT id,path FROM organizer_dirs ORDER BY id").fetchall()
    conn.close()
    items = []
    for row in rows:
        try:
            directory = organizer_safe_dir(row["path"])
            items.append({"id": row["id"], **organizer_dir_summary(directory), "missing": False})
        except OrganizerError as exc:
            items.append({"id": row["id"], "path": row["path"], "name": Path(row["path"]).name,
                          "loose_count": 0, "subdir_count": 0, "missing": True, "error": str(exc)})
    return JSONResponse({"dirs": items})


@app.post("/api/file-organizer/dirs")
def organizer_add_dir(body: OrganizerDirIn):
    try:
        directory = organizer_safe_dir(body.path)
    except OrganizerError as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO organizer_dirs(path,created_at) VALUES(?,?)",
                 (str(directory), datetime.now().isoformat(timespec="seconds")))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True, "path": str(directory)})


@app.delete("/api/file-organizer/dirs/{dir_id}")
def organizer_delete_dir(dir_id: int):
    """只从纳管列表移除，不碰磁盘上的目录。"""
    conn = get_db()
    conn.execute("DELETE FROM organizer_dirs WHERE id=?", (dir_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})


@app.get("/api/file-organizer/rules")
def organizer_get_rules():
    return JSONResponse({"rules": _organizer_rules()})


@app.post("/api/file-organizer/rules")
def organizer_create_rule(body: OrganizerRuleIn):
    if not body.category.strip() or not body.pattern.strip():
        return JSONResponse({"ok": False, "error": "分类名和匹配内容不能为空"}, status_code=400)
    if "/" in body.category:
        return JSONResponse({"ok": False, "error": "分类名不能包含斜杠"}, status_code=400)
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO organizer_rules(category,match_type,pattern,priority,enabled,updated_at)"
        " VALUES(?,?,?,?,?,?)",
        (body.category.strip(), body.match_type, body.pattern.strip(), body.priority,
         body.enabled, datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    rule_id = cur.lastrowid
    conn.close()
    return JSONResponse({"ok": True, "id": rule_id})


@app.patch("/api/file-organizer/rules/{rule_id}")
def organizer_update_rule(rule_id: int, body: OrganizerRuleIn):
    conn = get_db()
    conn.execute(
        "UPDATE organizer_rules SET category=?,match_type=?,pattern=?,priority=?,enabled=?,updated_at=?"
        " WHERE id=?",
        (body.category.strip(), body.match_type, body.pattern.strip(), body.priority,
         body.enabled, datetime.now().isoformat(timespec="seconds"), rule_id),
    )
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})


@app.delete("/api/file-organizer/rules/{rule_id}")
def organizer_delete_rule(rule_id: int):
    conn = get_db()
    conn.execute("DELETE FROM organizer_rules WHERE id=?", (rule_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})


@app.post("/api/file-organizer/preview")
def organizer_preview(body: OrganizerRunIn):
    """不动磁盘，只回一份「哪个文件去哪」的清单，供页面点整理前心里有数。"""
    try:
        directory = organizer_safe_dir(body.path)
    except OrganizerError as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)
    planned = organizer_plan(directory, _organizer_rules(), body.include_hidden)
    summary: dict[str, int] = {}
    for item in planned:
        summary[item.category] = summary.get(item.category, 0) + 1
    return JSONResponse({
        "ok": True,
        "total": len(planned),
        "summary": summary,
        "items": [{"filename": p.filename, "category": p.category, "reason": p.reason} for p in planned],
    })


@app.post("/api/file-organizer/organize")
def organizer_run(body: OrganizerRunIn):
    """一键整理。移动流水落 organizer_runs，供撤销使用。"""
    try:
        directory = organizer_safe_dir(body.path)
    except OrganizerError as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)
    result = organizer_organize(directory, _organizer_rules(), body.include_hidden)
    run_id = None
    if result.moved:
        conn = get_db()
        cur = conn.execute(
            "INSERT INTO organizer_runs(directory,moves,moved_count,created_at) VALUES(?,?,?,?)",
            (str(directory), json.dumps(result.moved, ensure_ascii=False), len(result.moved),
             datetime.now().isoformat(timespec="seconds")),
        )
        conn.commit()
        run_id = cur.lastrowid
        conn.close()
    return JSONResponse({
        "ok": True,
        "run_id": run_id,
        "moved": len(result.moved),
        "skipped": result.skipped,
        "categories": result.categories,
        "details": result.moved,
    })


@app.get("/api/file-organizer/runs")
def organizer_list_runs():
    conn = get_db()
    rows = conn.execute(
        "SELECT id,directory,moved_count,created_at,undone_at FROM organizer_runs"
        " ORDER BY id DESC LIMIT 20"
    ).fetchall()
    conn.close()
    return JSONResponse({"runs": [dict(r) for r in rows]})


@app.post("/api/file-organizer/runs/{run_id}/undo")
def organizer_undo_run(run_id: int):
    conn = get_db()
    row = conn.execute("SELECT moves,undone_at FROM organizer_runs WHERE id=?", (run_id,)).fetchone()
    if not row:
        conn.close()
        return JSONResponse({"ok": False, "error": "找不到这次整理记录"}, status_code=404)
    if row["undone_at"]:
        conn.close()
        return JSONResponse({"ok": False, "error": "这次整理已经撤销过了"}, status_code=400)
    outcome = organizer_undo(json.loads(row["moves"]))
    conn.execute("UPDATE organizer_runs SET undone_at=? WHERE id=?",
                 (datetime.now().isoformat(timespec="seconds"), run_id))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True, **outcome})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8080, reload=True)
